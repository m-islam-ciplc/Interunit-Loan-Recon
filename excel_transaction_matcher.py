import pandas as pd
# import numpy as np  # ❌ UNUSED - commenting out
import re
from typing import List, Dict, Any, Tuple
# import logging  # ❌ UNUSED - commenting out
import os
import sys
import argparse
from openpyxl.styles import Alignment
import openpyxl
from lc_matching_logic import LCMatchingLogic
from po_matching_logic import POMatchingLogic
from usd_matching_logic import USDMatchingLogic
from interunit_loan_matching_logic import InterunitLoanMatcher
from transaction_block_identifier import TransactionBlockIdentifier

# =============================================================================
# CONFIGURATION SECTION
# =============================================================================
# Import configuration from dedicated config module
from config import (
    INPUT_FILE1_PATH, INPUT_FILE2_PATH, OUTPUT_FOLDER, OUTPUT_SUFFIX,
    SIMPLE_SUFFIX, CREATE_SIMPLE_FILES, 
    # CREATE_ALT_FILES,  # ❌ UNUSED - commenting out
    VERBOSE_DEBUG
    # AMOUNT_TOLERANCE  # ❌ UNUSED - removed since all matching uses exact amounts
)

# Import patterns from their respective modules
from lc_matching_logic import LC_PATTERN
from po_matching_logic import PO_PATTERN
from usd_matching_logic import USD_PATTERN

def print_configuration():
    """Print current configuration settings."""
    print("=" * 60)
    print("CURRENT CONFIGURATION")
    print("=" * 60)
    print(f"Input File 1: {INPUT_FILE1_PATH}")
    print(f"Input File 2: {INPUT_FILE2_PATH}")
    print(f"Output Folder: {OUTPUT_FOLDER}")
    print(f"Output Suffix: {OUTPUT_SUFFIX}")
    print(f"Simple Files: {'Yes' if CREATE_SIMPLE_FILES else 'No'}")
    # print(f"Alternative Files: {'Yes' if CREATE_ALT_FILES else 'No'}")  # ❌ UNUSED - commenting out
    print(f"Verbose Debug: {'Yes' if VERBOSE_DEBUG else 'No'}")
    print(f"LC Pattern: {LC_PATTERN}")
    print(f"PO Pattern: {PO_PATTERN}")
    print(f"USD Pattern: {USD_PATTERN}")
    # print(f"Amount Tolerance: {AMOUNT_TOLERANCE}")  # ❌ UNUSED - removed
    print("=" * 60)

def update_configuration():
    """Interactive configuration update (for future use)."""
    print("To update configuration, modify the variables at the top of this file:")
    print("1. INPUT_FILE1_PATH - Path to your first Excel file")
    print("2. INPUT_FILE2_PATH - Path to your second Excel file")
    print("3. OUTPUT_FOLDER - Where to save output files")
    print("4. OUTPUT_SUFFIX - Suffix for matched files")
    print("5. SIMPLE_SUFFIX - Suffix for simple test files")
    print("6. CREATE_SIMPLE_FILES - Whether to create simple test files")
    # print("7. CREATE_ALT_FILES - Whether to create alternative files")  # ❌ UNUSED - commenting out
    print("8. VERBOSE_DEBUG - Whether to show detailed debug output")
    print("9. LC_PATTERN - Regex pattern for LC number extraction (defined in lc_matching_logic.py)")
    print("10. PO_PATTERN - Regex pattern for PO number extraction (defined in po_matching_logic.py)")
    print("11. USD_PATTERN - Regex pattern for USD amount extraction (defined in usd_matching_logic.py)")
    # print("12. AMOUNT_TOLERANCE - Tolerance for amount matching (0 for exact)")  # ❌ UNUSED - removed

class   ExcelTransactionMatcher:
    """
    Handles complex Excel files with metadata rows and transaction data.
    """
    
    def __init__(self, file1_path: str, file2_path: str):
        self.file1_path = file1_path
        self.file2_path = file2_path
        self.metadata1 = None
        self.transactions1 = None
        self.metadata2 = None
        self.transactions2 = None
        self.lc_matching_logic = LCMatchingLogic()
        self.po_matching_logic = POMatchingLogic()
        self.usd_matching_logic = USDMatchingLogic()
        self.interunit_loan_matcher = InterunitLoanMatcher()
        self.block_identifier = TransactionBlockIdentifier()
        
        # ❌ UNUSED INSTANCE VARIABLES - commenting out
        # self.lc_parent_mapping = None
        # self.po_parent_mapping = None
        
    def read_complex_excel(self, file_path: str):
        """Read Excel file with metadata + transaction structure."""
        # Read everything first - preserve date format by reading as strings
        full_df = pd.read_excel(file_path, header=None, converters={0: str})

        # Extract metadata (rows 0-7, which are Excel rows 1-8)
        metadata = full_df.iloc[0:8, :]

        # Extract transaction data (rows 8+, which are Excel rows 9+)
        transactions = full_df.iloc[8:, :]

        # Set first row as headers and remove it from data
        transactions.columns = transactions.iloc[0]
        transactions = transactions.iloc[1:].reset_index(drop=True)

        # DEBUG: Show what columns we actually have
        print(f"DEBUG: Columns after transformation: {list(transactions.columns)}")

        # DEBUG: Show actual date values from first few rows
        print(f"DEBUG: First 5 date values (raw): {transactions.iloc[:5, 0].tolist()}")
        print(f"DEBUG: Date column data type: {transactions.iloc[0, 0].__class__.__name__}")

        return metadata, transactions
    
    def extract_lc_numbers(self, description_series):
        """Extract LC numbers from transaction descriptions."""
        def extract_single_lc(description):
            if pd.isna(description):
                return None
            
            # Pattern for LC numbers: L/C-123/456, LC-123/456, or similar formats
            match = re.search(LC_PATTERN, str(description).upper())
            return match.group() if match else None
        
        return description_series.apply(extract_single_lc)
    
    def extract_po_numbers(self, description_series):
        """Extract PO numbers from transaction descriptions."""
        def extract_single_po(description):
            if pd.isna(description):
                return None
            
            # Pattern for PO numbers: XXX/PO/YYYY/M/NNNNN format
            match = re.search(PO_PATTERN, str(description).upper())
            return match.group() if match else None
        
        return description_series.apply(extract_single_po)
    
    def extract_lc_numbers_from_narration(self, file_path):
        """Extract LC numbers from narration rows (regular text Column C - not bold, not italic) using openpyxl formatting."""
        lc_numbers = []
        lc_parent_rows = []
        
        # Load workbook with openpyxl to access formatting
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        for row in range(9, ws.max_row + 1):  # Start from row 9 (after headers)
            particulars_cell = ws.cell(row=row, column=2)
            desc_cell = ws.cell(row=row, column=3)
            
            # Check if this is a narration row (italic text Column C - not bold, but italic)
            is_narration = (desc_cell.value and 
                           desc_cell.font and 
                           not desc_cell.font.bold and 
                           desc_cell.font.italic)
            
            if is_narration:
                # This is a narration row, check for LC numbers
                narration_text = str(desc_cell.value)
                lc = self.extract_lc_numbers(pd.Series([narration_text])).iloc[0]
                
                if lc is not None:
                    # Found LC in narration row, need to find parent transaction row
                    parent_row = self.find_parent_transaction_row_with_formatting(ws, row)
                    if parent_row is not None:
                        print(f"DEBUG: LC {lc} at narration row {row} linked to parent row {parent_row}")
                        lc_numbers.append(lc)
                        lc_parent_rows.append(parent_row)
                    else:
                        print(f"DEBUG: LC {lc} at narration row {row} - NO PARENT FOUND!")
                        lc_numbers.append(None)
                        lc_parent_rows.append(None)
                else:
                    lc_numbers.append(None)
                    lc_parent_rows.append(None)
            else:
                lc_numbers.append(None)
                lc_parent_rows.append(None)
        
        wb.close()
        
        # Store parent row mapping for later use
        
        return pd.Series(lc_numbers)
    
    def extract_po_numbers_from_narration(self, file_path):
        """Extract PO numbers from narration rows (italic text Column C - not bold, but italic) using openpyxl formatting."""
        # Load workbook with openpyxl to access formatting
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Create a list to store PO numbers for each row in the DataFrame
        # We need to align with the transactions DataFrame structure
        po_numbers = []
        po_parent_rows = []
        
        # First, get the transactions DataFrame to know how many rows we need
        transactions_df = self.read_complex_excel(file_path)[1]  # Get transactions part
        total_rows = len(transactions_df)
        
        # Initialize with None for all rows
        for i in range(total_rows):
            po_numbers.append(None)
            po_parent_rows.append(None)
        
        # Now scan for PO numbers in narration rows and map them to DataFrame indices
        for excel_row in range(9, ws.max_row + 1):  # Excel rows start from 9
            particulars_cell = ws.cell(row=excel_row, column=2)
            desc_cell = ws.cell(row=excel_row, column=3)
            
            # Check if this is a narration row (italic text Column C - not bold, but italic)
            is_narration = (desc_cell.value and 
                           desc_cell.font and 
                           not desc_cell.font.bold and 
                           desc_cell.font.italic)
            
            if is_narration:
                # This is a narration row, check for PO numbers
                narration_text = str(desc_cell.value)
                po = self.extract_po_numbers(pd.Series([narration_text])).iloc[0]
                
                if po is not None:
                    # Found PO in narration row, need to find parent transaction row
                    parent_row = self.find_parent_transaction_row_with_formatting(ws, excel_row)
                    if parent_row is not None:
                        # Convert Excel row to DataFrame index
                        df_index = parent_row - 9  # Excel row 9 = DataFrame index 0
                        if 0 <= df_index < total_rows:
                            print(f"DEBUG: PO {po} at Excel row {excel_row} -> DataFrame index {df_index}")
                            po_numbers[df_index] = po
                            po_parent_rows[df_index] = df_index
                        else:
                            print(f"DEBUG: PO {po} at Excel row {excel_row} - INVALID DataFrame index {df_index}")
                    else:
                        print(f"DEBUG: PO {po} at Excel row {excel_row} - NO PARENT FOUND!")
        
        wb.close()
        
        # Store parent row mapping for later use
        # self.po_parent_mapping = dict(zip(range(len(po_numbers)), po_parent_rows))  # ❌ UNUSED - commenting out
        
        return pd.Series(po_numbers)
    
    # ❌ UNUSED METHOD - commenting out
    # def find_parent_transaction_row(self, current_row, transactions_df):
    #     """Find the parent transaction row for a description row."""
    #     # Look backwards from current row to find the most recent transaction row
    #     for row_idx in range(current_row - 1, -1, -1):  # Start from current_row - 1, not current_row
    #         row = transactions_df.iloc[row_idx]
    #         has_date = pd.notna(row.iloc[0])  # Date column
    #         has_debit = pd.notna(row.iloc[7]) and float(row.iloc[7]) > 0  # Debit column
    #         has_credit = pd.notna(row.iloc[8]) and float(row.iloc[8]) > 0  # Credit column
    #             
    #         if has_date and (has_debit or has_credit):
    #             return row_idx
    #     
    #     # If no parent found looking backwards, look forwards
    #     for row_idx in range(current_row + 1, len(transactions_df)):
    #         row = transactions_df.iloc[row_idx]
    #         has_date = pd.notna(row.iloc[0])  # Date column
    #         has_debit = pd.notna(row.iloc[7]) and float(row.iloc[7]) > 0  # Debit column
    #         has_credit = pd.notna(row.iloc[8]) and float(row.iloc[8]) > 0  # Credit column
    #             
    #         if has_date and (has_debit or has_credit):
    #             return row_idx
    #     
    #     return None
    
    # ❌ UNUSED METHOD - commenting out
    # def identify_transaction_blocks(self, transactions_df):
    #     """Identify transaction blocks based on date+Dr/Cr start and next date+Dr/Cr end."""
    #     blocks = []
    #     current_block = []
    #     in_block = False
    #     
    #     for idx, row in transactions_df.iterrows():
    #         # Check if row has date in Col A and Dr/Cr in Col B (block start/end)
    #         has_date = pd.notna(row.iloc[0])  # Col A (Date)
    #         has_dr_cr = pd.notna(row.iloc[1]) and str(row.iloc[1]).strip() in ['Dr', 'Cr']  # Col B (Particulars)
    #         
    #         # Check if this is a new block start (date + Dr/Cr)
    #         is_new_block_start = has_date and has_dr_cr
    #         
    #         if is_new_block_start:
    #             # If we're already in a block, end the current one
    #             if in_block and current_block:
    #                 blocks.append(current_block)
    #             
    #             # Start new block
    #             current_block = [row]
    #                 in_block = True
    #         elif in_block:
    #             # Continue adding rows to current block
    #             current_block.append(row)
    #     
    #         # Add the last block if it exists
    #         if current_block:
    #             blocks.append(current_block)
    #     
    #         return blocks
    
    # ❌ UNUSED METHOD - commenting out
    # def identify_transaction_blocks_with_formatting(self, file_path):
    #     """Identify transaction blocks using openpyxl to check bold formatting in Column C."""
    #     blocks = []
    #     
    #     # Load workbook with openpyxl to access formatting
    #     wb = openpyxl.load_workbook(file_path)
    #     ws = wb.active
    #     
    #     current_block = []
    #     in_block = False
    #     
    #     for row in range(9, ws.max_row + 1):  # Start from row 9 (after headers)
    #         date_cell = ws.cell(row=row, column=1)
    #         particulars_cell = ws.cell(row=row, column=2)
    #         desc_cell = ws.cell(row=row, column=3)
    #         
    #         # Check if this is a transaction block header (Date + Dr/Cr + BOLD Col C)
    #         has_date = date_cell.value is not None
    #         has_dr_cr = particulars_cell.value and str(particulars_cell.value).strip() in ['Dr', 'Cr']
    #         has_bold_desc = desc_cell.font and desc_cell.font.bold
    #             
    #         if has_date and has_dr_cr and has_bold_desc:
    #             # If we're already in a block, end the current one
    #             if in_block and current_block:
    #                 blocks.append(current_block)
    #             
    #             # Start new block
    #                 current_block = [row]
    #                 in_block = True
    #         elif in_block:
    #             # Continue adding rows to current block
    #                 current_block.append(row)
    #     
    #         # Add the last block if it exists
    #         if current_block:
    #             blocks.append(current_block)
    #     
    #         wb.close()
    #         return blocks
    
    def load_workbooks_and_extract_data(self):
        """
        Load Excel workbooks once and extract all required data in a single pass.
        This optimization reduces file I/O operations from 4 to 1 per file.
        """
        print("Loading workbooks and extracting data...")
        
        # Load File 1 workbook once
        wb1 = openpyxl.load_workbook(self.file1_path, data_only=True)
        ws1 = wb1.active
        
        # Load File 2 workbook once  
        wb2 = openpyxl.load_workbook(self.file2_path, data_only=True)
        ws2 = wb2.active
        
        # Extract all data from File 1
        lc_numbers1 = []
        po_numbers1 = []
        usd_amounts1 = []
        interunit_accounts1 = []
        
        # Process File 1 rows 9 onwards (same logic as individual methods)
        for row in range(9, ws1.max_row + 1):
            narration = ws1.cell(row=row, column=3).value  # Column C is narration
            if narration:
                # Extract LC numbers
                lc_matches = re.findall(LC_PATTERN, str(narration).upper())
                if lc_matches:
                    lc_numbers1.append((row, lc_matches[0]))
                
                # Extract PO numbers
                po_matches = re.findall(PO_PATTERN, str(narration).upper())
                if po_matches:
                    po_numbers1.append((row, po_matches[0]))
                
                # Extract USD amounts
                usd_matches = re.findall(USD_PATTERN, str(narration).upper())
                if usd_matches:
                    usd_amounts1.append((row, usd_matches[0]))
                
                # Extract interunit accounts (using the same pattern as interunit_loan_matching_logic)
                interunit_matches = re.findall(r'([A-Z]{2,4})#(\d{4,6})', str(narration).upper())
                if interunit_matches:
                    interunit_accounts1.append((row, f"{interunit_matches[0][0]}#{interunit_matches[0][1]}"))
        
        # Convert to Series with proper indexing (matching original logic)
        # Create Series with same length as transactions DataFrame, initialized with None
        total_rows1 = len(self.transactions1)
        lc_numbers1_series = pd.Series([None] * total_rows1, index=range(total_rows1))
        po_numbers1_series = pd.Series([None] * total_rows1, index=range(total_rows1))
        usd_amounts1_series = pd.Series([None] * total_rows1, index=range(total_rows1))
        interunit_accounts1_series = pd.Series([None] * total_rows1, index=range(total_rows1))
        
        # Now populate the found items at their correct DataFrame indices
        for row, lc_num in lc_numbers1:
            df_index = row - 9  # Excel row 9 = DataFrame index 0
            if 0 <= df_index < total_rows1:
                lc_numbers1_series.iloc[df_index] = lc_num
        
        for row, po_num in po_numbers1:
            df_index = row - 9  # Excel row 9 = DataFrame index 0
            if 0 <= df_index < total_rows1:
                po_numbers1_series.iloc[df_index] = po_num
        
        for row, usd_amount in usd_amounts1:
            df_index = row - 9  # Excel row 9 = DataFrame index 0
            if 0 <= df_index < total_rows1:
                usd_amounts1_series.iloc[df_index] = usd_amount
        
        for row, account in interunit_accounts1:
            df_index = row - 9  # Excel row 9 = DataFrame index 0
            if 0 <= df_index < total_rows1:
                interunit_accounts1_series.iloc[df_index] = account
        
        # Extract all data from File 2
        lc_numbers2 = []
        po_numbers2 = []
        usd_amounts2 = []
        interunit_accounts2 = []
        
        # Process File 2 rows 9 onwards
        for row in range(9, ws2.max_row + 1):
            narration = ws2.cell(row=row, column=3).value  # Column C is narration
            if narration:
                # Extract LC numbers
                lc_matches = re.findall(LC_PATTERN, str(narration).upper())
                if lc_matches:
                    lc_numbers2.append((row, lc_matches[0]))
                
                # Extract PO numbers
                po_matches = re.findall(PO_PATTERN, str(narration).upper())
                if po_matches:
                    po_numbers2.append((row, po_matches[0]))
                
                # Extract USD amounts
                usd_matches = re.findall(USD_PATTERN, str(narration).upper())
                if usd_matches:
                    usd_amounts2.append((row, usd_matches[0]))
                
                # Extract interunit accounts (using the same pattern as interunit_loan_matching_logic)
                interunit_matches = re.findall(r'([A-Z]{2,4})#(\d{4,6})', str(narration).upper())
                if interunit_matches:
                    interunit_accounts2.append((row, f"{interunit_matches[0][0]}#{interunit_matches[0][1]}"))
        
        # Convert to Series with proper indexing (matching original logic)
        # Create Series with same length as transactions DataFrame, initialized with None
        total_rows2 = len(self.transactions2)
        lc_numbers2_series = pd.Series([None] * total_rows2, index=range(total_rows2))
        po_numbers2_series = pd.Series([None] * total_rows2, index=range(total_rows2))
        usd_amounts2_series = pd.Series([None] * total_rows2, index=range(total_rows2))
        interunit_accounts2_series = pd.Series([None] * total_rows2, index=range(total_rows2))
        
        # Now populate the found items at their correct DataFrame indices
        for row, lc_num in lc_numbers2:
            df_index = row - 9  # Excel row 9 = DataFrame index 0
            if 0 <= df_index < total_rows2:
                lc_numbers2_series.iloc[df_index] = lc_num
        
        for row, po_num in po_numbers2:
            df_index = row - 9  # Excel row 9 = DataFrame index 0
            if 0 <= df_index < total_rows2:
                po_numbers2_series.iloc[df_index] = po_num
        
        for row, usd_amount in usd_amounts2:
            df_index = row - 9  # Excel row 9 = DataFrame index 0
            if 0 <= df_index < total_rows2:
                usd_amounts2_series.iloc[df_index] = usd_amount
        
        for row, account in interunit_accounts2:
            df_index = row - 9  # Excel row 9 = DataFrame index 0
            if 0 <= df_index < total_rows2:
                interunit_accounts2_series.iloc[df_index] = account
        
        # Close workbooks
        wb1.close()
        wb2.close()
        
        print(f"Data extraction complete:")
        print(f"  File 1: {len(lc_numbers1)} LC, {len(po_numbers1)} PO, {len(usd_amounts1)} USD, {len(interunit_accounts1)} Interunit")
        print(f"  File 2: {len(lc_numbers2)} LC, {len(po_numbers2)} PO, {len(usd_amounts2)} USD, {len(interunit_accounts2)} Interunit")
        
        return {
            'lc_numbers1': lc_numbers1_series,
            'po_numbers1': po_numbers1_series,
            'usd_amounts1': usd_amounts1_series,
            'interunit_accounts1': interunit_accounts1_series,
            'lc_numbers2': lc_numbers2_series,
            'po_numbers2': po_numbers2_series,
            'usd_amounts2': usd_amounts2_series,
            'interunit_accounts2': interunit_accounts2_series
        }

    def process_files(self):
        """Process both files and prepare for matching."""
        print("Reading Pole Book STEEL.xlsx...")
        self.metadata1, self.transactions1 = self.read_complex_excel(self.file1_path)
        
        print("Reading Steel Book POLE.xlsx...")
        self.metadata2, self.transactions2 = self.read_complex_excel(self.file2_path)
        
        print(f"File 1: {len(self.transactions1)} rows")
        print(f"File 2: {len(self.transactions2)} rows")
        
        # DEBUG: Show column names for both files
        print(f"File 1 columns: {list(self.transactions1.columns)}")
        print(f"File 2 columns: {list(self.transactions2.columns)}")
        
        # Find the description column (should be the 3rd column, index 2)
        # Let's check what's actually in the columns
        print(f"File 1 first row: {list(self.transactions1.iloc[0, :])}")
        
        # Load workbooks once and extract all data in a single pass
        print("Loading workbooks and extracting all data...")
        extracted_data = self.load_workbooks_and_extract_data()
        
        # Extract LC numbers from both files
        lc_numbers1 = extracted_data['lc_numbers1']
        lc_numbers2 = extracted_data['lc_numbers2']
        
        # Extract PO numbers from both files
        po_numbers1 = extracted_data['po_numbers1']
        po_numbers2 = extracted_data['po_numbers2']
        
        # Extract interunit loan accounts from both files
        interunit_accounts1 = extracted_data['interunit_accounts1']
        interunit_accounts2 = extracted_data['interunit_accounts2']
        
        # Extract USD amounts from both files
        usd_amounts1 = extracted_data['usd_amounts1']
        usd_amounts2 = extracted_data['usd_amounts2']
        
        # Identify transaction blocks using formatting
        print("Identifying transaction blocks using formatting...")
        blocks1 = self.block_identifier.identify_transaction_blocks(self.transactions1, self.file1_path)
        blocks2 = self.block_identifier.identify_transaction_blocks(self.transactions2, self.file2_path)
        
        print(f"File 1: {len(blocks1)} transaction blocks")
        print(f"File 2: {len(blocks2)} transaction blocks")
        
        return self.transactions1, self.transactions2, blocks1, blocks2, lc_numbers1, lc_numbers2, po_numbers1, po_numbers2, interunit_accounts1, interunit_accounts2, usd_amounts1, usd_amounts2
    
    def find_potential_matches(self):
        """Find potential LC, PO, Interunit, and USD matches between the two files (sequential approach)."""
        transactions1, transactions2, blocks1, blocks2, lc_numbers1, lc_numbers2, po_numbers1, po_numbers2, interunit_accounts1, interunit_accounts2, usd_amounts1, usd_amounts2 = self.process_files()
        
        print("\n" + "="*60)
        print("STEP 1: LC MATCHING")
        print("="*60)
        
        # Initialize shared state for consistent Match IDs across LC and PO matching
        shared_existing_matches = {}
        shared_match_counter = 0
        
        # Step 1: Find LC matches
        lc_matches = self.lc_matching_logic.find_potential_matches(
            transactions1, transactions2, lc_numbers1, lc_numbers2,
            shared_existing_matches, shared_match_counter
        )
        
        # Update the shared counter after LC matching
        if lc_matches:
            shared_match_counter = max(int(match['match_id'][1:]) for match in lc_matches)
        
        print(f"\nLC Matching Results: {len(lc_matches)} matches found")
        
        # Step 2: Find PO matches on UNMATCHED records
        print("\n" + "="*60)
        print("STEP 2: PO MATCHING (ON UNMATCHED RECORDS)")
        print("="*60)
        
        # Create masks for unmatched records
        lc_matched_indices1 = set()
        lc_matched_indices2 = set()
        
        for match in lc_matches:
            lc_matched_indices1.add(match['File1_Index'])
            lc_matched_indices2.add(match['File2_Index'])
        
        # Filter PO numbers to only unmatched records
        po_numbers1_unmatched = po_numbers1.copy()
        po_numbers2_unmatched = po_numbers2.copy()
        
        # Mark matched records as None in PO numbers
        for idx in lc_matched_indices1:
            if idx < len(po_numbers1_unmatched):
                po_numbers1_unmatched.iloc[idx] = None
        
        for idx in lc_matched_indices2:
            if idx < len(po_numbers2_unmatched):
                po_numbers2_unmatched.iloc[idx] = None
        
        print(f"File 1: {len(po_numbers1_unmatched[po_numbers1_unmatched.notna()])} unmatched PO numbers")
        print(f"File 2: {len(po_numbers2_unmatched[po_numbers2_unmatched.notna()])} unmatched PO numbers")
        
        # Find PO matches on unmatched records with shared state
        po_matches = self.po_matching_logic.find_potential_matches(
            transactions1, transactions2, po_numbers1_unmatched, po_numbers2_unmatched,
            shared_existing_matches, shared_match_counter
        )
        
        # Update the shared counter after PO matching
        if po_matches:
            shared_match_counter = max(int(match['match_id'][1:]) for match in po_matches)
        
        print(f"\nPO Matching Results: {len(po_matches)} matches found")
        
        # Step 3: Find Interunit Loan matches on UNMATCHED records
        print("\n" + "="*60)
        print("STEP 3: INTERUNIT LOAN MATCHING (ON UNMATCHED RECORDS)")
        print("="*60)
        
        # Create masks for unmatched records (after LC and PO matching)
        lc_po_matched_indices1 = set()
        lc_po_matched_indices2 = set()
        
        for match in lc_matches + po_matches:
            lc_po_matched_indices1.add(match['File1_Index'])
            lc_po_matched_indices2.add(match['File2_Index'])
        
        # Filter interunit accounts to only unmatched records
        interunit_accounts1_unmatched = interunit_accounts1.copy()
        interunit_accounts2_unmatched = interunit_accounts2.copy()
        
        # Mark matched records as None in interunit accounts
        for idx in lc_po_matched_indices1:
            if idx < len(interunit_accounts1_unmatched):
                interunit_accounts1_unmatched.iloc[idx] = None
        
        for idx in lc_po_matched_indices2:
            if idx < len(interunit_accounts2_unmatched):
                interunit_accounts2_unmatched.iloc[idx] = None
        
        print(f"File 1: {len(interunit_accounts1_unmatched[interunit_accounts1_unmatched.notna()])} unmatched interunit accounts")
        print(f"File 2: {len(interunit_accounts2_unmatched[interunit_accounts2_unmatched.notna()])} unmatched interunit accounts")
        
        # Find interunit loan matches on unmatched records with shared state
        interunit_matches = self.interunit_loan_matcher.find_potential_matches(
            transactions1, transactions2, interunit_accounts1_unmatched, interunit_accounts2_unmatched,
            self.file1_path, self.file2_path, shared_existing_matches, shared_match_counter
        )
        
        print(f"\nInterunit Loan Matching Results: {len(interunit_matches)} matches found")
        
        # Step 4: Find USD matches on UNMATCHED records
        print("\n" + "="*60)
        print("STEP 4: USD MATCHING (ON UNMATCHED RECORDS)")
        print("="*60)
        
        # Create masks for unmatched records (after LC, PO, and Interunit matching)
        lc_po_interunit_matched_indices1 = set()
        lc_po_interunit_matched_indices2 = set()
        
        for match in lc_matches + po_matches + interunit_matches:
            lc_po_interunit_matched_indices1.add(match['File1_Index'])
            lc_po_interunit_matched_indices2.add(match['File2_Index'])
        
        # Filter USD amounts to only unmatched records
        usd_amounts1_unmatched = usd_amounts1.copy()
        usd_amounts2_unmatched = usd_amounts2.copy()
        
        # Mark matched records as None in USD amounts
        for idx in lc_po_interunit_matched_indices1:
            if idx < len(usd_amounts1_unmatched):
                usd_amounts1_unmatched.iloc[idx] = None
        
        for idx in lc_po_interunit_matched_indices2:
            if idx < len(usd_amounts2_unmatched):
                usd_amounts2_unmatched.iloc[idx] = None
        
        print(f"File 1: {len(usd_amounts1_unmatched[usd_amounts1_unmatched.notna()])} unmatched USD amounts")
        print(f"File 2: {len(usd_amounts2_unmatched[usd_amounts2_unmatched.notna()])} unmatched USD amounts")
        
        # Find USD matches on unmatched records with shared state
        usd_matches = self.usd_matching_logic.find_potential_matches(
            transactions1, transactions2, usd_amounts1_unmatched, usd_amounts2_unmatched,
            shared_existing_matches, shared_match_counter
        )
        
        # Update the shared counter after USD matching
        if usd_matches:
            shared_match_counter = max(int(match['match_id'][1:]) for match in usd_matches)
        
        print(f"\nUSD Matching Results: {len(usd_matches)} matches found")
        
        # Combine all matches
        all_matches = lc_matches + po_matches + interunit_matches + usd_matches
        
        print(f"\n" + "="*60)
        print("FINAL RESULTS")
        print("="*60)
        print(f"Total Matches: {len(all_matches)}")
        print(f"  - LC Matches: {len(lc_matches)}")
        print(f"  - PO Matches: {len(po_matches)}")
        print(f"  - Interunit Loan Matches: {len(interunit_matches)}")
        print(f"  - USD Matches: {len(usd_matches)}")
        
        return all_matches
    
    # ❌ UNUSED METHOD - commenting out
    # def find_transaction_block_header(self, current_row, transactions_df):
    #     """Find the transaction block header row (with date and particulars) for a given row."""
    #     # Look backwards from current row to find the most recent block header
    #     for row_idx in range(current_row, -1, -1):
    #         row = transactions_df.iloc[row_idx]
    #         has_date = pd.notna(row.iloc[0])  # Col A (Date)
    #         has_particulars = pd.notna(row.iloc[1]).strip() in ['Dr', 'Cr']  # Col B (Particulars)
    #             
    #         if has_date and has_particulars:
    #             return row_idx
    #     
    #     # If no header found looking backwards, look forwards
    #     for row_idx in range(current_row + 1, len(transactions_df)):
    #         row = transactions_df.iloc[row_idx]
    #         has_date = pd.notna(row.iloc[0])  # Col A (Date)
    #         has_particulars = pd.notna(row.iloc[1]).strip() in ['Dr', 'Cr']  # Col B (Particulars)
    #             
    #         if has_date and has_particulars:
    #             return row_idx
    #     
    #         return current_row  # Fallback to current row if no header found
    
    def find_parent_transaction_row_with_formatting(self, ws, current_row):
        """Find the parent transaction row for a narration row using openpyxl formatting."""
        # Look backwards from current row to find the most recent transaction block header
        for row_idx in range(current_row, 8, -1):  # Start from current_row, go back to row 9
            date_cell = ws.cell(row=row_idx, column=1)
            particulars_cell = ws.cell(row=row_idx, column=2)
            desc_cell = ws.cell(row=row_idx, column=3)
            
            # Check if this is a transaction block header (Date + Dr/Cr + BOLD Col C)
            has_date = date_cell.value is not None
            has_dr_cr = particulars_cell.value and str(particulars_cell.value).strip() in ['Dr', 'Cr']
            has_bold_desc = desc_cell.font and desc_cell.font.bold
            
            if has_date and has_dr_cr and has_bold_desc:
                return row_idx
        
        return None
    

    

    
    def create_audit_info(self, match):
        """Create audit info in clean, readable plaintext format for LC, PO, Interunit, and USD matches."""
        # Determine match type and create appropriate audit info
        if 'Match_Type' in match:
            # Use explicit match type if available
            match_type = match['Match_Type']
            amount = match.get('File1_Amount', match.get('File2_Amount', 0))
            
            if match_type == 'LC':
                lc_number = match.get('LC_Number', 'Unknown')
                audit_info = f"LC Match: {lc_number}\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
            elif match_type == 'PO':
                po_number = match.get('PO_Number', 'Unknown')
                audit_info = f"PO Match: {po_number}\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
            elif match_type == 'Interunit':
                interunit_account = match.get('Interunit_Account', 'Unknown')
                audit_info = f"Interunit Loan Match: {interunit_account}\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
            elif match_type == 'USD':
                usd_amount = match.get('USD_Amount', 'Unknown')
                audit_info = f"USD Match: {usd_amount}\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
            else:
                audit_info = f"{match_type} Match\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
        else:
            # Fallback to old logic for backward compatibility
            if 'LC_Number' in match and match['LC_Number']:
                # This is an LC match - use File1_Amount or File2_Amount
                amount = match.get('File1_Amount', match.get('File2_Amount', 0))
                audit_info = f"LC Match: {match['LC_Number']}\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
            elif 'PO_Number' in match and match['PO_Number']:
                # This is a PO match - use File1_Amount or File2_Amount
                amount = match.get('File1_Amount', match.get('File2_Amount', 0))
                audit_info = f"PO Match: {match['PO_Number']}\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
            elif 'Interunit_Account' in match and match['Interunit_Account']:
                # This is an Interunit Loan match - use File1_Amount or File2_Amount
                amount = match.get('File1_Amount', match.get('File2_Amount', 0))
                audit_info = f"Interunit Loan Match: {match['Interunit_Account']}\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
            elif 'USD_Amount' in match and match['USD_Amount']:
                # This is a USD match - use File1_Amount or File2_Amount
                amount = match.get('File1_Amount', match.get('File2_Amount', 0))
                audit_info = f"USD Match: {match['USD_Amount']}\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
            else:
                # Fallback for unknown match type
                amount = match.get('File1_Amount', match.get('File2_Amount', 0))
                audit_info = f"Unknown Match Type\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
        
        return audit_info
    
    def _preserve_tally_date_format(self, transactions_df: pd.DataFrame):
        """Ensure dates are in Tally format (e.g., '01/Jul/2024') before saving."""
        if len(transactions_df.columns) > 2:  # After adding Match ID and Audit Info columns
            # Date column is now at index 2 (third column) after adding Match ID and Audit Info
            date_col = transactions_df.iloc[:, 2]  # Third column is date
            
            # Convert any datetime objects or datetime strings back to Tally format strings
            def format_tally_date(date_val):
                if pd.isna(date_val):
                    return date_val
                
                # If it's already a Tally format string, keep it
                if isinstance(date_val, str) and '/' in str(date_val) and len(str(date_val)) <= 12:
                    return date_val
                
                # If it's a datetime object, convert to Tally format
                if hasattr(date_val, 'strftime'):
                    return date_val.strftime('%d/%b/%Y')
                
                # If it's a datetime string (like '2024-07-01 00:00:00'), parse and convert
                if isinstance(date_val, str) and ('-' in str(date_val) or ':' in str(date_val)):
                    try:
                        # Parse the datetime string and convert to Tally format
                        from datetime import datetime
                        parsed_date = pd.to_datetime(date_val)
                        return parsed_date.strftime('%d/%b/%Y')
                    except:
                        return date_val
                
                return date_val
            
            # Apply formatting to date column
            transactions_df.iloc[:, 2] = date_col.apply(format_tally_date)
            
            if VERBOSE_DEBUG:
                print(f"DEBUG: Date format preservation applied. Sample dates: {transactions_df.iloc[:3, 2].tolist()}")
                print(f"DEBUG: Date column index: 2, column name: {transactions_df.columns[2]}")
                print(f"DEBUG: Date types after conversion: {[type(x) for x in transactions_df.iloc[:3, 2]]}")

    def _format_amount_columns(self, worksheet):
        """Format amount columns (Debit and Credit) to prevent scientific notation."""
        # Debit column (J) and Credit column (K) - after adding Match ID and Audit Info
        debit_col = 9  # Column J (0-indexed)
        credit_col = 10  # Column K (0-indexed)
        
        # Format all data rows (starting from row 9)
        for row in range(9, worksheet.max_row + 1):
            try:
                # Format Debit column
                debit_cell = worksheet.cell(row=row, column=debit_col + 1)  # openpyxl uses 1-indexed
                if debit_cell.value is not None and debit_cell.value != '':
                    debit_cell.number_format = '#,##0.00'
                
                # Format Credit column
                credit_cell = worksheet.cell(row=row, column=credit_col + 1)  # openpyxl uses 1-indexed
                if credit_cell.value is not None and credit_cell.value != '':
                    credit_cell.number_format = '#,##0.00'
                    
            except Exception as e:
                print(f"Error formatting amount columns for row {row}: {e}")

    def _set_column_widths(self, worksheet):
        """Set column widths for the worksheet"""
        worksheet.column_dimensions['A'].width = 9.00
        worksheet.column_dimensions['B'].width = 30.00
        worksheet.column_dimensions['C'].width = 12.00
        worksheet.column_dimensions['D'].width = 10.33
        worksheet.column_dimensions['E'].width = 60.00
        worksheet.column_dimensions['F'].width = 5.00
        worksheet.column_dimensions['G'].width = 5.00
        worksheet.column_dimensions['H'].width = 12.78
        worksheet.column_dimensions['I'].width = 9.00
        worksheet.column_dimensions['J'].width = 13.78
        worksheet.column_dimensions['K'].width = 14.22
        worksheet.column_dimensions['L'].width = 11.22

    def _apply_top_alignment(self, worksheet):
        """Apply top alignment and text wrapping to ALL cells in the worksheet."""
        print(f"Setting top alignment for {worksheet.max_row} rows × {worksheet.max_column} columns...")
        
        for row in range(1, worksheet.max_row + 1):  # ALL rows from 1 to max
            for col in range(1, worksheet.max_column + 1):  # ALL columns
                try:
                    cell = worksheet.cell(row=row, column=col)
                    
                    # Always create a new alignment object to avoid style conflicts
                    new_alignment = Alignment(vertical='top')
                    
                    # Enable text wrapping for columns B (Audit Info) and E (Description)
                    if col in [2, 5]:  # Columns B and E
                        new_alignment.wrap_text = True
                    
                    # Apply the new alignment (this overwrites any existing alignment)
                    cell.alignment = new_alignment
                        
                except Exception as e:
                    print(f"Error setting alignment for row {row}, col {col}: {e}")
                    # Continue with next cell instead of stopping
                    continue
        
        print(f"Top alignment applied successfully!")

    def _apply_filters_to_header(self, worksheet):
        """Apply filters to the header row (Row 9) for easy data filtering and sorting."""
        try:
            # Apply filters to Row 9 (header row)
            # Note: openpyxl uses 1-based indexing, so Row 9 is actually row 9
            worksheet.auto_filter.ref = f"A9:L9"
            print(f"Filters applied to header row (Row 9) successfully!")
        except Exception as e:
            print(f"Error applying filters to header row: {e}")
    
    def _apply_alternating_background_colors(self, worksheet, file_matched_df):
        """Apply alternating background colors to matched transaction blocks."""
        try:
            from openpyxl.styles import PatternFill
            
            # Define two alternating colors
            color1 = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")  # Very light blue
            color2 = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")  # Very light lemon yellow
            
            # Get all rows with Match IDs
            match_id_column = file_matched_df.iloc[:, 0]  # First column (Match ID)
            populated_rows = match_id_column.notna()
            
            if not populated_rows.any():
                print("No matched rows found for background coloring")
                return
            
            # Get unique Match IDs in order they appear
            unique_match_ids = []
            seen_ids = set()
            for idx, match_id in enumerate(match_id_column):
                if pd.notna(match_id) and match_id not in seen_ids:
                    unique_match_ids.append(match_id)
                    seen_ids.add(match_id)
            
            print(f"Applying alternating background colors to {len(unique_match_ids)} matched transaction blocks")
            
            # Apply alternating colors to each Match ID block
            for block_index, match_id in enumerate(unique_match_ids):
                # Choose color based on block index (alternating)
                color = color1 if block_index % 2 == 0 else color2
                
                # Find all rows with this Match ID
                block_rows = file_matched_df[file_matched_df.iloc[:, 0] == match_id].index
                
                # Apply color to all rows in this block
                for df_row_idx in block_rows:
                    excel_row = df_row_idx + 10  # Convert DataFrame index to Excel row (metadata + header offset)
                    
                    # Color all columns in this row
                    for col in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=excel_row, column=col)
                        cell.fill = color
                
                print(f"  Block {match_id}: Applied {'Color 1' if block_index % 2 == 0 else 'Color 2'} to {len(block_rows)} rows")
            
            print("Background colors applied successfully!")
            
        except Exception as e:
            print(f"Error applying background colors: {e}")

    def _format_output_file_transaction_blocks(self, worksheet):
        """Format output file transaction blocks: make ledger text bold, narration italic, and Entered By person's name bold+italic."""
        try:
            from openpyxl.styles import Font
            
            print("Formatting output file transaction blocks...")
            
            # Create fonts for different formatting
            bold_font = Font(bold=True)
            italic_font = Font(italic=True)
            bold_italic_font = Font(bold=True, italic=True)
            
            # Process all rows starting from row 10 (after metadata and header)
            for row in range(10, worksheet.max_row + 1):
                # Check if this row is the end of a transaction block
                cell_d = worksheet.cell(row=row, column=4)  # Column D (Particulars)
                cell_e = worksheet.cell(row=row, column=5)  # Column E (Description)
                
                # Check if this row contains "Entered By :" in Column D
                if (cell_d.value and 
                    isinstance(cell_d.value, str) and 
                    "Entered By :" in str(cell_d.value)):
                    
                    # This is the end of a transaction block
                    # Make the Entered By person's name bold and italic
                    if cell_e.value:
                        cell_e.font = bold_italic_font
                        print(f"  Row {row}: Made Entered By person's name bold+italic: '{str(cell_e.value)[:50]}...'")
                    
                    # The row above this contains narration text
                    narration_row = row - 1
                    
                    if narration_row >= 10:  # Ensure we don't go below row 10
                        narration_cell_e = worksheet.cell(row=narration_row, column=5)  # Column E
                        
                        # Make the narration text italic
                        if narration_cell_e.value:
                            narration_cell_e.font = italic_font
                            print(f"  Row {narration_row}: Made narration text italic: '{str(narration_cell_e.value)[:50]}...'")
                        
                        # Now find the transaction block start and make all ledger text bold
                        # Look backwards from narration row to find block start
                        for ledger_row in range(narration_row - 1, 9, -1):  # Go back from narration to row 10
                            # Check if this is a block start row
                            date_cell = worksheet.cell(row=ledger_row, column=3)  # Column C (Date)
                            particulars_cell = worksheet.cell(row=ledger_row, column=4)  # Column D (Particulars)
                            vch_type_cell = worksheet.cell(row=ledger_row, column=8)  # Column H (Vch Type)
                            vch_no_cell = worksheet.cell(row=ledger_row, column=9)  # Column I (Vch No)
                            
                            # Check if this is a block start (has date, Dr/Cr, Vch Type, Vch No)
                            is_block_start = (date_cell.value and 
                                            particulars_cell.value and 
                                            str(particulars_cell.value).strip() in ['Dr', 'Cr'] and
                                            vch_type_cell.value and 
                                            vch_no_cell.value)
                            
                            if is_block_start:
                                 # Found block start, now make all rows from here to narration bold
                                 for bold_row in range(ledger_row, narration_row):
                                     bold_cell_e = worksheet.cell(row=bold_row, column=5)  # Column E
                                     if bold_cell_e.value:
                                         bold_cell_e.font = bold_font
                                         print(f"  Row {bold_row}: Made ledger text bold: '{str(bold_cell_e.value)[:50]}...'")
                                 
                                 # Also make Column H (Vch Type) bold in the first row of the transaction block
                                 vch_type_cell = worksheet.cell(row=ledger_row, column=8)  # Column H (Vch Type)
                                 if vch_type_cell.value:
                                     vch_type_cell.font = bold_font
                                     print(f"  Row {ledger_row}: Made Vch Type bold: '{str(vch_type_cell.value)[:50]}...'")
                                 
                                 # Make all Debit and Credit values (Columns J and K) bold in this transaction block
                                 for bold_row in range(ledger_row, narration_row):
                                     # Make Debit column (J) bold
                                     debit_cell = worksheet.cell(row=bold_row, column=10)  # Column J (Debit)
                                     if debit_cell.value and debit_cell.value != '':
                                         debit_cell.font = bold_font
                                         print(f"  Row {bold_row}: Made Debit value bold: '{str(debit_cell.value)[:20]}...'")
                                     
                                     # Make Credit column (K) bold
                                     credit_cell = worksheet.cell(row=bold_row, column=11)  # Column K (Credit)
                                     if credit_cell.value and credit_cell.value != '':
                                         credit_cell.font = bold_font
                                         print(f"  Row {bold_row}: Made Credit value bold: '{str(credit_cell.value)[:20]}...'")
                                 
                                 break
            
            # Also check for "Opening Balance" text and make it bold along with its Debit/Credit values
            print("Checking for Opening Balance entries...")
            for row in range(10, worksheet.max_row + 1):
                cell_e = worksheet.cell(row=row, column=5)  # Column E (Description)
                
                # Check if this row contains "Opening Balance" text
                if (cell_e.value and 
                    isinstance(cell_e.value, str) and 
                    "Opening Balance" in str(cell_e.value)):
                    
                    # Make the Opening Balance text bold
                    cell_e.font = bold_font
                    print(f"  Row {row}: Made Opening Balance text bold: '{str(cell_e.value)[:50]}...'")
                    
                    # Make the associated Debit and Credit values bold
                    debit_cell = worksheet.cell(row=row, column=10)  # Column J (Debit)
                    if debit_cell.value and debit_cell.value != '':
                        debit_cell.font = bold_font
                        print(f"  Row {row}: Made Opening Balance Debit value bold: '{str(debit_cell.value)[:20]}...'")
                    
                    credit_cell = worksheet.cell(row=row, column=11)  # Column K (Credit)
                    if credit_cell.value and credit_cell.value != '':
                        credit_cell.font = bold_font
                        print(f"  Row {row}: Made Opening Balance Credit value bold: '{str(credit_cell.value)[:20]}...'")
            
            print("Output file transaction block formatting completed successfully!")
            
        except Exception as e:
            print(f"Error formatting output file transaction blocks: {e}")



















    def create_matched_files(self, matches, transactions1, transactions2):
        """Create matched versions of both files with new columns."""
        if not matches:
            print("No matches found. Cannot create matched files.")
            return
        
        # Create file1 with new columns
        file1_matched = transactions1.copy()
        
        # Create new columns with proper names
        match_id_col = pd.Series([None] * len(file1_matched), name='Match ID')
        audit_info_col = pd.Series([None] * len(file1_matched), name='Audit Info')
        match_type_col = pd.Series([None] * len(file1_matched), name='Match Type')
        
        # Concatenate new columns with existing data
        file1_matched = pd.concat([match_id_col, audit_info_col, file1_matched, match_type_col], axis=1)
        
        print(f"DEBUG: File1 DataFrame created with shape: {file1_matched.shape}")
        print(f"DEBUG: File1 columns: {list(file1_matched.columns)}")
        
        # Create file2 with new columns
        file2_matched = transactions2.copy()
        
        # Create new columns with proper names
        match_id_col2 = pd.Series([None] * len(file2_matched), name='Match ID')
        audit_info_col2 = pd.Series([None] * len(file2_matched), name='Audit Info')
        match_type_col2 = pd.Series([None] * len(file2_matched), name='Match Type')
        
        # Concatenate new columns with existing data
        file2_matched = pd.concat([match_id_col2, audit_info_col2, file2_matched, match_type_col2], axis=1)
        
        print(f"DEBUG: File2 DataFrame created with shape: {file2_matched.shape}")
        print(f"DEBUG: File2 columns: {list(file2_matched.columns)}")
        
        print(f"DEBUG: Added Match Type column to both DataFrames")
        print(f"DEBUG: File1 columns: {list(file1_matched.columns)}")
        print(f"DEBUG: File2 columns: {list(file2_matched.columns)}")
        
        # Verify the new columns are actually there
        print(f"DEBUG: File1 first few rows of Match ID column:")
        print(file1_matched.iloc[:5, 0].tolist())
        print(f"DEBUG: File1 first few rows of Audit Info column:")
        print(file1_matched.iloc[:5, 1].tolist())
        
        print(f"\n=== DEBUG: MATCH DATA POPULATION ===")
        
        # Populate match information
        for match in matches:
            match_id = match['match_id']  # Use the pre-assigned match ID
            audit_info = self.create_audit_info(match)
            
            print(f"Match {match_id}:")
            # Use the explicit Match_Type field if available, otherwise fall back to inference
            if 'Match_Type' in match and match['Match_Type']:
                match_type = match['Match_Type']
                print(f"  Match Type: {match_type} (from explicit field)")
            elif 'LC_Number' in match and match['LC_Number']:
                print(f"  LC Number: {match['LC_Number']}")
                match_type = 'LC'
            elif 'PO_Number' in match and match['PO_Number']:
                print(f"  PO Number: {match['PO_Number']}")
                match_type = 'PO'
            elif 'Interunit_Account' in match and match['Interunit_Account']:
                print(f"  Interunit Account: {match['Interunit_Account']}")
                match_type = 'Interunit'
            else:
                print(f"  Unknown Match Type")
                match_type = 'Unknown'
            print(f"  File1 Row {match['File1_Index']}: Debit={match['File1_Debit']}, Credit={match['File1_Credit']}")
            print(f"  File2 Row {match['File2_Index']}: Debit={match['File2_Debit']}, Credit={match['File2_Credit']}")
            print(f"  Audit Info: {audit_info}")
            print(f"  Match Type: {match_type}")
            
            # Update file1 - populate entire transaction block with Match ID and Audit Info
            file1_row_idx = match['File1_Index']
            print(f"    DEBUG: Setting File1 row {file1_row_idx} col 0 to '{match_id}'")
            print(f"    DEBUG: Setting File1 row {file1_row_idx} col 1 to '{audit_info[:50]}...'")
            print(f"    DEBUG: Setting File1 row {file1_row_idx} col -1 to '{match_type}' (last column)")
            
            # Find the entire transaction block for file1 and populate all rows
            file1_block_rows = self.block_identifier.get_transaction_block_rows(file1_row_idx, self.file1_path)
            print(f"    DEBUG: File1 transaction block spans rows: {file1_block_rows}")
            
            # Populate ALL rows of the transaction block with Match ID and Match Type, but Audit Info only in second-to-last row
            for i, block_row in enumerate(file1_block_rows):
                if 0 <= block_row < len(file1_matched):
                    file1_matched.iloc[block_row, 0] = match_id  # Match ID column (index 0)
                    file1_matched.iloc[block_row, -1] = match_type  # Match Type column (last column) - ALL ROWS
                    
                    # Audit Info goes ONLY in the second-to-last row of the transaction block
                    if i == len(file1_block_rows) - 2:  # Second-to-last row
                        file1_matched.iloc[block_row, 1] = audit_info  # Audit Info column (index 1)
                        print(f"    DEBUG: Populated File1 row {block_row} with Match ID '{match_id}', Audit Info, and Match Type '{match_type}' (second-to-last row)")
                    else:
                        print(f"    DEBUG: Populated File1 row {block_row} with Match ID '{match_id}' and Match Type '{match_type}'")
            

            
            # Update file2 - populate entire transaction block with Match ID and Audit Info
            file2_row_idx = match['File2_Index']
            print(f"    DEBUG: Setting File2 row {file2_row_idx} col 0 to '{match_id}'")
            print(f"    DEBUG: Setting File2 row {file2_row_idx} col 1 to '{audit_info[:50]}...'")
            print(f"    DEBUG: Setting File2 row {file2_row_idx} col -1 to '{match_type}' (last column)")
            
            # Find the entire transaction block for file2 and populate all rows
            file2_block_rows = self.block_identifier.get_transaction_block_rows(file2_row_idx, self.file2_path)
            print(f"    DEBUG: File2 transaction block spans rows: {file2_block_rows}")
            
            # Populate ALL rows of the transaction block with Match ID and Match Type, but Audit Info only in second-to-last row
            for i, block_row in enumerate(file2_block_rows):
                if 0 <= block_row < len(file2_matched):
                    file2_matched.iloc[block_row, 0] = match_id  # Match ID column (index 0)
                    file2_matched.iloc[block_row, -1] = match_type  # Match Type column (last column) - ALL ROWS
                    
                    # Audit Info goes ONLY in the second-to-last row of the transaction block
                    if i == len(file2_block_rows) - 2:  # Second-to-last row
                        file2_matched.iloc[block_row, 1] = audit_info  # Audit Info column (index 1)
                        print(f"    DEBUG: Populated File2 row {block_row} with Match ID '{match_id}', Audit Info, and Match Type '{match_type}' (second-to-last row)")
                    else:
                        print(f"    DEBUG: Populated File2 row {block_row} with Match ID '{match_id}' and Match Type '{match_type}'")
        
        # Save matched files using configuration variables
        base_name1 = os.path.splitext(os.path.basename(self.file1_path))[0]
        base_name2 = os.path.splitext(os.path.basename(self.file2_path))[0]
        
        output_file1 = os.path.join(OUTPUT_FOLDER, f"{base_name1}{OUTPUT_SUFFIX}")
        output_file2 = os.path.join(OUTPUT_FOLDER, f"{base_name2}{OUTPUT_SUFFIX}")
        
        if VERBOSE_DEBUG:
            print(f"\n=== DEBUG: BEFORE SAVING ===")
            print(f"File1 - Rows with Match IDs: {file1_matched.iloc[:, 0].notna().sum()}")
            print(f"File1 - Rows with Audit Info: {file1_matched.iloc[:, 1].notna().sum()}")
            print(f"File1 - Rows with Match Type: {file1_matched.iloc[:, -1].notna().sum()}")
            print(f"File2 - Rows with Match IDs: {file2_matched.iloc[:, 0].notna().sum()}")
            print(f"File2 - Rows with Audit Info: {file2_matched.iloc[:, 1].notna().sum()}")
            print(f"File2 - Rows with Match Type: {file2_matched.iloc[:, -1].notna().sum()}")
            
            # Show some actual values to verify they're there
            print(f"\n=== DEBUG: ACTUAL VALUES IN DATAFRAME ===")
            
            # Get the actual populated rows dynamically
            populated_rows = file1_matched.iloc[:, 0].notna()
            if populated_rows.any():
                populated_indices = file1_matched[populated_rows].index
                for idx in populated_indices[:4]:  # Show first 4 populated rows
                    print(f"File1 - Row {idx} Match ID: '{file1_matched.iloc[idx, 0]}'")
                    print(f"File1 - Row {idx} Audit Info: '{file1_matched.iloc[idx, 1]}'")
            else:
                print("No populated rows found in File1")
        
        # Preserve Tally date format before saving
        print("\n=== PRESERVING TALLY DATE FORMAT ===")
        self._preserve_tally_date_format(file1_matched)
        self._preserve_tally_date_format(file2_matched)
        
        # Create output with metadata + matched transactions
        with pd.ExcelWriter(output_file1, engine='openpyxl') as writer:
            # Write metadata
            self.metadata1.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            # Write matched transactions (skip first 8 rows to make room for metadata)
            file1_matched.to_excel(writer, sheet_name='Sheet1', index=False, header=True, startrow=8)
            
            # Get the worksheet to set column widths
            worksheet = writer.sheets['Sheet1']
            self._set_column_widths(worksheet)
            self._format_amount_columns(worksheet) # Apply amount formatting
            
            # Apply top alignment and text wrapping to ALL cells in the worksheet
            self._apply_top_alignment(worksheet)
            
            # Apply filters to the header row for easy data filtering and sorting
            self._apply_filters_to_header(worksheet)
            
            # Apply alternating background colors to matched transaction blocks
            self._apply_alternating_background_colors(worksheet, file1_matched)
            
            # Format output file transaction blocks (make narration italic)
            self._format_output_file_transaction_blocks(worksheet)
            
                    
        with pd.ExcelWriter(output_file2, engine='openpyxl') as writer:
            # Write metadata
            self.metadata2.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            # Write matched transactions (skip first 8 rows to make room for metadata)
            file2_matched.to_excel(writer, sheet_name='Sheet1', index=False, header=True, startrow=8)
            
            # Get the worksheet to set column widths
            worksheet = writer.sheets['Sheet1']
            self._set_column_widths(worksheet)
            self._format_amount_columns(worksheet) # Apply amount formatting
            
            # Apply top alignment and text wrapping to ALL cells in the worksheet
            self._apply_top_alignment(worksheet)
            
            # Apply filters to the header row for easy data filtering and sorting
            self._apply_filters_to_header(worksheet)
            
            # Apply alternating background colors to matched transaction blocks
            self._apply_alternating_background_colors(worksheet, file2_matched)
            
            # Format output file transaction blocks (make narration italic)
            self._format_output_file_transaction_blocks(worksheet)
            
        
        # Also create a simple version without metadata to test (if enabled)
        if CREATE_SIMPLE_FILES:
            simple_output1 = os.path.join(OUTPUT_FOLDER, f"{base_name1}{SIMPLE_SUFFIX}")
            simple_output2 = os.path.join(OUTPUT_FOLDER, f"{base_name2}{SIMPLE_SUFFIX}")
            
            print(f"\nCreating simple test files without metadata...")
            file1_matched.to_excel(simple_output1, index=False, header=True)
            file2_matched.to_excel(simple_output2, index=False, header=True)
            
            print(f"Created simple test files:")
            print(f"  {simple_output1}")
            print(f"  {simple_output2}")
        

        
        print(f"\n=== DEBUG: AFTER SAVING ===")
        print(f"Checking if files were actually written...")
        
        # Verify the files were written correctly
        try:
            df_check1 = pd.read_excel(output_file1, header=8)
            print(f"File1 loaded successfully, shape: {df_check1.shape}")
            print(f"File1 - Rows with Match IDs: {df_check1.iloc[:, 0].notna().sum()}")
            print(f"File1 - Rows with Audit Info: {df_check1.iloc[:, 1].notna().sum()}")
            print(f"File1 - Rows with Match Type: {df_check1.iloc[:, -1].notna().sum()}")
            
            # Check if text wrapping was applied by reading the Excel file with openpyxl
            print(f"\n=== VERIFYING TEXT WRAPPING IN FILE 1 ===")
            wb1 = openpyxl.load_workbook(output_file1)
            ws1 = wb1.active
            print(f"Worksheet: {ws1.title}")
            print(f"Max row: {ws1.max_row}, Max column: {ws1.max_column}")
            
            # Check a few cells in columns B and E for text wrapping
            for row in range(9, min(15, ws1.max_row + 1)):
                cell_b = ws1.cell(row=row, column=2)
                cell_e = ws1.cell(row=row, column=5)
                print(f"Row {row}:")
                print(f"  Column B: value='{cell_b.value}', wrap_text={cell_b.alignment.wrap_text if cell_b.alignment else 'None'}")
                print(f"  Column E: value='{cell_e.value}', wrap_text={cell_e.alignment.wrap_text if cell_e.alignment else 'None'}")
            
            wb1.close()
            
        except Exception as e:
            print(f"Error reading File1: {e}")
        
        try:
            df_check2 = pd.read_excel(output_file2, header=8)
            print(f"File2 loaded successfully, shape: {df_check2.shape}")
            print(f"File2 - Rows with Match IDs: {df_check2.iloc[:, 0].notna().sum()}")
            print(f"File2 - Rows with Audit Info: {df_check2.iloc[:, 1].notna().sum()}")
            print(f"File2 - Rows with Match Type: {df_check2.iloc[:, -1].notna().sum()}")
            
            # Check if text wrapping was applied by reading the Excel file with openpyxl
            print(f"\n=== VERIFYING TEXT WRAPPING IN FILE 2 ===")
            wb2 = openpyxl.load_workbook(output_file2)
            ws2 = wb2.active
            print(f"Worksheet: {ws2.title}")
            print(f"Max row: {ws2.max_row}, Max column: {ws2.max_column}")
            
            # Check a few cells in columns B and E for text wrapping
            for row in range(9, min(15, ws2.max_row + 1)):
                cell_b = ws2.cell(row=row, column=2)
                cell_e = ws2.cell(row=row, column=5)
                print(f"Row {row}:")
                print(f"  Column B: value='{cell_b.value}', wrap_text={cell_b.alignment.wrap_text if cell_b.alignment else 'None'}")
                print(f"  Column E: value='{cell_e.value}', wrap_text={cell_e.alignment.wrap_text if cell_e.alignment else 'None'}")
            
            wb2.close()
            
        except Exception as e:
            print(f"Error reading File2: {e}")
        
        print(f"\nCreated matched files:")
        print(f"  {output_file1}")
        print(f"  {output_file2}")
        
        # Verify the data was populated correctly
        self.verify_match_data(file1_matched, file2_matched, matches)
    

    
    def verify_match_data(self, file1_matched, file2_matched, matches):
        """Verify that Match ID and Audit Info columns are properly populated."""
        print(f"\n=== VERIFICATION RESULTS ===")
        
        # Check Match ID column population - only count non-empty, non-NaN values
        match_ids_1 = file1_matched.iloc[:, 0].replace('', None).dropna()
        match_ids_2 = file2_matched.iloc[:, 0].replace('', None).dropna()
        
        print(f"File 1 - Match IDs populated: {len(match_ids_1)}")
        print(f"File 2 - Match IDs populated: {len(match_ids_2)}")
        
        # Check Audit Info column population - only count non-empty, non-NaN values
        audit_info_1 = file1_matched.iloc[:, 1].replace('', None).dropna()
        audit_info_2 = file2_matched.iloc[:, 1].replace('', None).dropna()
        
        print(f"File 1 - Audit Info populated: {len(audit_info_1)}")
        print(f"File 2 - Audit Info populated: {len(audit_info_2)}")
        
        # Check Match Type column population - only count non-empty, non-NaN values
        match_types_1 = file1_matched.iloc[:, -1].replace('', None).dropna()
        match_types_2 = file2_matched.iloc[:, -1].replace('', None).dropna()
        
        print(f"File 1 - Match Types populated: {len(match_types_1)}")
        print(f"File 2 - Match Types populated: {len(match_types_2)}")
        
        # Show sample populated data
        if len(match_ids_1) > 0:
            print(f"\nFile 1 - Sample populated rows:")
            for i, match_id in enumerate(match_ids_1[:3]):
                # Find rows that actually have this match_id (not empty strings)
                mask = (file1_matched.iloc[:, 0] == match_id) & (file1_matched.iloc[:, 0] != '')
                if mask.any():
                    row_idx = file1_matched[mask].index[0]
                    print(f"  Row {row_idx}: Match ID = {match_id}")
                    print(f"    Date: {file1_matched.iloc[row_idx, 3]}")
                    print(f"    Description: {str(file1_matched.iloc[row_idx, 4])[:50]}...")
                    print(f"    Debit: {file1_matched.iloc[row_idx, 10]}, Credit: {file1_matched.iloc[row_idx, 11]}")
                    print(f"    Match Type: {file1_matched.iloc[row_idx, -1]}")
        
        if len(match_ids_2) > 0:
            print(f"\nFile 2 - Sample populated rows:")
            for i, match_id in enumerate(match_ids_2[:3]):
                # Find rows that actually have this match_id (not empty strings)
                mask = (file2_matched.iloc[:, 0] == match_id) & (file2_matched.iloc[:, 0] != '')
                if mask.any():
                    row_idx = file2_matched[mask].index[0]
                    print(f"  Row {row_idx}: Match ID = {match_id}")
                    print(f"    Date: {file2_matched.iloc[row_idx, 3]}")
                    print(f"    Description: {str(file2_matched.iloc[row_idx, 4])[:50]}...")
                    print(f"    Debit: {file2_matched.iloc[row_idx, 10]}, Credit: {file2_matched.iloc[row_idx, 11]}")
                    print(f"    Match Type: {file2_matched.iloc[row_idx, -1]}")

def main():
    # Show current configuration
    print_configuration()
    print()
    
    # Use configuration variables from the top of the file
    print(f"=== PROCESSING FILES ===")
    
    # Check if input files exist
    if not os.path.exists(INPUT_FILE1_PATH):
        print(f"ERROR: Input file 1 not found: {INPUT_FILE1_PATH}")
        return
    if not os.path.exists(INPUT_FILE2_PATH):
        print(f"ERROR: Input file 2 not found: {INPUT_FILE2_PATH}")
        return
    
    # Create output folder if it doesn't exist
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    
    # Create matcher instance
    matcher = ExcelTransactionMatcher(INPUT_FILE1_PATH, INPUT_FILE2_PATH)
    matches = matcher.find_potential_matches()
    # Get the data from the matcher instance
    transactions1 = matcher.transactions1
    transactions2 = matcher.transactions2
    
    print(f"\n=== SUMMARY ===")
    print(f"Total potential matches found: {len(matches)}")
    
    if matches:
        print("\nCreating matched output files...")
        matcher.create_matched_files(matches, transactions1, transactions2)
        print("\nOutput files created successfully!")
    else:
        print("\nNo matches found. No output files created.")

if __name__ == "__main__":
    # Parse command line arguments
    parser = argparse.ArgumentParser(description='Excel Transaction Matcher for LC Numbers')
    parser.add_argument('--file1', help='Path to first Excel file (overrides config)')
    parser.add_argument('--file2', help='Path to second Excel file (overrides config)')
    parser.add_argument('--output', help='Output folder (overrides config)')
    parser.add_argument('--config', action='store_true', help='Show current configuration')
    parser.add_argument('--help-config', action='store_true', help='Show configuration help')
    
    args = parser.parse_args()
    
    # Handle special arguments
    if args.help_config:
        update_configuration()
        sys.exit(0)
    
    if args.config:
        print_configuration()
        sys.exit(0)
    
    # Override configuration with command line arguments if provided
    if args.file1:
        INPUT_FILE1_PATH = args.file1
    if args.file2:
        INPUT_FILE2_PATH = args.file2
    if args.output:
        OUTPUT_FOLDER = args.output
    
    # Run the main function
    main()
    