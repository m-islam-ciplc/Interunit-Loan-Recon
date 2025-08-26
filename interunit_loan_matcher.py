import pandas as pd
import re

import os
import sys
from openpyxl.styles import Alignment
import openpyxl
from openpyxl import load_workbook
from matching_logic import (
    LCMatchingLogic, POMatchingLogic, USDMatchingLogic,
    InterunitLoanMatcher, AggregatedPOMatchingLogic, NarrationMatchingLogic
)
from transaction_block_identifier import TransactionBlockIdentifier

# =============================================================================
# CONFIGURATION SECTION
# =============================================================================
# Import configuration from dedicated config module
from config import (
    OUTPUT_SUFFIX, SIMPLE_SUFFIX, CREATE_SIMPLE_FILES, VERBOSE_DEBUG
)

# Import patterns from the matching logic package
from matching_logic import LC_PATTERN, PO_PATTERN, USD_PATTERN


class   ExcelTransactionMatcher:
    """
    Handles complex Excel files with metadata rows and transaction data.
    """
    
    def __init__(self, file1_path: str, file2_path: str):
        self.file1_path = file1_path
        self.file2_path = file2_path
        self.metadata1 = None
        self.transactions1 = None
        self.metadata2 = None
        self.transactions2 = None
        self.block_identifier = TransactionBlockIdentifier()
        self.lc_matching_logic = LCMatchingLogic(self.block_identifier)
        self.po_matching_logic = POMatchingLogic(self.block_identifier)
        self.usd_matching_logic = USDMatchingLogic(self.block_identifier)
        self.interunit_loan_matcher = InterunitLoanMatcher(self.block_identifier)
        self.aggregated_po_matching_logic = AggregatedPOMatchingLogic(self.block_identifier)
        self.narration_matching_logic = NarrationMatchingLogic(self.block_identifier)
        
        # Performance optimization caches
        self._block_header_cache1 = {}
        self._block_header_cache2 = {}
        self._amount_cache1 = {}
        self._amount_cache2 = {}
        
        # Cached workbook data for performance
        self._cached_wb1 = None
        self._cached_ws1 = None
        self._cached_wb2 = None
        self._cached_ws2 = None
        self._cached_blocks1 = None
        self._cached_blocks2 = None
        self._cached_formatting_data1 = None
        self._cached_formatting_data2 = None
        
        # Compiled regex patterns for performance
        self._compiled_lc_pattern = re.compile(LC_PATTERN)
        self._compiled_po_pattern = re.compile(PO_PATTERN)
        self._compiled_usd_pattern = re.compile(USD_PATTERN)
        self._compiled_interunit_pattern = re.compile(r'([A-Z]{2,4})#(\d{4,6})')
        
        # Cached extracted data
        self._cached_extracted_data = None
        
        # Optimized data access caches
        self._lc_numbers1_array = None
        self._lc_numbers2_array = None
        self._po_numbers1_array = None
        self._po_numbers2_array = None
        self._usd_amounts1_array = None
        self._usd_amounts2_array = None
        self._interunit_accounts1_array = None
        self._interunit_accounts2_array = None
        
        # Additional performance caches
        self._block_header_cache = {}  # Universal block header cache
        self._description_row_cache = {}  # Universal description row cache
        self._narration_cache = {}  # Cached narration strings
        self._amount_cache = {}  # Universal amount cache
        
        # ULTRA-OPTIMIZED: Matching logic registry for future scalability
        self._matching_logic_registry = self._initialize_matching_logic_registry()
        
        # ULTRA-OPTIMIZED: Pre-computed data caches
        self._precomputed_data = {}
        self._block_pair_cache = {}
        
        # ULTRA-OPTIMIZED: Performance monitoring
        self._performance_stats = {
            'block_pairs_analyzed': 0,
            'matches_found': 0,
            'cache_hits': 0,
            'cache_misses': 0,
            'processing_time': 0
        }
        
        # ULTRA-OPTIMIZED: Memory management
        self._max_cache_size = 10000  # Maximum cache entries
        self._cache_cleanup_threshold = 0.8  # Cleanup when 80% full
        
    def read_complex_excel(self, file_path: str):
        """Read Excel file with metadata + transaction structure."""
        # Read everything as strings to preserve all formatting
        full_df = pd.read_excel(file_path, header=None, dtype=str)

        # Extract metadata (rows 0-7, which are Excel rows 1-8)
        metadata = full_df.iloc[0:8, :]

        # Extract transaction data (rows 8+, which are Excel rows 9+)
        transactions = full_df.iloc[8:, :]

        # Set first row as headers and remove it from data
        transactions.columns = transactions.iloc[0]
        transactions = transactions.iloc[1:].reset_index(drop=True)

        return metadata, transactions
    
    def extract_amounts_from_strings(self, row):
        """Extract amounts from row data that's loaded as strings - OPTIMIZED."""
        # Get raw values once
        debit_raw = row.iloc[7] if pd.notna(row.iloc[7]) else '0'
        credit_raw = row.iloc[8] if pd.notna(row.iloc[8]) else '0'
        
        # Convert to strings once
        debit_str = str(debit_raw)
        credit_str = str(credit_raw)
        
        # Optimized numeric conversion
        try:
            # Remove commas and check if numeric in one pass
            debit_clean = debit_str.replace(',', '')
            credit_clean = credit_str.replace(',', '')
            
            # Use faster numeric check
            debit = float(debit_clean) if debit_clean.replace('.', '').isdigit() else 0.0
            credit = float(credit_clean) if credit_clean.replace('.', '').isdigit() else 0.0
        except (ValueError, TypeError):
            debit, credit = 0.0, 0.0
        
        return debit, credit
    
    def load_and_cache_workbooks(self):
        """Load workbooks once and cache them for performance."""
        if self._cached_wb1 is None:
            print("Loading and caching workbooks for performance...")
            self._cached_wb1 = load_workbook(self.file1_path)
            self._cached_ws1 = self._cached_wb1.active
            self._cached_wb2 = load_workbook(self.file2_path)
            self._cached_ws2 = self._cached_wb2.active
            print("Workbooks cached successfully")
    
    def get_cached_blocks(self, file_num):
        """Get cached transaction blocks or identify them if not cached."""
        if file_num == 1:
            if self._cached_blocks1 is None:
                print("Identifying transaction blocks for File 1...")
                self._cached_blocks1 = self.block_identifier.identify_transaction_blocks(
                    self.transactions1, self.file1_path
                )
                print(f"File 1: {len(self._cached_blocks1)} transaction blocks cached")
            return self._cached_blocks1
        else:
            if self._cached_blocks2 is None:
                print("Identifying transaction blocks for File 2...")
                self._cached_blocks2 = self.block_identifier.identify_transaction_blocks(
                    self.transactions2, self.file2_path
                )
                print(f"File 2: {len(self._cached_blocks2)} transaction blocks cached")
            return self._cached_blocks2
    
    def get_cached_formatting_data(self, file_num):
        """Get cached formatting data or analyze it if not cached."""
        if file_num == 1:
            if self._cached_formatting_data1 is None:
                print("Analyzing formatting data for File 1...")
                self._cached_formatting_data1 = self._analyze_formatting_data(
                    self._cached_ws1, self._cached_blocks1
                )
                print(f"File 1: {len(self._cached_formatting_data1)} blocks with formatting data cached")
            return self._cached_formatting_data1
        else:
            if self._cached_formatting_data2 is None:
                print("Analyzing formatting data for File 2...")
                self._cached_formatting_data2 = self._analyze_formatting_data(
                    self._cached_ws2, self._cached_blocks2
                )
                print(f"File 2: {len(self._cached_formatting_data2)} blocks with formatting data cached")
            return self._cached_formatting_data2
    
    def _analyze_formatting_data(self, worksheet, blocks):
        """Analyze formatting data for all blocks at once."""
        formatting_data = []
        
        for i, block in enumerate(blocks):
            block_data = {
                'block_index': i,
                'block_rows': block,
                'ledger_accounts': [],
                'narration_short_codes': [],
                'amounts': {}
            }
            
            # Check each row in the block
            for row_idx in block:
                excel_row = row_idx + 10  # Convert to Excel row number
                
                if excel_row <= worksheet.max_row:
                    cell_c = worksheet.cell(row=excel_row, column=3)  # Column C
                    
                    # Check for ledger accounts (Bold but not italic)
                    if (cell_c.value and 
                        cell_c.font and 
                        cell_c.font.bold and 
                        not cell_c.font.italic):
                        
                        # Check if this is an interunit account
                        for full_account, short_codes in self.interunit_loan_matcher.interunit_account_mapping.items():
                            if full_account.upper() in str(cell_c.value).upper():
                                # Handle both list and single string formats for backward compatibility
                                codes_list = short_codes if isinstance(short_codes, list) else [short_codes]
                                # Add all short codes for this account
                                for short_code in codes_list:
                                    block_data['ledger_accounts'].append({
                                        'full_account': full_account,
                                        'short_code': short_code,
                                        'cell_value': cell_c.value
                                    })
                    
                    # Check for narration rows (Italic but not bold)
                    elif (cell_c.value and 
                          cell_c.font and 
                          not cell_c.font.bold and 
                          cell_c.font.italic):
                        
                        # Look for short codes in narration
                        for short_codes in self.interunit_loan_matcher.interunit_account_mapping.values():
                            # Handle both list and single string formats for backward compatibility
                            codes_list = short_codes if isinstance(short_codes, list) else [short_codes]
                            for short_code in codes_list:
                                if short_code in str(cell_c.value):
                                    block_data['narration_short_codes'].append({
                                        'short_code': short_code,
                                        'narration': cell_c.value
                                    })
                    
                    # Check for amounts (Debit/Credit columns)
                    debit_cell = worksheet.cell(row=excel_row, column=8)  # Column H
                    credit_cell = worksheet.cell(row=excel_row, column=9)  # Column I
                    
                    if (debit_cell.value is not None and debit_cell.value != 0) or \
                       (credit_cell.value is not None and credit_cell.value != 0):
                        block_data['amounts'] = {
                            'debit': debit_cell.value if debit_cell.value else None,
                            'credit': credit_cell.value if credit_cell.value else None,
                            'row': row_idx
                        }
            
            if block_data['ledger_accounts'] or block_data['narration_short_codes']:
                formatting_data.append(block_data)
        
        return formatting_data
    
    def close_cached_workbooks(self):
        """Close cached workbooks to free memory."""
        if self._cached_wb1:
            self._cached_wb1.close()
            self._cached_wb1 = None
            self._cached_ws1 = None
        if self._cached_wb2:
            self._cached_wb2.close()
            self._cached_wb2 = None
            self._cached_ws2 = None
    
    def clear_all_caches(self):
        """Clear all caches to free memory."""
        # Clear workbook caches
        self.close_cached_workbooks()
        
        # Clear other caches
        self._cached_blocks1 = None
        self._cached_blocks2 = None
        self._cached_formatting_data1 = None
        self._cached_formatting_data2 = None
        self._cached_extracted_data = None
        
        # Clear amount caches
        self._amount_cache1.clear()
        self._amount_cache2.clear()
        self._block_header_cache1.clear()
        self._block_header_cache2.clear()
        
        # Clear additional performance caches
        self._block_header_cache.clear()
        self._description_row_cache.clear()
        self._narration_cache.clear()
        self._amount_cache.clear()
        
        print("All caches cleared for memory management")
    
    
    def _apply_batch_formatting(self, worksheet, num_rows):
        """Apply formatting in batch for better performance."""
        # Set column widths
        column_widths = {
            'A': 9.00, 'B': 30.00, 'C': 12.00, 'D': 10.33, 'E': 60.00,
            'F': 5.00, 'G': 5.00, 'H': 12.78, 'I': 9.00, 'J': 13.78, 'K': 14.22
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Format amount columns in batch
        for row in range(2, num_rows + 2):  # Skip header row
            for col in ['H', 'I', 'J', 'K']:  # Amount columns
                cell = worksheet[f"{col}{row}"]
                if cell.value is not None:
                    cell.number_format = '#,##0.00'
    
    def create_optimized_arrays(self, lc_numbers1, lc_numbers2, po_numbers1, po_numbers2, 
                               usd_amounts1, usd_amounts2, interunit_accounts1, interunit_accounts2):
        """Create optimized numpy arrays for faster access in matching loops."""
        print("Creating optimized arrays for faster data access...")
        
        # Convert Series to numpy arrays for faster access
        self._lc_numbers1_array = lc_numbers1.values
        self._lc_numbers2_array = lc_numbers2.values
        self._po_numbers1_array = po_numbers1.values
        self._po_numbers2_array = po_numbers2.values
        self._usd_amounts1_array = usd_amounts1.values
        self._usd_amounts2_array = usd_amounts2.values
        self._interunit_accounts1_array = interunit_accounts1.values
        self._interunit_accounts2_array = interunit_accounts2.values
        
        print("Optimized arrays created successfully")
    
    def create_unmatched_indices_optimized(self, *match_lists):
        """Create unmatched indices efficiently without list concatenation."""
        matched_indices1 = set()
        matched_indices2 = set()
        
        # Process all match lists in one pass
        for match_list in match_lists:
            for match in match_list:
                if 'File1_Index' in match:
                    matched_indices1.add(match['File1_Index'])
                if 'File2_Index' in match:
                    matched_indices2.add(match['File2_Index'])
        
        return matched_indices1, matched_indices2
    
    def get_cached_block_header_universal(self, row_idx, transactions_df):
        """Get block header with universal caching for maximum performance."""
        cache_key = f"{id(transactions_df)}_{row_idx}"
        
        if cache_key not in self._block_header_cache:
            self._block_header_cache[cache_key] = self.block_identifier.find_transaction_block_header(row_idx, transactions_df)
        
        return self._block_header_cache[cache_key]
    
    def get_cached_description_row_universal(self, row_idx, transactions_df):
        """Get description row with universal caching for maximum performance."""
        cache_key = f"{id(transactions_df)}_{row_idx}"
        
        if cache_key not in self._description_row_cache:
            self._description_row_cache[cache_key] = self.block_identifier.find_description_row_in_block(row_idx, transactions_df)
        
        return self._description_row_cache[cache_key]
    
    def get_cached_narration(self, row_idx, transactions_df):
        """Get narration with caching for maximum performance."""
        cache_key = f"{id(transactions_df)}_{row_idx}"
        
        if cache_key not in self._narration_cache:
            # Check bounds before accessing
            if row_idx < len(transactions_df):
                narration = str(transactions_df.iloc[row_idx, 2]).strip()
            else:
                narration = ""
            self._narration_cache[cache_key] = narration
        
        return self._narration_cache[cache_key]
    
    def get_cached_amounts_universal(self, row_idx, transactions_df):
        """Get amounts with universal caching for maximum performance."""
        cache_key = f"{id(transactions_df)}_{row_idx}"
        
        if cache_key not in self._amount_cache:
            # Check bounds before accessing
            if row_idx < len(transactions_df):
                row = transactions_df.iloc[row_idx]
                self._amount_cache[cache_key] = self.extract_amounts_from_strings(row)
            else:
                self._amount_cache[cache_key] = (0.0, 0.0)
        
        return self._amount_cache[cache_key]
    
    def find_matches_block_based_optimized(self, transactions1, transactions2, blocks1, blocks2, 
                                         lc_numbers1, lc_numbers2, po_numbers1, po_numbers2, 
                                         interunit_accounts1, interunit_accounts2, usd_amounts1, usd_amounts2):
        """ULTRA-OPTIMIZED: Pre-filter by amount matching, then run matching logic only on filtered records."""
        
        print("\n" + "="*70)
        print("🚀 ULTRA-OPTIMIZED: PRE-FILTER BY AMOUNT MATCHING")
        print("="*70)
        print("Step 1: Find transaction blocks where lender debit = borrower credit")
        print("Step 2: Save those records in filtered DataFrames")
        print("Step 3: Run matching logic only on filtered records")
        print("This eliminates unnecessary processing of non-matching amounts.")
        
        print(f"\n📊 INITIAL DATA:")
        print(f"File 1: {len(transactions1)} transactions, {len(blocks1)} blocks")
        print(f"File 2: {len(transactions2)} transactions, {len(blocks2)} blocks")
        
        # STEP 1: Pre-filter by amount matching
        print("\n🔍 STEP 1: PRE-FILTERING BY AMOUNT MATCHING...")
        filtered_records1, filtered_records2 = self._prefilter_by_amount_matching(
            transactions1, transactions2, blocks1, blocks2
        )
        
        print(f"✅ PRE-FILTERING COMPLETE:")
        print(f"  - Filtered File 1: {len(filtered_records1)} records")
        print(f"  - Filtered File 2: {len(filtered_records2)} records")
        print(f"  - Reduction: {len(transactions1) - len(filtered_records1)} records eliminated from File 1")
        print(f"  - Reduction: {len(transactions2) - len(filtered_records2)} records eliminated from File 2")
        
        # STEP 2: Extract data only from filtered records
        print("\n🔍 STEP 2: EXTRACTING DATA FROM FILTERED RECORDS...")
        filtered_lc1, filtered_lc2, filtered_po1, filtered_po2, filtered_inter1, filtered_inter2, filtered_usd1, filtered_usd2 = self._extract_data_from_filtered_records(
            filtered_records1, filtered_records2, lc_numbers1, lc_numbers2, po_numbers1, po_numbers2,
            interunit_accounts1, interunit_accounts2, usd_amounts1, usd_amounts2
        )
        
        # STEP 3: Run matching logic only on filtered records
        print("\n🔍 STEP 3: RUNNING MATCHING LOGIC ON FILTERED RECORDS...")
        all_matches = self._run_matching_on_filtered_records(
            filtered_records1, filtered_records2, filtered_lc1, filtered_lc2, 
            filtered_po1, filtered_po2, filtered_inter1, filtered_inter2, 
            filtered_usd1, filtered_usd2
        )
        
        print(f"\n📈 ULTRA-OPTIMIZED MATCHING RESULTS:")
        print(f"Total matches found: {len(all_matches)}")
        print(f"Processing efficiency: {len(all_matches)} matches from {len(filtered_records1) + len(filtered_records2)} filtered records")
        
        return all_matches
    
    def _prefilter_by_amount_matching(self, transactions1, transactions2, blocks1, blocks2):
        """Pre-filter records where lender debit amount = borrower credit amount."""
        print("  - Creating amount-to-record mapping...")
        
        # Create amount-to-record mapping for both files
        amount_to_records1 = {}  # {amount: [(row_idx, role, block_idx)]}
        amount_to_records2 = {}
        
        # Map File 1 records by amount
        for block_idx, block in enumerate(blocks1):
            for row_idx in block:
                if row_idx < len(transactions1):
                    amounts = self.get_cached_amounts_universal(row_idx, transactions1)
                    if amounts[0] > 0:  # Debit amount (lender)
                        if amounts[0] not in amount_to_records1:
                            amount_to_records1[amounts[0]] = []
                        amount_to_records1[amounts[0]].append((row_idx, 'lender', block_idx))
                    if amounts[1] > 0:  # Credit amount (borrower)
                        if amounts[1] not in amount_to_records1:
                            amount_to_records1[amounts[1]] = []
                        amount_to_records1[amounts[1]].append((row_idx, 'borrower', block_idx))
        
        # Map File 2 records by amount
        for block_idx, block in enumerate(blocks2):
            for row_idx in block:
                if row_idx < len(transactions2):
                    amounts = self.get_cached_amounts_universal(row_idx, transactions2)
                    if amounts[0] > 0:  # Debit amount (lender)
                        if amounts[0] not in amount_to_records2:
                            amount_to_records2[amounts[0]] = []
                        amount_to_records2[amounts[0]].append((row_idx, 'lender', block_idx))
                    if amounts[1] > 0:  # Credit amount (borrower)
                        if amounts[1] not in amount_to_records2:
                            amount_to_records2[amounts[1]] = []
                        amount_to_records2[amounts[1]].append((row_idx, 'borrower', block_idx))
        
        print(f"  - File 1: {len(amount_to_records1)} unique amounts")
        print(f"  - File 2: {len(amount_to_records2)} unique amounts")
        
        # Find common amounts
        common_amounts = set(amount_to_records1.keys()) & set(amount_to_records2.keys())
        print(f"  - Common amounts: {len(common_amounts)}")
        
        # Collect records that have matching amounts
        filtered_indices1 = set()
        filtered_indices2 = set()
        
        for amount in common_amounts:
            file1_records = amount_to_records1[amount]
            file2_records = amount_to_records2[amount]
            
            # Check for lender-borrower pairs
            for row1_idx, role1, block1_idx in file1_records:
                for row2_idx, role2, block2_idx in file2_records:
                    # Only process lender-borrower pairs (not lender-lender or borrower-borrower)
                    if role1 != role2:
                        filtered_indices1.add(row1_idx)
                        filtered_indices2.add(row2_idx)
        
        print(f"  - Filtered indices File 1: {len(filtered_indices1)}")
        print(f"  - Filtered indices File 2: {len(filtered_indices2)}")
        
        # Create filtered DataFrames
        filtered_records1 = transactions1.iloc[list(filtered_indices1)].copy() if filtered_indices1 else transactions1.iloc[0:0].copy()
        filtered_records2 = transactions2.iloc[list(filtered_indices2)].copy() if filtered_indices2 else transactions2.iloc[0:0].copy()
        
        return filtered_records1, filtered_records2
    
    def _extract_data_from_filtered_records(self, filtered_records1, filtered_records2, 
                                          lc_numbers1, lc_numbers2, po_numbers1, po_numbers2,
                                          interunit_accounts1, interunit_accounts2, usd_amounts1, usd_amounts2):
        """Extract matching data only from filtered records."""
        print("  - Extracting LC numbers from filtered records...")
        filtered_lc1 = self._extract_filtered_series(lc_numbers1, filtered_records1, lc_numbers1)
        filtered_lc2 = self._extract_filtered_series(lc_numbers2, filtered_records2, lc_numbers2)
        
        print("  - Extracting PO numbers from filtered records...")
        filtered_po1 = self._extract_filtered_series(po_numbers1, filtered_records1, po_numbers1)
        filtered_po2 = self._extract_filtered_series(po_numbers2, filtered_records2, po_numbers2)
        
        print("  - Extracting Interunit accounts from filtered records...")
        filtered_inter1 = self._extract_filtered_series(interunit_accounts1, filtered_records1, interunit_accounts1)
        filtered_inter2 = self._extract_filtered_series(interunit_accounts2, filtered_records2, interunit_accounts2)
        
        print("  - Extracting USD amounts from filtered records...")
        filtered_usd1 = self._extract_filtered_series(usd_amounts1, filtered_records1, usd_amounts1)
        filtered_usd2 = self._extract_filtered_series(usd_amounts2, filtered_records2, usd_amounts2)
        
        return filtered_lc1, filtered_lc2, filtered_po1, filtered_po2, filtered_inter1, filtered_inter2, filtered_usd1, filtered_usd2
    
    def _extract_filtered_series(self, original_series, filtered_df, reference_series):
        """Extract series data for filtered records only."""
        if len(filtered_df) == 0:
            return pd.Series(dtype=object)
        
        # Get the indices of filtered records in the original series
        filtered_indices = filtered_df.index
        return original_series.iloc[filtered_indices].copy()
    
    def _run_matching_on_filtered_records(self, filtered_records1, filtered_records2,
                                        filtered_lc1, filtered_lc2, filtered_po1, filtered_po2,
                                        filtered_inter1, filtered_inter2, filtered_usd1, filtered_usd2):
        """Run matching logic only on pre-filtered records."""
        print("  - Running matching logic on filtered records...")
        
        all_matches = []
        
        # Create amount-to-record mapping for filtered records
        amount_to_records1 = {}
        amount_to_records2 = {}
        
        # Map filtered File 1 records by amount
        for idx, row in filtered_records1.iterrows():
            amounts = self.get_cached_amounts_universal(idx, filtered_records1)
            if amounts[0] > 0:  # Debit amount (lender)
                if amounts[0] not in amount_to_records1:
                    amount_to_records1[amounts[0]] = []
                amount_to_records1[amounts[0]].append((idx, 'lender'))
            if amounts[1] > 0:  # Credit amount (borrower)
                if amounts[1] not in amount_to_records1:
                    amount_to_records1[amounts[1]] = []
                amount_to_records1[amounts[1]].append((idx, 'borrower'))
        
        # Map filtered File 2 records by amount
        for idx, row in filtered_records2.iterrows():
            amounts = self.get_cached_amounts_universal(idx, filtered_records2)
            if amounts[0] > 0:  # Debit amount (lender)
                if amounts[0] not in amount_to_records2:
                    amount_to_records2[amounts[0]] = []
                amount_to_records2[amounts[0]].append((idx, 'lender'))
            if amounts[1] > 0:  # Credit amount (borrower)
                if amounts[1] not in amount_to_records2:
                    amount_to_records2[amounts[1]] = []
                amount_to_records2[amounts[1]].append((idx, 'borrower'))
        
        # Find common amounts in filtered records
        common_amounts = set(amount_to_records1.keys()) & set(amount_to_records2.keys())
        print(f"  - Common amounts in filtered records: {len(common_amounts)}")
        
        # Process each common amount
        for amount in common_amounts:
            file1_records = amount_to_records1[amount]
            file2_records = amount_to_records2[amount]
            
            # Check all possible record combinations for this amount
            for row1_idx, role1 in file1_records:
                for row2_idx, role2 in file2_records:
                    # Skip if both are lenders or both are borrowers
                    if role1 == role2:
                        continue
                    
                    # Determine lender and borrower
                    if role1 == 'lender':
                        lender_row, borrower_row = row1_idx, row2_idx
                        lender_file, borrower_file = 1, 2
                        lender_records, borrower_records = filtered_records1, filtered_records2
                    else:
                        lender_row, borrower_row = row2_idx, row1_idx
                        lender_file, borrower_file = 2, 1
                        lender_records, borrower_records = filtered_records2, filtered_records1
                    
                    # Analyze this record pair and determine best match type
                    match_type, match_data = self._analyze_filtered_record_pair(
                        lender_records, borrower_records, lender_row, borrower_row,
                        lender_file, borrower_file, filtered_lc1, filtered_lc2,
                        filtered_po1, filtered_po2, filtered_inter1, filtered_inter2,
                        filtered_usd1, filtered_usd2, amount
                    )
                    
                    if match_type:
                        # Create match record
                        match = {
                            'match_id': None,
                            'Match_Type': match_type,
                            'File1_Index': lender_row if lender_file == 1 else borrower_row,
                            'File2_Index': borrower_row if lender_file == 1 else lender_row,
                            'Amount': amount,
                            **match_data
                        }
                        all_matches.append(match)
                        print(f"    ✅ {match_type} match found: Amount {amount}")
                        break  # Move to next amount after first match
        
        return all_matches
    
    def _analyze_filtered_record_pair(self, lender_records, borrower_records, lender_row, borrower_row,
                                    lender_file, borrower_file, filtered_lc1, filtered_lc2,
                                    filtered_po1, filtered_po2, filtered_inter1, filtered_inter2,
                                    filtered_usd1, filtered_usd2, amount):
        """Analyze a filtered record pair and determine the best match type."""
        
        # Pre-compute all data for this record pair
        block_pair_key = f"{lender_file}_{lender_row}_{borrower_file}_{borrower_row}"
        
        if block_pair_key not in self._block_pair_cache:
            self._performance_stats['cache_misses'] += 1
            self._block_pair_cache[block_pair_key] = self._precompute_filtered_record_data(
                lender_records, borrower_records, lender_row, borrower_row,
                lender_file, borrower_file, filtered_lc1, filtered_lc2,
                filtered_po1, filtered_po2, filtered_inter1, filtered_inter2,
                filtered_usd1, filtered_usd2
            )
            self._cleanup_caches_if_needed()
        else:
            self._performance_stats['cache_hits'] += 1
        
        block_data = self._block_pair_cache[block_pair_key]
        
        # Process matching logics in priority order
        for logic_name, logic_config in sorted(self._matching_logic_registry.items(), 
                                             key=lambda x: x[1]['priority']):
            if not logic_config['enabled']:
                continue
                
            try:
                match_result = logic_config['function'](block_data, amount)
                if match_result:
                    self._performance_stats['matches_found'] += 1
                    return logic_name, match_result
            except Exception as e:
                print(f"⚠️ Warning: {logic_name} matching failed: {e}")
                continue
        
        return None, {}
    
    def _precompute_filtered_record_data(self, lender_records, borrower_records, lender_row, borrower_row,
                                       lender_file, borrower_file, filtered_lc1, filtered_lc2,
                                       filtered_po1, filtered_po2, filtered_inter1, filtered_inter2,
                                       filtered_usd1, filtered_usd2):
        """Pre-compute all data for a filtered record pair."""
        return {
            'lender_narration': self.get_cached_narration(lender_row, lender_records),
            'borrower_narration': self.get_cached_narration(borrower_row, borrower_records),
            'lender_lc': self._get_filtered_lc_number(lender_file, lender_row, filtered_lc1, filtered_lc2),
            'borrower_lc': self._get_filtered_lc_number(borrower_file, borrower_row, filtered_lc1, filtered_lc2),
            'lender_po': self._get_filtered_po_number(lender_file, lender_row, filtered_po1, filtered_po2),
            'borrower_po': self._get_filtered_po_number(borrower_file, borrower_row, filtered_po1, filtered_po2),
            'lender_interunit': self._get_filtered_interunit(lender_file, lender_row, filtered_inter1, filtered_inter2),
            'borrower_interunit': self._get_filtered_interunit(borrower_file, borrower_row, filtered_inter1, filtered_inter2),
            'lender_usd': self._get_filtered_usd(lender_file, lender_row, filtered_usd1, filtered_usd2),
            'borrower_usd': self._get_filtered_usd(borrower_file, borrower_row, filtered_usd1, filtered_usd2),
            'lender_file': lender_file,
            'borrower_file': borrower_file,
            'lender_row': lender_row,
            'borrower_row': borrower_row
        }
    
    def _get_filtered_lc_number(self, file_num, row_idx, filtered_lc1, filtered_lc2):
        """Get LC number from filtered series."""
        if file_num == 1 and row_idx in filtered_lc1.index:
            return filtered_lc1.loc[row_idx] if pd.notna(filtered_lc1.loc[row_idx]) else None
        elif file_num == 2 and row_idx in filtered_lc2.index:
            return filtered_lc2.loc[row_idx] if pd.notna(filtered_lc2.loc[row_idx]) else None
        return None
    
    def _get_filtered_po_number(self, file_num, row_idx, filtered_po1, filtered_po2):
        """Get PO number from filtered series."""
        if file_num == 1 and row_idx in filtered_po1.index:
            return filtered_po1.loc[row_idx] if pd.notna(filtered_po1.loc[row_idx]) else None
        elif file_num == 2 and row_idx in filtered_po2.index:
            return filtered_po2.loc[row_idx] if pd.notna(filtered_po2.loc[row_idx]) else None
        return None
    
    def _get_filtered_interunit(self, file_num, row_idx, filtered_inter1, filtered_inter2):
        """Get Interunit account from filtered series."""
        if file_num == 1 and row_idx in filtered_inter1.index:
            return filtered_inter1.loc[row_idx] if pd.notna(filtered_inter1.loc[row_idx]) else None
        elif file_num == 2 and row_idx in filtered_inter2.index:
            return filtered_inter2.loc[row_idx] if pd.notna(filtered_inter2.loc[row_idx]) else None
        return None
    
    def _get_filtered_usd(self, file_num, row_idx, filtered_usd1, filtered_usd2):
        """Get USD amount from filtered series."""
        if file_num == 1 and row_idx in filtered_usd1.index:
            return filtered_usd1.loc[row_idx] if pd.notna(filtered_usd1.loc[row_idx]) else None
        elif file_num == 2 and row_idx in filtered_usd2.index:
            return filtered_usd2.loc[row_idx] if pd.notna(filtered_usd2.loc[row_idx]) else None
        return None
    
    def add_matching_logic(self, name, priority, function, enabled=True, description=""):
        """ULTRA-OPTIMIZED: Add new matching logic dynamically for future scalability."""
        self._matching_logic_registry[name] = {
            'priority': priority,
            'function': function,
            'enabled': enabled,
            'description': description
        }
        print(f"✅ Added matching logic: {name} (Priority: {priority}, Enabled: {enabled})")
    
    def enable_matching_logic(self, name):
        """ULTRA-OPTIMIZED: Enable a matching logic."""
        if name in self._matching_logic_registry:
            self._matching_logic_registry[name]['enabled'] = True
            print(f"✅ Enabled matching logic: {name}")
        else:
            print(f"❌ Matching logic not found: {name}")
    
    def disable_matching_logic(self, name):
        """ULTRA-OPTIMIZED: Disable a matching logic."""
        if name in self._matching_logic_registry:
            self._matching_logic_registry[name]['enabled'] = False
            print(f"✅ Disabled matching logic: {name}")
        else:
            print(f"❌ Matching logic not found: {name}")
    
    def get_matching_logic_status(self):
        """ULTRA-OPTIMIZED: Get status of all matching logics."""
        status = {}
        for name, config in self._matching_logic_registry.items():
            status[name] = {
                'enabled': config['enabled'],
                'priority': config['priority'],
                'description': config['description']
            }
        return status
    
    def _cleanup_caches_if_needed(self):
        """ULTRA-OPTIMIZED: Clean up caches when they get too large."""
        total_cache_size = (len(self._block_pair_cache) + 
                           len(self._block_header_cache) + 
                           len(self._narration_cache) + 
                           len(self._amount_cache))
        
        if total_cache_size > self._max_cache_size * self._cache_cleanup_threshold:
            print(f"🧹 Cleaning up caches (size: {total_cache_size})...")
            
            # Keep only the most recent 50% of entries
            keep_size = self._max_cache_size // 2
            
            # Clean block pair cache
            if len(self._block_pair_cache) > keep_size:
                items = list(self._block_pair_cache.items())
                self._block_pair_cache = dict(items[-keep_size:])
            
            # Clean other caches
            if len(self._block_header_cache) > keep_size:
                items = list(self._block_header_cache.items())
                self._block_header_cache = dict(items[-keep_size:])
            
            if len(self._narration_cache) > keep_size:
                items = list(self._narration_cache.items())
                self._narration_cache = dict(items[-keep_size:])
            
            if len(self._amount_cache) > keep_size:
                items = list(self._amount_cache.items())
                self._amount_cache = dict(items[-keep_size:])
            
            print(f"✅ Cache cleanup completed (new size: {len(self._block_pair_cache) + len(self._block_header_cache) + len(self._narration_cache) + len(self._amount_cache)})")
    
    def get_performance_stats(self):
        """ULTRA-OPTIMIZED: Get performance statistics."""
        return self._performance_stats.copy()
    
    def reset_performance_stats(self):
        """ULTRA-OPTIMIZED: Reset performance statistics."""
        self._performance_stats = {
            'block_pairs_analyzed': 0,
            'matches_found': 0,
            'cache_hits': 0,
            'cache_misses': 0,
            'processing_time': 0
        }
        print("✅ Performance statistics reset")
    
    def optimize_for_large_files(self):
        """ULTRA-OPTIMIZED: Apply additional optimizations for large files."""
        print("🚀 Applying large file optimizations...")
        
        # Increase cache sizes for large files
        self._max_cache_size = 50000
        
        # Enable more aggressive pre-filtering
        print("✅ Large file optimizations applied")
    
    def optimize_for_small_files(self):
        """ULTRA-OPTIMIZED: Apply optimizations for small files."""
        print("🚀 Applying small file optimizations...")
        
        # Reduce cache sizes for small files
        self._max_cache_size = 1000
        
        # Disable some heavy optimizations
        print("✅ Small file optimizations applied")
        
    def _initialize_matching_logic_registry(self):
        """Initialize the matching logic registry for maximum performance and future scalability."""
        return {
            'Narration': {
                'priority': 1,
                'function': self._match_narration_ultra_optimized,
                'enabled': True,
                'description': 'Exact narration matching (highest priority)'
            },
            'LC': {
                'priority': 2,
                'function': self._match_lc_ultra_optimized,
                'enabled': True,
                'description': 'Letter of Credit number matching'
            },
            'PO': {
                'priority': 3,
                'function': self._match_po_ultra_optimized,
                'enabled': True,
                'description': 'Purchase Order number matching'
            },
            'Interunit': {
                'priority': 4,
                'function': self._match_interunit_ultra_optimized,
                'enabled': True,
                'description': 'Interunit account matching'
            },
            'USD': {
                'priority': 5,
                'function': self._match_usd_ultra_optimized,
                'enabled': True,
                'description': 'USD amount matching'
            },
        }
    
    def analyze_block_pair(self, transactions1, transactions2, blocks1, blocks2,
                          lender_row, borrower_row, lender_file, borrower_file,
                          lc_numbers1, lc_numbers2, po_numbers1, po_numbers2,
                          interunit_accounts1, interunit_accounts2, usd_amounts1, usd_amounts2,
                          amount):
        """ULTRA-OPTIMIZED: Analyze a block pair using the matching logic registry."""
        
        # ULTRA-OPTIMIZED: Performance monitoring
        self._performance_stats['block_pairs_analyzed'] += 1
        
        # ULTRA-OPTIMIZED: Pre-compute all data once for this block pair
        block_pair_key = f"{lender_file}_{lender_row}_{borrower_file}_{borrower_row}"
        
        if block_pair_key not in self._block_pair_cache:
            self._performance_stats['cache_misses'] += 1
            self._block_pair_cache[block_pair_key] = self._precompute_block_pair_data(
                transactions1, transactions2, lender_row, borrower_row, 
                lender_file, borrower_file, lc_numbers1, lc_numbers2, 
                po_numbers1, po_numbers2, interunit_accounts1, interunit_accounts2, 
                usd_amounts1, usd_amounts2
            )
            
            # ULTRA-OPTIMIZED: Memory management
            self._cleanup_caches_if_needed()
        else:
            self._performance_stats['cache_hits'] += 1
        
        block_data = self._block_pair_cache[block_pair_key]
        
        # ULTRA-OPTIMIZED: Process matching logics in priority order
        for logic_name, logic_config in sorted(self._matching_logic_registry.items(), 
                                             key=lambda x: x[1]['priority']):
            if not logic_config['enabled']:
                continue
                
            try:
                match_result = logic_config['function'](block_data, amount)
                if match_result:
                    self._performance_stats['matches_found'] += 1
                    return logic_name, match_result
            except Exception as e:
                print(f"⚠️ Warning: {logic_name} matching failed: {e}")
                continue
        
        # No match found
        return None, {}
    
    def _precompute_block_pair_data(self, transactions1, transactions2, lender_row, borrower_row,
                                   lender_file, borrower_file, lc_numbers1, lc_numbers2,
                                   po_numbers1, po_numbers2, interunit_accounts1, interunit_accounts2,
                                   usd_amounts1, usd_amounts2):
        """ULTRA-OPTIMIZED: Pre-compute all data for a block pair to avoid repeated calculations."""
        return {
            'lender_narration': self.get_cached_narration(lender_row, transactions1 if lender_file == 1 else transactions2),
            'borrower_narration': self.get_cached_narration(borrower_row, transactions2 if borrower_file == 1 else transactions1),
            'lender_lc': self.get_optimized_lc_numbers(lender_file, lender_row),
            'borrower_lc': self.get_optimized_lc_numbers(borrower_file, borrower_row),
            'lender_po': self.get_optimized_po_numbers(lender_file, lender_row),
            'borrower_po': self.get_optimized_po_numbers(borrower_file, borrower_row),
            'lender_interunit': self.get_optimized_interunit_accounts(lender_file, lender_row),
            'borrower_interunit': self.get_optimized_interunit_accounts(borrower_file, borrower_row),
            'lender_usd': self.get_optimized_usd_amounts(lender_file, lender_row),
            'borrower_usd': self.get_optimized_usd_amounts(borrower_file, borrower_row),
            'lender_file': lender_file,
            'borrower_file': borrower_file,
            'lender_row': lender_row,
            'borrower_row': borrower_row
        }
    
    def _match_narration_ultra_optimized(self, block_data, amount):
        """ULTRA-OPTIMIZED: Narration matching with maximum performance."""
        lender_narration = block_data['lender_narration']
        borrower_narration = block_data['borrower_narration']
        
        if (len(lender_narration) > 10 and len(borrower_narration) > 10 and 
            lender_narration.lower() not in ['nan', 'none', ''] and
            borrower_narration.lower() not in ['nan', 'none', ''] and
            lender_narration == borrower_narration):
            return {'Narration': lender_narration}
        return None
    
    def _match_lc_ultra_optimized(self, block_data, amount):
        """ULTRA-OPTIMIZED: LC matching with maximum performance."""
        lender_lc = block_data['lender_lc']
        borrower_lc = block_data['borrower_lc']
        
        if lender_lc and borrower_lc and lender_lc == borrower_lc:
            return {'LC_Number': lender_lc}
        return None
    
    def _match_po_ultra_optimized(self, block_data, amount):
        """ULTRA-OPTIMIZED: PO matching with maximum performance."""
        lender_po = block_data['lender_po']
        borrower_po = block_data['borrower_po']
        
        if lender_po and borrower_po and lender_po == borrower_po:
            return {'PO_Number': lender_po}
        return None
    
    def _match_interunit_ultra_optimized(self, block_data, amount):
        """ULTRA-OPTIMIZED: Interunit matching with maximum performance."""
        lender_interunit = block_data['lender_interunit']
        borrower_interunit = block_data['borrower_interunit']
        
        if lender_interunit and borrower_interunit and lender_interunit == borrower_interunit:
            return {'Interunit_Account': lender_interunit}
        return None
    
    def _match_usd_ultra_optimized(self, block_data, amount):
        """ULTRA-OPTIMIZED: USD matching with maximum performance."""
        lender_usd = block_data['lender_usd']
        borrower_usd = block_data['borrower_usd']
        
        if lender_usd and borrower_usd and lender_usd == borrower_usd:
            return {'USD_Amount': lender_usd}
        return None
    
    
    def precompute_all_block_data(self, transactions_df):
        """Precompute all block headers, description rows, and narrations for maximum performance."""
        print("Precomputing all block data for maximum performance...")
        
        for idx in range(len(transactions_df)):
            # Precompute block header
            self.get_cached_block_header_universal(idx, transactions_df)
            
            # Precompute description row
            self.get_cached_description_row_universal(idx, transactions_df)
            
            # Precompute narration
            self.get_cached_narration(idx, transactions_df)
            
            # Precompute amounts
            self.get_cached_amounts_universal(idx, transactions_df)
        
        print(f"Precomputed block data for {len(transactions_df)} rows")
    
    def get_optimized_lc_numbers(self, file_num, idx):
        """Get LC numbers using optimized array access."""
        if file_num == 1:
            return self._lc_numbers1_array[idx] if idx < len(self._lc_numbers1_array) else None
        else:
            return self._lc_numbers2_array[idx] if idx < len(self._lc_numbers2_array) else None
    
    def get_optimized_po_numbers(self, file_num, idx):
        """Get PO numbers using optimized array access."""
        if file_num == 1:
            return self._po_numbers1_array[idx] if idx < len(self._po_numbers1_array) else None
        else:
            return self._po_numbers2_array[idx] if idx < len(self._po_numbers2_array) else None
    
    def get_optimized_usd_amounts(self, file_num, idx):
        """Get USD amounts using optimized array access."""
        if file_num == 1:
            return self._usd_amounts1_array[idx] if idx < len(self._usd_amounts1_array) else None
        else:
            return self._usd_amounts2_array[idx] if idx < len(self._usd_amounts2_array) else None
    
    def get_optimized_interunit_accounts(self, file_num, idx):
        """Get interunit accounts using optimized array access."""
        if file_num == 1:
            return self._interunit_accounts1_array[idx] if idx < len(self._interunit_accounts1_array) else None
        else:
            return self._interunit_accounts2_array[idx] if idx < len(self._interunit_accounts2_array) else None
    
    def get_cached_block_header(self, idx, transactions_df, file_num):
        """Get transaction block header with caching for performance."""
        cache = self._block_header_cache1 if file_num == 1 else self._block_header_cache2
        
        if idx not in cache:
            cache[idx] = self.block_identifier.find_transaction_block_header(idx, transactions_df)
        
        return cache[idx]
    
    def get_cached_amounts(self, idx, transactions_df, file_num):
        """Get amounts with caching for performance."""
        cache = self._amount_cache1 if file_num == 1 else self._amount_cache2
        
        if idx not in cache:
            row = transactions_df.iloc[idx]
            cache[idx] = self.extract_amounts_from_strings(row)
        
        return cache[idx]
    
    def preprocess_all_amounts(self, transactions_df, file_num):
        """Preprocess all amounts for a file to populate cache."""
        print(f"Preprocessing amounts for File {file_num}...")
        cache = self._amount_cache1 if file_num == 1 else self._amount_cache2
        
        for idx in range(len(transactions_df)):
            if idx not in cache:
                row = transactions_df.iloc[idx]
                cache[idx] = self.extract_amounts_from_strings(row)
        
        print(f"Preprocessed {len(cache)} amounts for File {file_num}")
    
    def create_amount_index(self, transactions_df, file_num):
        """Create an index of transactions by amount for fast lookup."""
        print(f"Creating amount index for File {file_num}...")
        amount_index = {}
        cache = self._amount_cache1 if file_num == 1 else self._amount_cache2
        
        for idx in range(len(transactions_df)):
            if idx in cache:
                debit, credit = cache[idx]
                
                if debit > 0:  # Lender transaction
                    if debit not in amount_index:
                        amount_index[debit] = {'lenders': [], 'borrowers': []}
                    amount_index[debit]['lenders'].append(idx)
                
                if credit > 0:  # Borrower transaction
                    if credit not in amount_index:
                        amount_index[credit] = {'lenders': [], 'borrowers': []}
                    amount_index[credit]['borrowers'].append(idx)
        
        print(f"Created amount index with {len(amount_index)} unique amounts for File {file_num}")
        return amount_index
    
    def find_lc_matches_optimized(self, transactions1, transactions2, lc_numbers1, lc_numbers2, existing_matches=None, match_id_manager=None):
        """Optimized LC matching using amount pre-filtering."""
        matches = []
        
        print(f"\n=== OPTIMIZED LC MATCHING ===")
        print(f"Using amount pre-filtering for performance...")
        
        # Get all unique amounts that exist in both files
        common_amounts = set(self.amount_index1.keys()) & set(self.amount_index2.keys())
        print(f"Found {len(common_amounts)} common amounts between files")
        
        for amount in common_amounts:
            # Get all lender/borrower pairs for this amount
            file1_lenders = self.amount_index1[amount]['lenders']
            file1_borrowers = self.amount_index1[amount]['borrowers']
            file2_lenders = self.amount_index2[amount]['lenders']
            file2_borrowers = self.amount_index2[amount]['borrowers']
            
            # Check all possible lender-borrower combinations
            for lender_idx in file1_lenders:
                for borrower_idx in file2_borrowers:
                    lc1 = self.get_optimized_lc_numbers(1, lender_idx)
                    lc2 = self.get_optimized_lc_numbers(2, borrower_idx)
                    if lc1 and lc2 and lc1 == lc2:
                        # Found LC match
                        matches.append({
                            'match_id': None,
                            'Match_Type': 'LC',
                            'File1_Index': lender_idx,
                            'File2_Index': borrower_idx,
                            'LC_Number': lc1,
                            'Amount': amount
                        })
            
            # Check reverse combinations
            for lender_idx in file2_lenders:
                for borrower_idx in file1_borrowers:
                    lc1 = self.get_optimized_lc_numbers(1, borrower_idx)
                    lc2 = self.get_optimized_lc_numbers(2, lender_idx)
                    if lc1 and lc2 and lc1 == lc2:
                        # Found LC match
                        matches.append({
                            'match_id': None,
                            'Match_Type': 'LC',
                            'File1_Index': borrower_idx,
                            'File2_Index': lender_idx,
                            'LC_Number': lc1,
                            'Amount': amount
                        })
        
        print(f"Found {len(matches)} LC matches using optimized method")
        return matches
    
    def find_po_matches_optimized(self, transactions1, transactions2, po_numbers1, po_numbers2, existing_matches=None, match_id_manager=None):
        """Optimized PO matching using amount pre-filtering."""
        matches = []
        
        print(f"\n=== OPTIMIZED PO MATCHING ===")
        print(f"Using amount pre-filtering for performance...")
        
        # Get all unique amounts that exist in both files
        common_amounts = set(self.amount_index1.keys()) & set(self.amount_index2.keys())
        print(f"Found {len(common_amounts)} common amounts between files")
        
        for amount in common_amounts:
            # Get all lender/borrower pairs for this amount
            file1_lenders = self.amount_index1[amount]['lenders']
            file1_borrowers = self.amount_index1[amount]['borrowers']
            file2_lenders = self.amount_index2[amount]['lenders']
            file2_borrowers = self.amount_index2[amount]['borrowers']
            
            # Check all possible lender-borrower combinations
            for lender_idx in file1_lenders:
                for borrower_idx in file2_borrowers:
                    po1 = self.get_optimized_po_numbers(1, lender_idx)
                    po2 = self.get_optimized_po_numbers(2, borrower_idx)
                    if po1 and po2 and po1 == po2:
                        # Found PO match
                        matches.append({
                            'match_id': None,
                            'Match_Type': 'PO',
                            'File1_Index': lender_idx,
                            'File2_Index': borrower_idx,
                            'PO_Number': po1,
                            'Amount': amount
                        })
            
            # Check reverse combinations
            for lender_idx in file2_lenders:
                for borrower_idx in file1_borrowers:
                    po1 = self.get_optimized_po_numbers(1, borrower_idx)
                    po2 = self.get_optimized_po_numbers(2, lender_idx)
                    if po1 and po2 and po1 == po2:
                        # Found PO match
                        matches.append({
                            'match_id': None,
                            'Match_Type': 'PO',
                            'File1_Index': borrower_idx,
                            'File2_Index': lender_idx,
                            'PO_Number': po1,
                            'Amount': amount
                        })
        
        print(f"Found {len(matches)} PO matches using optimized method")
        return matches
    
    def find_usd_matches_optimized(self, transactions1, transactions2, usd_amounts1, usd_amounts2, existing_matches=None, match_id_manager=None):
        """Optimized USD matching using amount pre-filtering."""
        matches = []
        
        print(f"\n=== OPTIMIZED USD MATCHING ===")
        print(f"Using amount pre-filtering for performance...")
        
        # Get all unique amounts that exist in both files
        common_amounts = set(self.amount_index1.keys()) & set(self.amount_index2.keys())
        print(f"Found {len(common_amounts)} common amounts between files")
        
        for amount in common_amounts:
            # Get all lender/borrower pairs for this amount
            file1_lenders = self.amount_index1[amount]['lenders']
            file1_borrowers = self.amount_index1[amount]['borrowers']
            file2_lenders = self.amount_index2[amount]['lenders']
            file2_borrowers = self.amount_index2[amount]['borrowers']
            
            # Check all possible lender-borrower combinations
            for lender_idx in file1_lenders:
                for borrower_idx in file2_borrowers:
                    usd1 = self.get_optimized_usd_amounts(1, lender_idx)
                    usd2 = self.get_optimized_usd_amounts(2, borrower_idx)
                    if usd1 and usd2 and usd1 == usd2:
                        # Found USD match
                        matches.append({
                            'match_id': None,
                            'Match_Type': 'USD',
                            'File1_Index': lender_idx,
                            'File2_Index': borrower_idx,
                            'USD_Amount': usd1,
                            'Amount': amount
                        })
            
            # Check reverse combinations
            for lender_idx in file2_lenders:
                for borrower_idx in file1_borrowers:
                    usd1 = self.get_optimized_usd_amounts(1, borrower_idx)
                    usd2 = self.get_optimized_usd_amounts(2, lender_idx)
                    if usd1 and usd2 and usd1 == usd2:
                        # Found USD match
                        matches.append({
                            'match_id': None,
                            'Match_Type': 'USD',
                            'File1_Index': borrower_idx,
                            'File2_Index': lender_idx,
                            'USD_Amount': usd1,
                            'Amount': amount
                        })
        
        print(f"Found {len(matches)} USD matches using optimized method")
        return matches
    
    def find_narration_matches_optimized(self, transactions1, transactions2, existing_matches=None, match_id_manager=None):
        """Optimized Narration matching using amount pre-filtering."""
        matches = []
        
        print(f"\n=== OPTIMIZED NARRATION MATCHING ===")
        print(f"Using amount pre-filtering for performance...")
        
        # Get all unique amounts that exist in both files
        common_amounts = set(self.amount_index1.keys()) & set(self.amount_index2.keys())
        print(f"Found {len(common_amounts)} common amounts between files")
        
        for amount in common_amounts:
            # Get all lender/borrower pairs for this amount
            file1_lenders = self.amount_index1[amount]['lenders']
            file1_borrowers = self.amount_index1[amount]['borrowers']
            file2_lenders = self.amount_index2[amount]['lenders']
            file2_borrowers = self.amount_index2[amount]['borrowers']
            
            # Check all possible lender-borrower combinations
            for lender_idx in file1_lenders:
                for borrower_idx in file2_borrowers:
                    # Get narrations for comparison using cached method
                    lender_narration = self.get_cached_narration(lender_idx, transactions1)
                    borrower_narration = self.get_cached_narration(borrower_idx, transactions2)
                    
                    # Check for exact narration match
                    if (len(lender_narration) > 10 and len(borrower_narration) > 10 and 
                        lender_narration.lower() not in ['nan', 'none', ''] and
                        borrower_narration.lower() not in ['nan', 'none', ''] and
                        lender_narration == borrower_narration):
                        # Found narration match
                        matches.append({
                            'match_id': None,
                            'Match_Type': 'Narration',
                            'File1_Index': lender_idx,
                            'File2_Index': borrower_idx,
                            'Narration': lender_narration,
                            'Amount': amount
                        })
            
            # Check reverse combinations
            for lender_idx in file2_lenders:
                for borrower_idx in file1_borrowers:
                    # Get narrations for comparison using cached method
                    lender_narration = self.get_cached_narration(lender_idx, transactions2)
                    borrower_narration = self.get_cached_narration(borrower_idx, transactions1)
                    
                    # Check for exact narration match
                    if (len(lender_narration) > 10 and len(borrower_narration) > 10 and 
                        lender_narration.lower() not in ['nan', 'none', ''] and
                        borrower_narration.lower() not in ['nan', 'none', ''] and
                        lender_narration == borrower_narration):
                        # Found narration match
                        matches.append({
                            'match_id': None,
                            'Match_Type': 'Narration',
                            'File1_Index': borrower_idx,
                            'File2_Index': lender_idx,
                            'Narration': lender_narration,
                            'Amount': amount
                        })
        
        print(f"Found {len(matches)} Narration matches using optimized method")
        return matches
    
    def find_interunit_matches_optimized(self, transactions1, transactions2, interunit_accounts1, interunit_accounts2, existing_matches=None, match_id_manager=None):
        """Optimized Interunit matching using cached formatting data."""
        matches = []
        
        print(f"\n=== OPTIMIZED INTERUNIT MATCHING ===")
        print(f"Using cached formatting data for performance...")
        
        # Get cached formatting data
        file1_interunit_data = self.get_cached_formatting_data(1)
        file2_interunit_data = self.get_cached_formatting_data(2)
        
        print(f"File 1: {len(file1_interunit_data)} blocks with interunit data")
        print(f"File 2: {len(file2_interunit_data)} blocks with interunit data")
        
        # Look for cross-referenced matches using cached data
        print(f"\n--- Looking for cross-referenced matches ---")
        
        for block1 in file1_interunit_data:
            for block2 in file2_interunit_data:
                # Check if blocks have opposite transaction types (one debit, one credit)
                if (block1['amounts'] and block2['amounts'] and
                    ((block1['amounts']['debit'] and block2['amounts']['credit']) or
                     (block1['amounts']['credit'] and block2['amounts']['debit']))):
                    
                    # Check if amounts match exactly (NO TOLERANCE)
                    amount1 = block1['amounts']['debit'] if block1['amounts']['debit'] else block1['amounts']['credit']
                    amount2 = block2['amounts']['debit'] if block2['amounts']['debit'] else block2['amounts']['credit']
                    
                    if amount1 == amount2:
                        # Check for cross-referenced short codes
                        cross_reference_found = False
                        file1_narration_contains = None
                        file2_narration_contains = None
                        
                        # File 1's narration should contain File 2's short code
                        for narration1 in block1['narration_short_codes']:
                            for ledger2 in block2['ledger_accounts']:
                                if narration1['short_code'] == ledger2['short_code']:
                                    cross_reference_found = True
                                    file1_narration_contains = narration1['short_code']
                                    break
                            if cross_reference_found:
                                break
                        
                        # File 2's narration should contain File 1's short code
                        if cross_reference_found:
                            for narration2 in block2['narration_short_codes']:
                                for ledger1 in block1['ledger_accounts']:
                                    if narration2['short_code'] == ledger1['short_code']:
                                        file2_narration_contains = narration2['short_code']
                                        
                                        # We have a match!
                                        match = {
                                            'match_id': None,
                                            'Match_Type': 'Interunit',
                                            'Interunit_Account': f"{file1_narration_contains} <-> {file2_narration_contains}",
                                            'File1_Index': block1['amounts']['row'],
                                            'File2_Index': block2['amounts']['row'],
                                            'File1_Debit': block1['amounts']['debit'],
                                            'File1_Credit': block1['amounts']['credit'],
                                            'File2_Debit': block2['amounts']['debit'],
                                            'File2_Credit': block2['amounts']['credit'],
                                            'File1_Amount': amount1,
                                            'File2_Amount': amount1,
                                            'Amount': amount1
                                        }
                                        
                                        matches.append(match)
                                        print(f"  MATCH: Amount {amount1}")
                                        print(f"    Cross-reference: File 1 narration contains {file1_narration_contains}")
                                        print(f"    Cross-reference: File 2 narration contains {file2_narration_contains}")
                                        break
                                
                                if file2_narration_contains:
                                    break
        
        print(f"Found {len(matches)} Interunit matches using optimized method")
        return matches
    
    def find_aggregated_po_matches_optimized(self, transactions1, transactions2, po_numbers1, po_numbers2, existing_matches=None, match_id_manager=None):
        """Optimized Aggregated PO matching using amount pre-filtering."""
        matches = []
        
        print(f"\n=== OPTIMIZED AGGREGATED PO MATCHING ===")
        print(f"Using amount pre-filtering for performance...")
        
        # Get all unique amounts that exist in both files
        common_amounts = set(self.amount_index1.keys()) & set(self.amount_index2.keys())
        print(f"Found {len(common_amounts)} common amounts between files")
        
        # Import regex for PO pattern matching
        import re
        from .po_matching_logic import PO_PATTERN
        
        for amount in common_amounts:
            # Get all lender/borrower pairs for this amount
            file1_lenders = self.amount_index1[amount]['lenders']
            file1_borrowers = self.amount_index1[amount]['borrowers']
            file2_lenders = self.amount_index2[amount]['lenders']
            file2_borrowers = self.amount_index2[amount]['borrowers']
            
            # Check all possible lender-borrower combinations
            for lender_idx in file1_lenders:
                for borrower_idx in file2_borrowers:
                    # Get narrations for PO extraction using cached method
                    lender_narration = self.get_cached_narration(lender_idx, transactions1)
                    borrower_narration = self.get_cached_narration(borrower_idx, transactions2)
                    
                    # Extract PO numbers from narrations using compiled pattern
                    lender_pos = self._compiled_po_pattern.findall(lender_narration)
                    borrower_pos = self._compiled_po_pattern.findall(borrower_narration)
                    
                    # Check if lender has multiple POs and borrower has matching POs
                    if len(lender_pos) >= 2 and len(borrower_pos) >= 1:
                        # Check if all lender POs are present in borrower
                        if all(po in borrower_pos for po in lender_pos):
                            # Found aggregated PO match
                            matches.append({
                                'match_id': None,
                                'Match_Type': 'Aggregated_PO',
                                'File1_Index': lender_idx,
                                'File2_Index': borrower_idx,
                                'PO_Count': len(lender_pos),
                                'All_POs': lender_pos,
                                'Amount': amount
                            })
            
            # Check reverse combinations
            for lender_idx in file2_lenders:
                for borrower_idx in file1_borrowers:
                    # Get narrations for PO extraction using cached method
                    lender_narration = self.get_cached_narration(lender_idx, transactions2)
                    borrower_narration = self.get_cached_narration(borrower_idx, transactions1)
                    
                    # Extract PO numbers from narrations using compiled pattern
                    lender_pos = self._compiled_po_pattern.findall(lender_narration)
                    borrower_pos = self._compiled_po_pattern.findall(borrower_narration)
                    
                    # Check if lender has multiple POs and borrower has matching POs
                    if len(lender_pos) >= 2 and len(borrower_pos) >= 1:
                        # Check if all lender POs are present in borrower
                        if all(po in borrower_pos for po in lender_pos):
                            # Found aggregated PO match
                            matches.append({
                                'match_id': None,
                                'Match_Type': 'Aggregated_PO',
                                'File1_Index': borrower_idx,
                                'File2_Index': lender_idx,
                                'PO_Count': len(lender_pos),
                                'All_POs': lender_pos,
                                'Amount': amount
                            })
        
        print(f"Found {len(matches)} Aggregated PO matches using optimized method")
        return matches
    
    def extract_lc_numbers(self, description_series):
        """Extract LC numbers from transaction descriptions."""
        def extract_single_lc(description):
            if pd.isna(description):
                return None
            
            # Pattern for LC numbers: L/C-123/456, LC-123/456, or similar formats
            match = self._compiled_lc_pattern.search(str(description).upper())
            return match.group() if match else None
        
        return description_series.apply(extract_single_lc)
    
    def extract_po_numbers(self, description_series):
        """Extract PO numbers from transaction descriptions."""
        def extract_single_po(description):
            if pd.isna(description):
                return None
            
            # Pattern for PO numbers: XXX/PO/YYYY/M/NNNNN format
            match = self._compiled_po_pattern.search(str(description).upper())
            return match.group() if match else None
        
        return description_series.apply(extract_single_po)
    
    def extract_lc_numbers_from_narration(self, file_path):
        """Extract LC numbers from narration rows (regular text Column C - not bold, not italic) using openpyxl formatting."""
        lc_numbers = []
        lc_parent_rows = []
        
        # Load workbook with openpyxl to access formatting
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        for row in range(9, ws.max_row + 1):  # Start from row 9 (after headers)
            desc_cell = ws.cell(row=row, column=3)
            
            # Check if this is a narration row (italic text Column C - not bold, but italic)
            is_narration = (desc_cell.value and 
                           desc_cell.font and 
                           not desc_cell.font.bold and 
                           desc_cell.font.italic)
            
            if is_narration:
                # This is a narration row, check for LC numbers
                narration_text = str(desc_cell.value)
                lc = self.extract_lc_numbers(pd.Series([narration_text])).iloc[0]
                
                if lc is not None:
                    # Found LC in narration row, need to find parent transaction row
                    parent_row = self.find_parent_transaction_row_with_formatting(ws, row)
                    if parent_row is not None:

                        lc_numbers.append(lc)
                        lc_parent_rows.append(parent_row)
                    else:

                        lc_numbers.append(None)
                        lc_parent_rows.append(None)
                else:
                    lc_numbers.append(None)
                    lc_parent_rows.append(None)
            else:
                lc_numbers.append(None)
                lc_parent_rows.append(None)
        
        wb.close()
        
        # Store parent row mapping for later use
        
        return pd.Series(lc_numbers)
    
    def extract_po_numbers_from_narration(self, file_path):
        """Extract PO numbers from narration rows (italic text Column C - not bold, but italic) using openpyxl formatting."""
        # Load workbook with openpyxl to access formatting
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Create a list to store PO numbers for each row in the DataFrame
        # We need to align with the transactions DataFrame structure
        po_numbers = []
        po_parent_rows = []
        
        # First, get the transactions DataFrame to know how many rows we need
        transactions_df = self.read_complex_excel(file_path)[1]  # Get transactions part
        total_rows = len(transactions_df)
        
        # Initialize with None for all rows
        for _ in range(total_rows):
            po_numbers.append(None)
            po_parent_rows.append(None)
        
        # Now scan for PO numbers in narration rows and map them to DataFrame indices
        for excel_row in range(9, ws.max_row + 1):  # Excel rows start from 9
            desc_cell = ws.cell(row=excel_row, column=3)
            
            # Check if this is a narration row (italic text Column C - not bold, but italic)
            is_narration = (desc_cell.value and 
                           desc_cell.font and 
                           not desc_cell.font.bold and 
                           desc_cell.font.italic)
            
            if is_narration:
                # This is a narration row, check for PO numbers
                narration_text = str(desc_cell.value)
                po = self.extract_po_numbers(pd.Series([narration_text])).iloc[0]
                
                if po is not None:
                    # Found PO in narration row, need to find parent transaction row
                    parent_row = self.find_parent_transaction_row_with_formatting(ws, excel_row)
                    if parent_row is not None:
                        # Convert Excel row to DataFrame index
                        df_index = parent_row - 9  # Excel row 9 = DataFrame index 0
                        if 0 <= df_index < total_rows:
                            # print(f"DEBUG: PO {po} at Excel row {excel_row} -> DataFrame index {df_index}")
                            po_numbers[df_index] = po
                            po_parent_rows[df_index] = df_index
                        else:
                            # print(f"DEBUG: PO {po} at Excel row {excel_row} - INVALID DataFrame index {df_index}")
                            pass
                    else:

                        pass
        
        wb.close()
        
        # Store parent row mapping for later use
        
        return pd.Series(po_numbers)

    def load_workbooks_and_extract_data_optimized(self):
        """
        Extract all required data using cached workbooks and compiled regex patterns.
        This is the most optimized version that reuses cached data.
        """
        if self._cached_extracted_data is not None:
            print("Using cached extracted data...")
            return self._cached_extracted_data
        
        print("Extracting data using cached workbooks and compiled regex...")
        
        # Use cached workbooks (already loaded)
        ws1 = self._cached_ws1
        ws2 = self._cached_ws2
        
        # Extract all data from File 1
        lc_numbers1 = []
        po_numbers1 = []
        usd_amounts1 = []
        interunit_accounts1 = []
        
        # Process File 1 rows 9 onwards (optimized with compiled regex)
        for row in range(9, ws1.max_row + 1):
            narration = ws1.cell(row=row, column=3).value  # Column C is narration
            if narration:
                # Cache upper case conversion
                narration_upper = str(narration).upper()
                
                # Extract LC numbers using compiled pattern
                lc_matches = self._compiled_lc_pattern.findall(narration_upper)
                if lc_matches:
                    lc_numbers1.append((row, lc_matches[0]))
                
                # Extract PO numbers using compiled pattern
                po_matches = self._compiled_po_pattern.findall(narration_upper)
                if po_matches:
                    po_numbers1.append((row, po_matches[0]))
                
                # Extract USD amounts using compiled pattern
                usd_matches = self._compiled_usd_pattern.findall(narration_upper)
                if usd_matches:
                    usd_amounts1.append((row, usd_matches[0]))
                
                # Extract interunit accounts using compiled pattern
                interunit_matches = self._compiled_interunit_pattern.findall(narration_upper)
                if interunit_matches:
                    interunit_accounts1.append((row, f"{interunit_matches[0][0]}#{interunit_matches[0][1]}"))
        
        # Convert to Series with proper indexing (matching original logic)
        # Create Series with same length as transactions DataFrame, initialized with None
        total_rows1 = len(self.transactions1)
        lc_numbers1_series = pd.Series([None] * total_rows1, index=range(total_rows1))
        po_numbers1_series = pd.Series([None] * total_rows1, index=range(total_rows1))
        usd_amounts1_series = pd.Series([None] * total_rows1, index=range(total_rows1))
        interunit_accounts1_series = pd.Series([None] * total_rows1, index=range(total_rows1))
        
        # Now populate the found items at their correct DataFrame indices
        for row, lc_num in lc_numbers1:
            df_index = row - 9  # Excel row 9 = DataFrame index 0
            if 0 <= df_index < total_rows1:
                lc_numbers1_series.iloc[df_index] = lc_num
        
        for row, po_num in po_numbers1:
            df_index = row - 9  # Excel row 9 = DataFrame index 0
            if 0 <= df_index < total_rows1:
                po_numbers1_series.iloc[df_index] = po_num
        
        for row, usd_amount in usd_amounts1:
            df_index = row - 9  # Excel row 9 = DataFrame index 0
            if 0 <= df_index < total_rows1:
                usd_amounts1_series.iloc[df_index] = usd_amount
        
        for row, account in interunit_accounts1:
            df_index = row - 9  # Excel row 9 = DataFrame index 0
            if 0 <= df_index < total_rows1:
                interunit_accounts1_series.iloc[df_index] = account
        
        # Extract all data from File 2
        lc_numbers2 = []
        po_numbers2 = []
        usd_amounts2 = []
        interunit_accounts2 = []
        
        # Process File 2 rows 9 onwards (optimized with compiled regex)
        for row in range(9, ws2.max_row + 1):
            narration = ws2.cell(row=row, column=3).value  # Column C is narration
            if narration:
                # Cache upper case conversion
                narration_upper = str(narration).upper()
                
                # Extract LC numbers using compiled pattern
                lc_matches = self._compiled_lc_pattern.findall(narration_upper)
                if lc_matches:
                    lc_numbers2.append((row, lc_matches[0]))
                
                # Extract PO numbers using compiled pattern
                po_matches = self._compiled_po_pattern.findall(narration_upper)
                if po_matches:
                    po_numbers2.append((row, po_matches[0]))
                
                # Extract USD amounts using compiled pattern
                usd_matches = self._compiled_usd_pattern.findall(narration_upper)
                if usd_matches:
                    usd_amounts2.append((row, usd_matches[0]))
                
                # Extract interunit accounts using compiled pattern
                interunit_matches = self._compiled_interunit_pattern.findall(narration_upper)
                if interunit_matches:
                    interunit_accounts2.append((row, f"{interunit_matches[0][0]}#{interunit_matches[0][1]}"))
        
        # Convert to Series with proper indexing (matching original logic)
        # Create Series with same length as transactions DataFrame, initialized with None
        total_rows2 = len(self.transactions2)
        lc_numbers2_series = pd.Series([None] * total_rows2, index=range(total_rows2))
        po_numbers2_series = pd.Series([None] * total_rows2, index=range(total_rows2))
        usd_amounts2_series = pd.Series([None] * total_rows2, index=range(total_rows2))
        interunit_accounts2_series = pd.Series([None] * total_rows2, index=range(total_rows2))
        
        # Now populate the found items at their correct DataFrame indices
        for row, lc_num in lc_numbers2:
            df_index = row - 9  # Excel row 9 = DataFrame index 0
            if 0 <= df_index < total_rows2:
                lc_numbers2_series.iloc[df_index] = lc_num
        
        for row, po_num in po_numbers2:
            df_index = row - 9  # Excel row 9 = DataFrame index 0
            if 0 <= df_index < total_rows2:
                po_numbers2_series.iloc[df_index] = po_num
        
        for row, usd_amount in usd_amounts2:
            df_index = row - 9  # Excel row 9 = DataFrame index 0
            if 0 <= df_index < total_rows2:
                usd_amounts2_series.iloc[df_index] = usd_amount
        
        for row, account in interunit_accounts2:
            df_index = row - 9  # Excel row 9 = DataFrame index 0
            if 0 <= df_index < total_rows2:
                interunit_accounts2_series.iloc[df_index] = account
        
        print(f"Data extraction complete:")
        print(f"  File 1: {len(lc_numbers1)} LC, {len(po_numbers1)} PO, {len(usd_amounts1)} USD, {len(interunit_accounts1)} Interunit")
        print(f"  File 2: {len(lc_numbers2)} LC, {len(po_numbers2)} PO, {len(usd_amounts2)} USD, {len(interunit_accounts2)} Interunit")
        
        # Cache the extracted data
        self._cached_extracted_data = {
            'lc_numbers1': lc_numbers1_series,
            'po_numbers1': po_numbers1_series,
            'usd_amounts1': usd_amounts1_series,
            'interunit_accounts1': interunit_accounts1_series,
            'lc_numbers2': lc_numbers2_series,
            'po_numbers2': po_numbers2_series,
            'usd_amounts2': usd_amounts2_series,
            'interunit_accounts2': interunit_accounts2_series
        }
        
        return self._cached_extracted_data

    def process_files(self):
        """Process both files and prepare for matching with performance optimizations."""
        print("Reading Pole Book STEEL.xlsx...")
        self.metadata1, self.transactions1 = self.read_complex_excel(self.file1_path)
        
        print("Reading Steel Book POLE.xlsx...")
        self.metadata2, self.transactions2 = self.read_complex_excel(self.file2_path)
        
        print(f"File 1: {len(self.transactions1)} rows")
        print(f"File 2: {len(self.transactions2)} rows")
        
        # Load and cache workbooks once for all operations
        self.load_and_cache_workbooks()
        
        # Preprocess amounts for performance optimization
        print("Preprocessing amounts for performance...")
        self.preprocess_all_amounts(self.transactions1, 1)
        self.preprocess_all_amounts(self.transactions2, 2)
        
        # Create amount indexes for fast lookup
        print("Creating amount indexes for fast lookup...")
        self.amount_index1 = self.create_amount_index(self.transactions1, 1)
        self.amount_index2 = self.create_amount_index(self.transactions2, 2)

        print(f"File 1 columns: {list(self.transactions1.columns)}")
        print(f"File 2 columns: {list(self.transactions2.columns)}")
        
        # Find the description column (should be the 3rd column, index 2)
        # Let's check what's actually in the columns
        print(f"File 1 first row: {list(self.transactions1.iloc[0, :])}")
        
        # Extract all data using cached workbooks and compiled regex
        print("Extracting all data using optimized methods...")
        extracted_data = self.load_workbooks_and_extract_data_optimized()
        
        # Extract LC numbers from both files
        lc_numbers1 = extracted_data['lc_numbers1']
        lc_numbers2 = extracted_data['lc_numbers2']
        
        # Extract PO numbers from both files
        po_numbers1 = extracted_data['po_numbers1']
        po_numbers2 = extracted_data['po_numbers2']
        
        # Extract interunit loan accounts from both files
        interunit_accounts1 = extracted_data['interunit_accounts1']
        interunit_accounts2 = extracted_data['interunit_accounts2']
        
        # Extract USD amounts from both files
        usd_amounts1 = extracted_data['usd_amounts1']
        usd_amounts2 = extracted_data['usd_amounts2']
        
        # Get cached transaction blocks (identified once, reused everywhere)
        print("Getting cached transaction blocks...")
        blocks1 = self.get_cached_blocks(1)
        blocks2 = self.get_cached_blocks(2)
        
        print(f"File 1: {len(blocks1)} transaction blocks")
        print(f"File 2: {len(blocks2)} transaction blocks")
        
        # Create optimized arrays for faster data access
        self.create_optimized_arrays(
            lc_numbers1, lc_numbers2, po_numbers1, po_numbers2,
            usd_amounts1, usd_amounts2, interunit_accounts1, interunit_accounts2
        )
        
        # Selective precomputation based on file size
        total_rows = len(self.transactions1) + len(self.transactions2)
        if total_rows > 2000:  # Only precompute for large files
            print("Large files detected - precomputing block data for maximum performance...")
            self.precompute_all_block_data(self.transactions1)
            self.precompute_all_block_data(self.transactions2)
        else:
            print("Small files detected - using on-demand caching for optimal performance...")
        
        return self.transactions1, self.transactions2, blocks1, blocks2, lc_numbers1, lc_numbers2, po_numbers1, po_numbers2, interunit_accounts1, interunit_accounts2, usd_amounts1, usd_amounts2
    
    def find_potential_matches(self):
        """Find potential matches using block-based optimization (analyze each block once)."""

        transactions1, transactions2, blocks1, blocks2, lc_numbers1, lc_numbers2, po_numbers1, po_numbers2, interunit_accounts1, interunit_accounts2, usd_amounts1, usd_amounts2 = self.process_files()
        
        print("\n" + "="*70)
        print("🚀 BLOCK-BASED OPTIMIZATION - ANALYZE EACH BLOCK ONCE")
        print("="*70)
        print("Instead of running each match type sequentially, we analyze each transaction block")
        print("once and determine the best match type for that block. This eliminates redundant")
        print("processing and improves performance significantly.")
        
        # Use block-based optimization for maximum performance
        all_matches = self.find_matches_block_based_optimized(
            transactions1, transactions2, blocks1, blocks2,
            lc_numbers1, lc_numbers2, po_numbers1, po_numbers2,
            interunit_accounts1, interunit_accounts2, usd_amounts1, usd_amounts2
        )
        
        # ARCHITECTURAL FIX: Assign sequential Match IDs to all matches
        print(f"\n=== ASSIGNING SEQUENTIAL MATCH IDs ===")
        print(f"Total matches found: {len(all_matches)}")
        
        # Initialize Match ID counter
        match_counter = 1
        
        # Assign sequential Match IDs to all matches
        for i, match in enumerate(all_matches):
            match_id = f"M{match_counter:03d}"  # Format as M001, M002, M003, etc.
            old_match_id = match.get('match_id', 'None')
            match['match_id'] = match_id
            match_counter += 1
            print(f"Match {i+1}: Assigned {match_id} to {match.get('Match_Type', 'Unknown')} match (was {old_match_id})")
        
        print(f"Assigned {len(all_matches)} sequential Match IDs (M001 to M{match_counter-1:03d})")
        
        # Sort matches by the newly assigned sequential Match IDs
        all_matches.sort(key=lambda x: x['match_id'])
        print(f"Sorted matches by sequential Match IDs")
        
        # Show match breakdown
        match_types = {}
        for match in all_matches:
            match_type = match.get('Match_Type', 'Unknown')
            match_types[match_type] = match_types.get(match_type, 0) + 1
        
        print(f"\n" + "="*60)
        print("FINAL MATCH SUMMARY")
        print("="*60)
        print(f"Total matches found: {len(all_matches)}")
        print(f"Match IDs assigned: M001 to M{match_counter-1:03d}")
        
        print("="*60)
        print("FINAL RESULTS")
        print("="*60)
        print(f"Total Matches: {len(all_matches)}")
        for match_type, count in match_types.items():
            print(f"  - {match_type} Matches: {count}")
        
        # Clean up all caches to free memory
        self.clear_all_caches()
        
        return all_matches
    
    def find_parent_transaction_row_with_formatting(self, ws, current_row):
        """Find the parent transaction row for a narration row using openpyxl formatting."""
        # Look backwards from current row to find the most recent transaction block header
        for row_idx in range(current_row, 8, -1):  # Start from current_row, go back to row 9
            date_cell = ws.cell(row=row_idx, column=1)
            particulars_cell = ws.cell(row=row_idx, column=2)
            desc_cell = ws.cell(row=row_idx, column=3)
            
            # Check if this is a transaction block header (Date + Dr/Cr + BOLD Col C)
            has_date = date_cell.value is not None
            has_dr_cr = particulars_cell.value and str(particulars_cell.value).strip() in ['Dr', 'Cr']
            has_bold_desc = desc_cell.font and desc_cell.font.bold
            
            if has_date and has_dr_cr and has_bold_desc:
                return row_idx
        
        return None
    
    def create_audit_info(self, match):
        """Create audit info in clean, readable plaintext format for LC, PO, Interunit, and USD matches."""
        # Determine match type and create appropriate audit info
        if 'Match_Type' in match:
            # Use explicit match type if available
            match_type = match['Match_Type']
            amount = match.get('File1_Amount', match.get('File2_Amount', 0))
            
            if match_type == 'Narration':
                audit_info = f"Narration Match\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
            elif match_type == 'LC':
                lc_number = match.get('LC_Number', 'Unknown')
                audit_info = f"LC Match: {lc_number}\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
            elif match_type == 'PO':
                po_number = match.get('PO_Number', 'Unknown')
                audit_info = f"PO Match: {po_number}\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
            elif match_type == 'Interunit':
                interunit_account = match.get('Interunit_Account', 'Unknown')
                audit_info = f"Interunit Loan Match: {interunit_account}\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
            elif match_type == 'USD':
                usd_amount = match.get('USD_Amount', 'Unknown')
                audit_info = f"USD Match: {usd_amount}\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
            elif match_type == 'Aggregated_PO':
                po_count = match.get('PO_Count', 'Unknown')
                all_pos = match.get('All_POs', [])
                po_list = ', '.join(all_pos[:5])  # Show first 5 POs
                if len(all_pos) > 5:
                    po_list += f" ... and {len(all_pos) - 5} more"
                audit_info = f"Aggregated PO Match: {po_count} POs\nPOs: {po_list}\nLender Amount: {amount:.2f}\nTotal Borrower Amount: {amount:.2f}"
            elif match_type == 'Manual':
                audit_info = f"Manual Match\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
            else:
                audit_info = f"{match_type} Match\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
        else:
            # Fallback to old logic for backward compatibility
            if 'LC_Number' in match and match['LC_Number']:
                # This is an LC match - use File1_Amount or File2_Amount
                amount = match.get('File1_Amount', match.get('File2_Amount', 0))
                audit_info = f"LC Match: {match['LC_Number']}\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
            elif 'PO_Number' in match and match['PO_Number']:
                # This is a PO match - use File1_Amount or File2_Amount
                amount = match.get('File1_Amount', match.get('File2_Amount', 0))
                audit_info = f"PO Match: {match['PO_Number']}\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
            elif 'Interunit_Account' in match and match['Interunit_Account']:
                # This is an Interunit Loan match - use File1_Amount or File2_Amount
                amount = match.get('File1_Amount', match.get('File2_Amount', 0))
                audit_info = f"Interunit Loan Match: {match['Interunit_Account']}\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
            elif 'USD_Amount' in match and match['USD_Amount']:
                # This is a USD match - use File1_Amount or File2_Amount
                amount = match.get('File1_Amount', match.get('File2_Amount', 0))
                audit_info = f"USD Match: {match['USD_Amount']}\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
            else:
                # Fallback for unknown match type
                amount = match.get('File1_Amount', match.get('File2_Amount', 0))
                audit_info = f"Unknown Match Type\nLender Amount: {amount:.2f}\nBorrower Amount: {amount:.2f}"
        
        return audit_info

    def _format_amount_columns(self, worksheet):
        """Format amount columns (Debit and Credit) to prevent scientific notation."""
        # Debit column (J) and Credit column (K) - after adding Match ID and Audit Info
        debit_col = 9  # Column J (0-indexed)
        credit_col = 10  # Column K (0-indexed)
        
        # Format all data rows (starting from row 9)
        for row in range(9, worksheet.max_row + 1):
            try:
                # Format Debit column
                debit_cell = worksheet.cell(row=row, column=debit_col + 1)  # openpyxl uses 1-indexed
                if debit_cell.value is not None and debit_cell.value != '':
                    debit_cell.number_format = '#,##0.00'
                
                # Format Credit column
                credit_cell = worksheet.cell(row=row, column=credit_col + 1)  # openpyxl uses 1-indexed
                if credit_cell.value is not None and credit_cell.value != '':
                    credit_cell.number_format = '#,##0.00'
                    
            except Exception as e:
                print(f"Error formatting amount columns for row {row}: {e}")

    def _format_date_columns(self, worksheet):
        """Format date columns to display as dd/mmm/yyyy without timestamps."""
        # Date column (C) - after adding Match ID and Audit Info
        # Match ID (A) = column 1, Audit Info (B) = column 2, Date (C) = column 3
        date_col = 3  # Column C (1-indexed in Excel)
        
        from datetime import datetime
        
        # Format all data rows (starting from row 10, which is after header row 9)
        for row in range(10, worksheet.max_row + 1):
            try:
                date_cell = worksheet.cell(row=row, column=date_col)
                if date_cell.value is not None and date_cell.value != '':
                    # If the value is a datetime object, convert it to string in dd/mmm/yyyy format
                    if isinstance(date_cell.value, datetime):
                        # Convert datetime to string in dd/mmm/yyyy format
                        date_cell.value = date_cell.value.strftime('%d/%b/%Y')
                    elif isinstance(date_cell.value, str) and ' ' in date_cell.value:
                        # If it's a string with timestamp, parse and convert
                        try:
                            parsed_date = pd.to_datetime(date_cell.value)
                            date_cell.value = parsed_date.strftime('%d/%b/%Y')
                        except:
                            pass  # Keep original if parsing fails
                    
                    # Set cell as text format to preserve the string format
                    date_cell.number_format = '@'  # Text format
                    
            except Exception as e:
                print(f"Error formatting date column for row {row}: {e}")

    def _set_column_widths(self, worksheet):
        """Set column widths for the worksheet"""
        worksheet.column_dimensions['A'].width = 9.00
        worksheet.column_dimensions['B'].width = 30.00
        worksheet.column_dimensions['C'].width = 12.00
        worksheet.column_dimensions['D'].width = 10.33
        worksheet.column_dimensions['E'].width = 60.00
        worksheet.column_dimensions['F'].width = 5.00
        worksheet.column_dimensions['G'].width = 5.00
        worksheet.column_dimensions['H'].width = 12.78
        worksheet.column_dimensions['I'].width = 9.00
        worksheet.column_dimensions['J'].width = 13.78
        worksheet.column_dimensions['K'].width = 14.22
        worksheet.column_dimensions['L'].width = 11.22

    def _apply_top_alignment(self, worksheet):
        """Apply top alignment and text wrapping to ALL cells in the worksheet."""
        print(f"Setting top alignment for {worksheet.max_row} rows × {worksheet.max_column} columns...")
        
        for row in range(1, worksheet.max_row + 1):  # ALL rows from 1 to max
            for col in range(1, worksheet.max_column + 1):  # ALL columns
                try:
                    cell = worksheet.cell(row=row, column=col)
                    
                    # Always create a new alignment object to avoid style conflicts
                    new_alignment = Alignment(vertical='top')
                    
                    # Enable text wrapping for columns B (Audit Info) and E (Description)
                    if col in [2, 5]:  # Columns B and E
                        new_alignment.wrap_text = True
                    
                    # Apply the new alignment (this overwrites any existing alignment)
                    cell.alignment = new_alignment
                        
                except Exception as e:
                    print(f"Error setting alignment for row {row}, col {col}: {e}")
                    # Continue with next cell instead of stopping
                    continue
        
        print(f"Top alignment applied successfully!")

    def _apply_filters_to_header(self, worksheet):
        """Apply filters to the header row (Row 9) for easy data filtering and sorting."""
        try:
            # Apply filters to Row 9 (header row)
            # Note: openpyxl uses 1-based indexing, so Row 9 is actually row 9
            worksheet.auto_filter.ref = f"A9:L9"
            print(f"Filters applied to header row (Row 9) successfully!")
        except Exception as e:
            print(f"Error applying filters to header row: {e}")
    
    def _apply_alternating_background_colors(self, worksheet, file_matched_df):
        """Apply alternating background colors to matched transaction blocks."""
        try:
            from openpyxl.styles import PatternFill
            
            # Define two alternating colors
            color1 = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")  # Very light blue
            color2 = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")  # Very light lemon yellow
            
            # Get all rows with Match IDs
            match_id_column = file_matched_df.iloc[:, 0]  # First column (Match ID)
            populated_rows = match_id_column.notna()
            
            if not populated_rows.any():
                print("No matched rows found for background coloring")
                return
            
            # Get unique Match IDs in order they appear
            unique_match_ids = []
            seen_ids = set()
            for _, match_id in enumerate(match_id_column):
                if pd.notna(match_id) and match_id not in seen_ids:
                    unique_match_ids.append(match_id)
                    seen_ids.add(match_id)
            
            print(f"Applying alternating background colors to {len(unique_match_ids)} matched transaction blocks")
            
            # Apply alternating colors to each Match ID block
            for block_index, match_id in enumerate(unique_match_ids):
                # Choose color based on block index (alternating)
                color = color1 if block_index % 2 == 0 else color2
                
                # Find all rows with this Match ID
                block_rows = file_matched_df[file_matched_df.iloc[:, 0] == match_id].index
                
                # Apply color to all rows in this block
                for df_row_idx in block_rows:
                    excel_row = df_row_idx + 10  # Convert DataFrame index to Excel row (metadata + header offset)
                    
                    # Color all columns in this row
                    for col in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=excel_row, column=col)
                        cell.fill = color
                
                print(f"  Block {match_id}: Applied {'Color 1' if block_index % 2 == 0 else 'Color 2'} to {len(block_rows)} rows")
            
            print("Background colors applied successfully!")
            
        except Exception as e:
            print(f"Error applying background colors: {e}")

    def _format_output_file_transaction_blocks(self, worksheet):
        """Format output file transaction blocks: make ledger text bold, narration italic, and Entered By person's name bold+italic."""
        try:
            from openpyxl.styles import Font
            
            print("Formatting output file transaction blocks...")
            
            # Create fonts for different formatting
            bold_font = Font(bold=True)
            italic_font = Font(italic=True)
            bold_italic_font = Font(bold=True, italic=True)
            
            # Process all rows starting from row 10 (after metadata and header)
            for row in range(10, worksheet.max_row + 1):
                # Check if this row is the end of a transaction block
                cell_d = worksheet.cell(row=row, column=4)  # Column D (Particulars)
                cell_e = worksheet.cell(row=row, column=5)  # Column E (Description)
                
                # Check if this row contains "Entered By :" in Column D
                if (cell_d.value and 
                    isinstance(cell_d.value, str) and 
                    "Entered By :" in str(cell_d.value)):
                    
                    # This is the end of a transaction block
                    # Make the Entered By person's name bold and italic
                    if cell_e.value:
                        cell_e.font = bold_italic_font
                        print(f"  Row {row}: Made Entered By person's name bold+italic: '{str(cell_e.value)[:50]}...'")
                    
                    # The row above this contains narration text
                    narration_row = row - 1
                    
                    if narration_row >= 10:  # Ensure we don't go below row 10
                        narration_cell_e = worksheet.cell(row=narration_row, column=5)  # Column E
                        
                        # Make the narration text italic
                        if narration_cell_e.value:
                            narration_cell_e.font = italic_font
                            print(f"  Row {narration_row}: Made narration text italic: '{str(narration_cell_e.value)[:50]}...'")
                        
                        # Now find the transaction block start and make all ledger text bold
                        # Look backwards from narration row to find block start
                        for ledger_row in range(narration_row - 1, 9, -1):  # Go back from narration to row 10
                            # Check if this is a block start row
                            date_cell = worksheet.cell(row=ledger_row, column=3)  # Column C (Date)
                            particulars_cell = worksheet.cell(row=ledger_row, column=4)  # Column D (Particulars)
                            vch_type_cell = worksheet.cell(row=ledger_row, column=8)  # Column H (Vch Type)
                            vch_no_cell = worksheet.cell(row=ledger_row, column=9)  # Column I (Vch No)
                            
                            # Check if this is a block start (has date, Dr/Cr, Vch Type, Vch No)
                            is_block_start = (date_cell.value and 
                                            particulars_cell.value and 
                                            str(particulars_cell.value).strip() in ['Dr', 'Cr'] and
                                            vch_type_cell.value and 
                                            vch_no_cell.value)
                            
                            if is_block_start:
                                 # Found block start, now make all rows from here to narration bold
                                 for bold_row in range(ledger_row, narration_row):
                                     bold_cell_e = worksheet.cell(row=bold_row, column=5)  # Column E
                                     if bold_cell_e.value:
                                         bold_cell_e.font = bold_font
                                         print(f"  Row {bold_row}: Made ledger text bold: '{str(bold_cell_e.value)[:50]}...'")
                                 
                                 # Also make Column H (Vch Type) bold in the first row of the transaction block
                                 vch_type_cell = worksheet.cell(row=ledger_row, column=8)  # Column H (Vch Type)
                                 if vch_type_cell.value:
                                     vch_type_cell.font = bold_font
                                     print(f"  Row {ledger_row}: Made Vch Type bold: '{str(vch_type_cell.value)[:50]}...'")
                                 
                                 # Make all Debit and Credit values (Columns J and K) bold in this transaction block
                                 for bold_row in range(ledger_row, narration_row):
                                     # Make Debit column (J) bold
                                     debit_cell = worksheet.cell(row=bold_row, column=10)  # Column J (Debit)
                                     if debit_cell.value and debit_cell.value != '':
                                         debit_cell.font = bold_font
                                         print(f"  Row {bold_row}: Made Debit value bold: '{str(debit_cell.value)[:20]}...'")
                                     
                                     # Make Credit column (K) bold
                                     credit_cell = worksheet.cell(row=bold_row, column=11)  # Column K (Credit)
                                     if credit_cell.value and credit_cell.value != '':
                                         credit_cell.font = bold_font
                                         print(f"  Row {bold_row}: Made Credit value bold: '{str(credit_cell.value)[:20]}...'")
                                 
                                 break
            
            # Also check for "Opening Balance" text and make it bold along with its Debit/Credit values
            print("Checking for Opening Balance entries...")
            for row in range(10, worksheet.max_row + 1):
                cell_e = worksheet.cell(row=row, column=5)  # Column E (Description)
                
                # Check if this row contains "Opening Balance" text
                if (cell_e.value and 
                    isinstance(cell_e.value, str) and 
                    "Opening Balance" in str(cell_e.value)):
                    
                    # Make the Opening Balance text bold
                    cell_e.font = bold_font
                    print(f"  Row {row}: Made Opening Balance text bold: '{str(cell_e.value)[:50]}...'")
                    
                    # Make the associated Debit and Credit values bold
                    debit_cell = worksheet.cell(row=row, column=10)  # Column J (Debit)
                    if debit_cell.value and debit_cell.value != '':
                        debit_cell.font = bold_font
                        print(f"  Row {row}: Made Opening Balance Debit value bold: '{str(debit_cell.value)[:20]}...'")
                    
                    credit_cell = worksheet.cell(row=row, column=11)  # Column K (Credit)
                    if credit_cell.value and credit_cell.value != '':
                        credit_cell.font = bold_font
                        print(f"  Row {row}: Made Opening Balance Credit value bold: '{str(credit_cell.value)[:20]}...'")
            
            print("Output file transaction block formatting completed successfully!")
            
        except Exception as e:
            print(f"Error formatting output file transaction blocks: {e}")

    def create_matched_files(self, matches, transactions1, transactions2):
        """Create matched versions of both files with new columns."""
        if not matches:
            print("No matches found. Cannot create matched files.")
            return
        
        # Matches are already in sequential order from post-processing step
        print(f"\n=== USING SEQUENTIALLY ASSIGNED MATCHES ===")
        print(f"Total matches: {len(matches)}")
        print(f"First 10 Match IDs: {[m['match_id'] for m in matches[:10]]}")

        print(f"Last 10 Match IDs: {[m['match_id'] for m in matches[-10:]]}")
        
        # Create file1 with new columns
        file1_matched = transactions1.copy()
        
        # Create new columns with proper names
        match_id_col = pd.Series([None] * len(file1_matched), name='Match ID')
        audit_info_col = pd.Series([None] * len(file1_matched), name='Audit Info')
        match_type_col = pd.Series([None] * len(file1_matched), name='Match Type')
        
        # Concatenate new columns with existing data
        file1_matched = pd.concat([match_id_col, audit_info_col, file1_matched, match_type_col], axis=1)
        

        
        # Create file2 with new columns
        file2_matched = transactions2.copy()
        
        # Create new columns with proper names
        match_id_col2 = pd.Series([None] * len(file2_matched), name='Match ID')
        audit_info_col2 = pd.Series([None] * len(file2_matched), name='Audit Info')
        match_type_col2 = pd.Series([None] * len(file2_matched), name='Match Type')
        
        # Concatenate new columns with existing data
        file2_matched = pd.concat([match_id_col2, audit_info_col2, file2_matched, match_type_col2], axis=1)
        
        # print(f"DEBUG: File2 DataFrame created with shape: {file2_matched.shape}")
        # print(f"DEBUG: File2 columns: {list(file2_matched.columns)}")
        
        # print(f"DEBUG: Added Match Type column to both DataFrames")
        # print(f"DEBUG: File1 columns: {list(file1_matched.columns)}")
        # print(f"DEBUG: File2 columns: {list(file2_matched.columns)}")
        
        # Verify the new columns are actually there
        # print(f"DEBUG: File1 first few rows of Match ID column:")
        # print(file1_matched.iloc[:5, 0].tolist())
        # print(f"DEBUG: File1 first few rows of Audit Info column:")
        # print(file1_matched.iloc[:5, 1].tolist())
        
        # Populate match information - process matches in sequential order
        for i, match in enumerate(matches):
            match_id = match['match_id']  # Use the pre-assigned match ID
            audit_info = self.create_audit_info(match)
            

            # Use the explicit Match_Type field if available, otherwise fall back to inference
            if 'Match_Type' in match and match['Match_Type']:
                match_type = match['Match_Type']
                print(f"  Match Type: {match_type} (from explicit field)")
            elif 'LC_Number' in match and match['LC_Number']:
                print(f"  LC Number: {match['LC_Number']}")
                match_type = 'LC'
            elif 'PO_Number' in match and match['PO_Number']:
                print(f"  PO Number: {match['PO_Number']}")
                match_type = 'PO'
            elif 'Interunit_Account' in match and match['Interunit_Account']:
                print(f"  Interunit Account: {match['Interunit_Account']}")
                match_type = 'Interunit'
            else:
                print(f"  Unknown Match Type")
                match_type = 'Unknown'
            print(f"  File1 Row {match['File1_Index']}: Debit={match['File1_Debit']}, Credit={match['File1_Credit']}")
            print(f"  File2 Row {match['File2_Index']}: Debit={match['File2_Debit']}, Credit={match['File2_Credit']}")
            print(f"  Audit Info: {audit_info}")
            print(f"  Match Type: {match_type}")
            
            # Update file1 - populate entire transaction block with Match ID and Audit Info
            file1_row_idx = match['File1_Index']
            # Find the entire transaction block for file1 and populate all rows
            file1_block_rows = self.block_identifier.get_transaction_block_rows(file1_row_idx, self.file1_path)
            
            # Populate ALL rows of the transaction block with Match ID and Match Type, but only if not already set (preserve first/lowest Match ID)
            for i, block_row in enumerate(file1_block_rows):
                if 0 <= block_row < len(file1_matched):
                    # Only set Match ID if not already set (preserve first/lowest Match ID)
                    if pd.isna(file1_matched.iloc[block_row, 0]) or file1_matched.iloc[block_row, 0] == '':
                        file1_matched.iloc[block_row, 0] = match_id  # Match ID column (index 0)
                        file1_matched.iloc[block_row, -1] = match_type  # Match Type column (last column) - ALL ROWS
                    else:
                        print(f"      WARNING: Row {block_row} already has Match ID {file1_matched.iloc[block_row, 0]}, preserving it instead of {match_id}")
                    
                    # Audit Info goes ONLY in the second-to-last row of the transaction block
                    if i == len(file1_block_rows) - 2:  # Second-to-last row
                        file1_matched.iloc[block_row, 1] = audit_info  # Audit Info column (index 1)
            

            
            # Update file2 - populate entire transaction block with Match ID and Audit Info
            file2_row_idx = match['File2_Index']
            # Find the entire transaction block for file2 and populate all rows
            file2_block_rows = self.block_identifier.get_transaction_block_rows(file2_row_idx, self.file2_path)
            
            # Populate ALL rows of the transaction block with Match ID and Match Type, but only if not already set (preserve first/lowest Match ID)
            for i, block_row in enumerate(file2_block_rows):
                if 0 <= block_row < len(file2_matched):
                    # Only set Match ID if not already set (preserve first/lowest Match ID)
                    if pd.isna(file2_matched.iloc[block_row, 0]) or file2_matched.iloc[block_row, 0] == '':
                        file2_matched.iloc[block_row, 0] = match_id  # Match ID column (index 0)
                        file2_matched.iloc[block_row, -1] = match_type  # Match Type column (last column) - ALL ROWS
                    else:
                        print(f"      WARNING: Row {block_row} already has Match ID {file2_matched.iloc[block_row, 0]}, preserving it instead of {match_id}")
                    
                    # Audit Info goes ONLY in the second-to-last row of the transaction block
                    if i == len(file2_block_rows) - 2:  # Second-to-last row
                        file2_matched.iloc[block_row, 1] = audit_info  # Audit Info column (index 1)
        
        # Add unmatched audit info for records that didn't match
        print(f"\n=== ADDING UNMATCHED AUDIT INFO ===")
        try:
            from unmatched_tracker import get_unmatched_tracker
            unmatched_tracker = get_unmatched_tracker()
            
            # Process File 1 unmatched records
            for idx in range(len(file1_matched)):
                # Skip if already has Match ID (matched record)
                if pd.notna(file1_matched.iloc[idx, 0]) and file1_matched.iloc[idx, 0] != '':
                    continue
                
                # Get audit info for this unmatched record
                audit_info = unmatched_tracker.get_audit_info_for_unmatched(idx, 1)
                if audit_info:
                    # Find the transaction block for this row
                    try:
                        block_rows = self.block_identifier.get_transaction_block_rows(idx, self.file1_path)
                        if block_rows:
                            # Put audit info in second-to-last row of block
                            second_to_last_idx = len(block_rows) - 2
                            if 0 <= second_to_last_idx < len(block_rows):
                                block_row = block_rows[second_to_last_idx]
                                if 0 <= block_row < len(file1_matched):
                                    # Only set if not already set
                                    if pd.isna(file1_matched.iloc[block_row, 1]) or file1_matched.iloc[block_row, 1] == '':
                                        file1_matched.iloc[block_row, 1] = audit_info
                    except Exception as e:
                        print(f"  Warning: Could not add unmatched audit info for File 1 row {idx}: {e}")
            
            # Process File 2 unmatched records
            for idx in range(len(file2_matched)):
                # Skip if already has Match ID (matched record)
                if pd.notna(file2_matched.iloc[idx, 0]) and file2_matched.iloc[idx, 0] != '':
                    continue
                
                # Get audit info for this unmatched record
                audit_info = unmatched_tracker.get_audit_info_for_unmatched(idx, 2)
                if audit_info:
                    # Find the transaction block for this row
                    try:
                        block_rows = self.block_identifier.get_transaction_block_rows(idx, self.file2_path)
                        if block_rows:
                            # Put audit info in second-to-last row of block
                            second_to_last_idx = len(block_rows) - 2
                            if 0 <= second_to_last_idx < len(block_rows):
                                block_row = block_rows[second_to_last_idx]
                                if 0 <= block_row < len(file2_matched):
                                    # Only set if not already set
                                    if pd.isna(file2_matched.iloc[block_row, 1]) or file2_matched.iloc[block_row, 1] == '':
                                        file2_matched.iloc[block_row, 1] = audit_info
                    except Exception as e:
                        print(f"  Warning: Could not add unmatched audit info for File 2 row {idx}: {e}")
            
            print(f"  Added unmatched audit info to output files")
        except ImportError:
            print(f"  Warning: unmatched_tracker not available, skipping unmatched audit info")
        except Exception as e:
            print(f"  Warning: Error adding unmatched audit info: {e}")
        
        # Save matched files using configuration variables
        base_name1 = os.path.splitext(os.path.basename(self.file1_path))[0]
        base_name2 = os.path.splitext(os.path.basename(self.file2_path))[0]
        
        # Get the directory of the input files
        input_dir1 = os.path.dirname(self.file1_path)
        input_dir2 = os.path.dirname(self.file2_path)
        
        output_file1 = os.path.join(input_dir1, f"{base_name1}{OUTPUT_SUFFIX}")
        output_file2 = os.path.join(input_dir2, f"{base_name2}{OUTPUT_SUFFIX}")
        
        print(f"\n=== OUTPUT FILE LOCATIONS ===")
        print(f"Input File 1: {self.file1_path}")
        print(f"Input Directory 1: {input_dir1}")
        print(f"Output File 1: {output_file1}")
        print(f"Input File 2: {self.file2_path}")
        print(f"Input Directory 2: {input_dir2}")
        print(f"Output File 2: {output_file2}")
        
        if VERBOSE_DEBUG:
            print(f"File1 - Rows with Match IDs: {file1_matched.iloc[:, 0].notna().sum()}")
            print(f"File1 - Rows with Audit Info: {file1_matched.iloc[:, 1].notna().sum()}")
            print(f"File1 - Rows with Match Type: {file1_matched.iloc[:, -1].notna().sum()}")
            print(f"File2 - Rows with Match IDs: {file2_matched.iloc[:, 0].notna().sum()}")
            print(f"File2 - Rows with Audit Info: {file2_matched.iloc[:, 1].notna().sum()}")
            print(f"File2 - Rows with Match Type: {file2_matched.iloc[:, -1].notna().sum()}")
            
            # Get the actual populated rows dynamically
            populated_rows = file1_matched.iloc[:, 0].notna()
            if populated_rows.any():
                populated_indices = file1_matched[populated_rows].index
                for idx in populated_indices[:4]:  # Show first 4 populated rows
                    print(f"File1 - Row {idx} Match ID: '{file1_matched.iloc[idx, 0]}'")
                    print(f"File1 - Row {idx} Audit Info: '{file1_matched.iloc[idx, 1]}'")
            else:
                print("No populated rows found in File1")
        
        # Convert date columns to strings in dd/mmm/yyyy format to prevent timestamp display
        print("\n=== CONVERTING DATES TO dd/mmm/yyyy FORMAT ===")
        from datetime import datetime
        
        def convert_date_to_string(date_value):
            """Convert date value to dd/mmm/yyyy string format."""
            if pd.isna(date_value) or date_value == '':
                return date_value
            try:
                # If it's already a datetime object, format it
                if isinstance(date_value, datetime):
                    return date_value.strftime('%d/%b/%Y')
                # If it's a string with timestamp, parse and reformat
                elif isinstance(date_value, str):
                    if ' ' in date_value or 'T' in date_value:
                        # Has timestamp, parse and reformat
                        parsed_date = pd.to_datetime(date_value)
                        return parsed_date.strftime('%d/%b/%Y')
                    # Already in correct format or can't parse, return as-is
                    return date_value
                # Try to convert to datetime if it's a pandas Timestamp or other date-like object
                else:
                    parsed_date = pd.to_datetime(date_value)
                    return parsed_date.strftime('%d/%b/%Y')
            except:
                # If any conversion fails, return as string representation
                return str(date_value) if date_value else date_value
        
        # Date column is at index 2 (after Match ID at 0 and Audit Info at 1)
        date_col_idx = 2
        
        # Convert dates in file1_matched using vectorized operation
        if date_col_idx < len(file1_matched.columns):
            date_col_name = file1_matched.columns[date_col_idx]
            file1_matched[date_col_name] = file1_matched[date_col_name].apply(convert_date_to_string)
        
        # Convert dates in file2_matched using vectorized operation
        if date_col_idx < len(file2_matched.columns):
            date_col_name = file2_matched.columns[date_col_idx]
            file2_matched[date_col_name] = file2_matched[date_col_name].apply(convert_date_to_string)
        
        # Create output with metadata + matched transactions
        with pd.ExcelWriter(output_file1, engine='openpyxl') as writer:
            # Write metadata
            self.metadata1.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            # Write matched transactions (skip first 8 rows to make room for metadata)
            file1_matched.to_excel(writer, sheet_name='Sheet1', index=False, header=True, startrow=8)
            
            # Get the worksheet to set column widths
            worksheet = writer.sheets['Sheet1']
            self._set_column_widths(worksheet)
            self._format_amount_columns(worksheet) # Apply amount formatting
            self._format_date_columns(worksheet) # Apply date formatting (dd/mmm/yyyy)
            
            # Apply top alignment and text wrapping to ALL cells in the worksheet
            self._apply_top_alignment(worksheet)
            
            # Apply filters to the header row for easy data filtering and sorting
            self._apply_filters_to_header(worksheet)
            
            # Apply alternating background colors to matched transaction blocks
            self._apply_alternating_background_colors(worksheet, file1_matched)
            
            # Format output file transaction blocks (make narration italic)
            self._format_output_file_transaction_blocks(worksheet)
            
                    
        with pd.ExcelWriter(output_file2, engine='openpyxl') as writer:
            # Write metadata
            self.metadata2.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            # Write matched transactions (skip first 8 rows to make room for metadata)
            file2_matched.to_excel(writer, sheet_name='Sheet1', index=False, header=True, startrow=8)
            
            # Get the worksheet to set column widths
            worksheet = writer.sheets['Sheet1']
            self._set_column_widths(worksheet)
            self._format_amount_columns(worksheet) # Apply amount formatting
            self._format_date_columns(worksheet) # Apply date formatting (dd/mmm/yyyy)
            
            # Apply top alignment and text wrapping to ALL cells in the worksheet
            self._apply_top_alignment(worksheet)
            
            # Apply filters to the header row for easy data filtering and sorting
            self._apply_filters_to_header(worksheet)
            
            # Apply alternating background colors to matched transaction blocks
            self._apply_alternating_background_colors(worksheet, file2_matched)
            
            # Format output file transaction blocks (make narration italic)
            self._format_output_file_transaction_blocks(worksheet)
            
        
        # Also create a simple version without metadata to test (if enabled)
        if CREATE_SIMPLE_FILES:
            simple_output1 = os.path.join(input_dir1, f"{base_name1}{SIMPLE_SUFFIX}")
            simple_output2 = os.path.join(input_dir2, f"{base_name2}{SIMPLE_SUFFIX}")
            
            print(f"\nCreating simple test files without metadata...")
            file1_matched.to_excel(simple_output1, index=False, header=True)
            file2_matched.to_excel(simple_output2, index=False, header=True)
            
            print(f"Created simple test files:")
            print(f"  {simple_output1}")
            print(f"  {simple_output2}")
        

        
        print(f"Checking if files were actually written...")
        
        # Verify the files were written correctly
        try:
            df_check1 = pd.read_excel(output_file1, header=8)
            print(f"File1 loaded successfully, shape: {df_check1.shape}")
            print(f"File1 - Rows with Match IDs: {df_check1.iloc[:, 0].notna().sum()}")
            print(f"File1 - Rows with Audit Info: {df_check1.iloc[:, 1].notna().sum()}")
            print(f"File1 - Rows with Match Type: {df_check1.iloc[:, -1].notna().sum()}")
            
            # Check if text wrapping was applied by reading the Excel file with openpyxl
            print(f"\n=== VERIFYING TEXT WRAPPING IN FILE 1 ===")
            wb1 = openpyxl.load_workbook(output_file1)
            ws1 = wb1.active
            print(f"Worksheet: {ws1.title}")
            print(f"Max row: {ws1.max_row}, Max column: {ws1.max_column}")
            
            # Check a few cells in columns B and E for text wrapping
            for row in range(9, min(15, ws1.max_row + 1)):
                cell_b = ws1.cell(row=row, column=2)
                cell_e = ws1.cell(row=row, column=5)
                print(f"Row {row}:")
                print(f"  Column B: value='{cell_b.value}', wrap_text={cell_b.alignment.wrap_text if cell_b.alignment else 'None'}")
                print(f"  Column E: value='{cell_e.value}', wrap_text={cell_e.alignment.wrap_text if cell_e.alignment else 'None'}")
            
            wb1.close()
            
        except Exception as e:
            print(f"Error reading File1: {e}")
        
        try:
            df_check2 = pd.read_excel(output_file2, header=8)
            print(f"File2 loaded successfully, shape: {df_check2.shape}")
            print(f"File2 - Rows with Match IDs: {df_check2.iloc[:, 0].notna().sum()}")
            print(f"File2 - Rows with Audit Info: {df_check2.iloc[:, 1].notna().sum()}")
            print(f"File2 - Rows with Match Type: {df_check2.iloc[:, -1].notna().sum()}")
            
            # Check if text wrapping was applied by reading the Excel file with openpyxl
            print(f"\n=== VERIFYING TEXT WRAPPING IN FILE 2 ===")
            wb2 = openpyxl.load_workbook(output_file2)
            ws2 = wb2.active
            print(f"Worksheet: {ws2.title}")
            print(f"Max row: {ws2.max_row}, Max column: {ws2.max_column}")
            
            # Check a few cells in columns B and E for text wrapping
            for row in range(9, min(15, ws2.max_row + 1)):
                cell_b = ws2.cell(row=row, column=2)
                cell_e = ws2.cell(row=row, column=5)
                print(f"Row {row}:")
                print(f"  Column B: value='{cell_b.value}', wrap_text={cell_b.alignment.wrap_text if cell_b.alignment else 'None'}")
                print(f"  Column E: value='{cell_e.value}', wrap_text={cell_e.alignment.wrap_text if cell_e.alignment else 'None'}")
            
            wb2.close()
            
        except Exception as e:
            print(f"Error reading File2: {e}")
        
        print(f"\nCreated matched files:")
        print(f"  {output_file1}")
        print(f"  {output_file2}")
        
        # Verify the data was populated correctly
        self.verify_match_data(file1_matched, file2_matched)
    

    
    def verify_match_data(self, file1_matched, file2_matched):
        """Verify that Match ID and Audit Info columns are properly populated."""
        print(f"\n=== VERIFICATION RESULTS ===")
        
        # Check Match ID column population - only count non-empty, non-NaN values
        match_ids_1 = file1_matched.iloc[:, 0].replace('', None).dropna()
        match_ids_2 = file2_matched.iloc[:, 0].replace('', None).dropna()
        
        print(f"File 1 - Match IDs populated: {len(match_ids_1)}")
        print(f"File 2 - Match IDs populated: {len(match_ids_2)}")
        
        # Check Audit Info column population - only count non-empty, non-NaN values
        audit_info_1 = file1_matched.iloc[:, 1].replace('', None).dropna()
        audit_info_2 = file2_matched.iloc[:, 1].replace('', None).dropna()
        
        print(f"File 1 - Audit Info populated: {len(audit_info_1)}")
        print(f"File 2 - Audit Info populated: {len(audit_info_2)}")
        
        # Check Match Type column population - only count non-empty, non-NaN values
        match_types_1 = file1_matched.iloc[:, -1].replace('', None).dropna()
        match_types_2 = file2_matched.iloc[:, -1].replace('', None).dropna()
        
        print(f"File 1 - Match Types populated: {len(match_types_1)}")
        print(f"File 2 - Match Types populated: {len(match_types_2)}")
        
        # Show sample populated data
        if len(match_ids_1) > 0:
            print(f"\nFile 1 - Sample populated rows:")
            for _, match_id in enumerate(match_ids_1[:3]):
                # Find rows that actually have this match_id (not empty strings)
                mask = (file1_matched.iloc[:, 0] == match_id) & (file1_matched.iloc[:, 0] != '')
                if mask.any():
                    row_idx = file1_matched[mask].index[0]
                    print(f"  Row {row_idx}: Match ID = {match_id}")
                    print(f"    Date: {file1_matched.iloc[row_idx, 3]}")
                    print(f"    Description: {str(file1_matched.iloc[row_idx, 4])[:50]}...")
                    print(f"    Debit: {file1_matched.iloc[row_idx, 10]}, Credit: {file1_matched.iloc[row_idx, 11]}")
                    print(f"    Match Type: {file1_matched.iloc[row_idx, -1]}")
        
        if len(match_ids_2) > 0:
            print(f"\nFile 2 - Sample populated rows:")
            for i, match_id in enumerate(match_ids_2[:3]):
                # Find rows that actually have this match_id (not empty strings)
                mask = (file2_matched.iloc[:, 0] == match_id) & (file2_matched.iloc[:, 0] != '')
                if mask.any():
                    row_idx = file2_matched[mask].index[0]
                    print(f"  Row {row_idx}: Match ID = {match_id}")
                    print(f"    Date: {file2_matched.iloc[row_idx, 3]}")
                    print(f"    Description: {str(file2_matched.iloc[row_idx, 4])[:50]}...")
                    print(f"    Debit: {file2_matched.iloc[row_idx, 10]}, Credit: {file2_matched.iloc[row_idx, 11]}")
                    print(f"    Match Type: {file2_matched.iloc[row_idx, -1]}")

    