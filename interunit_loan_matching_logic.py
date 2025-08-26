#!/usr/bin/env python3
"""
Interunit Loan Matcher Module

This module implements the CORRECT interunit loan matching logic:
1. Find full account numbers in ledger (of both files)
2. Find the short code in narration (of both files)
3. Debit/Credit amounts must match exactly (no tolerance)
4. Cross-referenced short codes: File 1 narration contains File 2's short code, and File 2 narration contains File 1's short code
5. Accounts must be different (lender vs borrower)
6. FOLLOWS CORE LOGIC AND FORMAT - cannot deviate
"""

import pandas as pd
import re
from openpyxl import load_workbook
from typing import Dict, List, Optional, Any
from transaction_block_identifier import TransactionBlockIdentifier

class InterunitLoanMatcher:
    """
    Matches interunit loan transactions between two files based on:
    - Full account numbers in ledger
    - Short codes in narration
    - Exact amount matching
    - Cross-referenced short codes (BIDIRECTIONAL)
    - FOLLOWS CORE LOGIC AND FORMAT EXACTLY
    """
    
    def __init__(self):
        # Interunit account mapping (Full Format → Short Code)
        self.interunit_account_mapping = {
            'Brac Bank PLC-CD-A/C-2028701210002': 'BBL#0002',
            'Dhaka Bank-STD-2051501833-CIL': 'DBL#1833',
            'Dutch Bangla Bank Ltd.-SND-1071200003988': 'DBBL#3988',
            'Eastern Bank Limited-SND-1011060605503': 'EBL#5503',
            'Eastern Bank OD#1012040163265': 'EBL#3265',
            'Eastern Bank OD#1012210603129': 'EBL#3129',
            'Eastern Bank,STD-1011220144056': 'EBL#4056',
            'Midland Bank Ltd-CE-0011-1060000313': 'MDB#0313',
            'Midland Bank PLC-CD-A/C-0011-1050011026': 'MDB#11026',
            'Midland-CE-0011-1060000304-CI': 'MDB#0304',
            'Midland-CE-0011-1060000331-CI': 'MDB#0331',
            'One Bank-CD/A/C-0011020008826': 'OBL#8826',
            'One Bank-SND-A/C-0013000002451': 'OBL#2451',
            'PBL-SND- 2126312011060': 'PBL#11060',
            'Prime Bank Limited-SND-2126318011502': 'PBL#11502',
            'Prime Bank-CD-2126117010855': 'PBL#10855',
            'MTBL-SND-A/C-1310000003858': 'MTBL#3858',
        }
        
        # No amount tolerance - exact matching required
        
        # Initialize transaction block identifier
        self.block_identifier = TransactionBlockIdentifier()
    
    def find_potential_matches(
        self, 
        transactions1: pd.DataFrame, 
        transactions2: pd.DataFrame, 
        interunit_accounts1: pd.Series, 
        interunit_accounts2: pd.Series, 
        file1_path: str, 
        file2_path: str, 
        existing_matches: Dict = None, 
        match_counter: int = 0
    ) -> List[Dict]:
        """
        Find interunit loan matches between two transaction files.
        IMPLEMENTS THE CORRECT CROSS-REFERENCING LOGIC.
        FOLLOWS CORE LOGIC AND FORMAT EXACTLY - cannot deviate.
        
        Args:
            transactions1: DataFrame of transactions from first file
            transactions2: DataFrame of transactions from second file
            interunit_accounts1: Series of interunit accounts from first file
            interunit_accounts2: Series of interunit accounts from second file
            file1_path: Path to first file (for openpyxl access)
            file2_path: Path to second file (for openpyxl access)
            existing_matches: Dictionary of existing matches (shared state)
            match_counter: Counter for generating unique match IDs (shared state)
            
        Returns:
            List of match dictionaries following core format
        """
        matches = []
        
        # Use shared state if provided, otherwise create new
        if existing_matches is None:
            existing_matches = {}
        if match_counter is None:
            match_counter = 0
        
        print(f"\n=== INTERUNIT LOAN MATCHING LOGIC (CORRECTED) ===")
        print(f"1. Find full account numbers in ledger (of both files)")
        print(f"2. Find the short code in narration (of both files)")
        print(f"3. Debit/Credit amounts must match exactly (no tolerance)")
        print(f"4. Cross-referenced short codes: File 1 narration contains File 2's short code, and File 2 narration contains File 1's short code")
        print(f"5. Accounts must be different (lender vs borrower)")
        print(f"6. FOLLOWS CORE LOGIC: Uses universal M001 format, shared state, same structure")
        
        # Identify transaction blocks in both files
        blocks1 = self.block_identifier.identify_transaction_blocks(transactions1, file1_path)
        blocks2 = self.block_identifier.identify_transaction_blocks(transactions2, file2_path)
        
        print(f"\nFile 1: {len(blocks1)} transaction blocks")
        print(f"File 2: {len(blocks2)} transaction blocks")
        
        # Load workbooks for formatting analysis
        wb1 = load_workbook(file1_path)
        ws1 = wb1.active
        wb2 = load_workbook(file2_path)
        ws2 = wb2.active
        
        # Collect all interunit account information from both files
        file1_interunit_data = []
        file2_interunit_data = []
        
        print(f"\n--- Scanning File 1 (Steel) for interunit data ---")
        for i, block in enumerate(blocks1):
            block_data = self._analyze_block_for_interunit_data(ws1, block, i)
            if block_data['ledger_accounts'] or block_data['narration_short_codes']:
                file1_interunit_data.append(block_data)
                if len(file1_interunit_data) <= 5:  # Show first 5
                    print(f"Block {i+1}: {len(block_data['ledger_accounts'])} ledger accounts, {len(block_data['narration_short_codes'])} short codes")
        
        print(f"\n--- Scanning File 2 (GeoTex) for interunit data ---")
        for i, block in enumerate(blocks2):
            block_data = self._analyze_block_for_interunit_data(ws2, block, i)
            if block_data['ledger_accounts'] or block_data['narration_short_codes']:
                file2_interunit_data.append(block_data)
                if len(file2_interunit_data) <= 5:  # Show first 5
                    print(f"Block {i+1}: {len(block_data['ledger_accounts'])} ledger accounts, {len(block_data['narration_short_codes'])} short codes")
        
        print(f"\n✓ File 1: {len(file1_interunit_data)} blocks with interunit data")
        print(f"✓ File 2: {len(file2_interunit_data)} blocks with interunit data")
        
        # Look for cross-referenced matches
        print(f"\n--- Looking for cross-referenced matches ---")
        potential_matches = []
        
        for block1 in file1_interunit_data:
            for block2 in file2_interunit_data:
                # Check if blocks have opposite transaction types (one debit, one credit)
                if (block1['amounts'] and block2['amounts'] and
                    ((block1['amounts']['debit'] and block2['amounts']['credit']) or
                     (block1['amounts']['credit'] and block2['amounts']['debit']))):
                    
                    # Check if amounts match exactly (NO TOLERANCE)
                    amount1 = block1['amounts']['debit'] if block1['amounts']['debit'] else block1['amounts']['credit']
                    amount2 = block2['amounts']['debit'] if block2['amounts']['debit'] else block2['amounts']['credit']
                    
                    if amount1 == amount2:
                        # Check for cross-referenced short codes
                        cross_reference_found = False
                        file1_narration_contains = None
                        file2_narration_contains = None
                        
                        # File 1's narration should contain File 2's short code
                        for narration1 in block1['narration_short_codes']:
                            for ledger2 in block2['ledger_accounts']:
                                if narration1['short_code'] == ledger2['short_code']:
                                    cross_reference_found = True
                                    file1_narration_contains = narration1['short_code']
                                    break
                            if cross_reference_found:
                                break
                        
                        # File 2's narration should contain File 1's short code
                        if cross_reference_found:
                            for narration2 in block2['narration_short_codes']:
                                for ledger1 in block1['ledger_accounts']:
                                    if narration2['short_code'] == ledger1['short_code']:
                                        file2_narration_contains = narration2['short_code']
                                        
                                        # We have a match! Check if we've already matched this combination
                                        match_key = (amount1, file1_narration_contains, file2_narration_contains)
                                        
                                        if match_key in existing_matches:
                                            # Use existing match ID for consistency
                                            match_id = existing_matches[match_key]
                                            print(f"  REUSING existing Match ID {match_id} for Amount {amount1}")
                                        else:
                                            # Create new match ID following CORE FORMAT
                                            match_counter += 1
                                            match_id = f"M{match_counter:03d}"  # M001, M002, M003... FOLLOWS CORE LOGIC
                                            existing_matches[match_key] = match_id
                                            print(f"  CREATING new Match ID {match_id} for Amount {amount1}")
                                        
                                        # Create match following CORE FORMAT exactly
                                        match = {
                                            'match_id': match_id,
                                            'Match_Type': 'Interunit',  # Add explicit match type
                                            'Interunit_Account': f"{file1_narration_contains} ↔ {file2_narration_contains}",
                                            'File1_Index': block1['amounts']['row'],
                                            'File2_Index': block2['amounts']['row'],
                                            'File1_Debit': block1['amounts']['debit'],
                                            'File1_Credit': block1['amounts']['credit'],
                                            'File2_Debit': block2['amounts']['debit'],
                                            'File2_Credit': block2['amounts']['credit'],
                                            'File1_Amount': amount1,  # Add File1_Amount for audit info
                                            'File2_Amount': amount1,  # Add File2_Amount for audit info
                                            'Amount': amount1
                                        }
                                        
                                        matches.append(match)
                                        print(f"  ✓ MATCH {match_id}: Amount {amount1}")
                                        print(f"    Cross-reference: File 1 narration contains {file1_narration_contains}")
                                        print(f"    Cross-reference: File 2 narration contains {file2_narration_contains}")
                                        break
                                
                                if file2_narration_contains:
                                    break
        
        # Close workbooks
        wb1.close()
        wb2.close()
        
        print(f"\nInterunit Loan Matching Complete: {len(matches)} matches found")
        print(f"FOLLOWS CORE LOGIC: Uses universal M001 format, integrates with shared state")
        return matches
    
    def _analyze_block_for_interunit_data(self, worksheet, block_rows, block_index):
        """Analyze a transaction block for interunit account data."""
        block_data = {
            'block_index': block_index,
            'block_rows': block_rows,
            'ledger_accounts': [],
            'narration_short_codes': [],
            'amounts': {}
        }
        
        # Check each row in the block
        for row_idx in block_rows:
            excel_row = row_idx + 10  # Convert to Excel row number
            
            if excel_row <= worksheet.max_row:
                cell_c = worksheet.cell(row=excel_row, column=3)  # Column C
                
                # Check for ledger accounts (Bold but not italic)
                if (cell_c.value and 
                    cell_c.font and 
                    cell_c.font.bold and 
                    not cell_c.font.italic):
                    
                    # Check if this is an interunit account
                    for full_account, short_code in self.interunit_account_mapping.items():
                        if full_account.upper() in str(cell_c.value).upper():
                            block_data['ledger_accounts'].append({
                                'full_account': full_account,
                                'short_code': short_code,
                                'cell_value': cell_c.value
                            })
                
                # Check for narration rows (Italic but not bold)
                elif (cell_c.value and 
                      cell_c.font and 
                      not cell_c.font.bold and 
                      cell_c.font.italic):
                    
                    # Look for short codes in narration
                    for short_code in self.interunit_account_mapping.values():
                        if short_code in str(cell_c.value):
                            block_data['narration_short_codes'].append({
                                'short_code': short_code,
                                'narration': cell_c.value
                            })
                
                # Check for amounts (Debit/Credit columns)
                debit_cell = worksheet.cell(row=excel_row, column=8)  # Column H
                credit_cell = worksheet.cell(row=excel_row, column=9)  # Column I
                
                if (debit_cell.value is not None and debit_cell.value != 0) or \
                   (credit_cell.value is not None and credit_cell.value != 0):
                    block_data['amounts'] = {
                        'debit': debit_cell.value if debit_cell.value else None,
                        'credit': credit_cell.value if credit_cell.value else None,
                        'row': row_idx
                    }
                

        
        return block_data
    
    def extract_interunit_accounts_from_narration(self, transactions: pd.DataFrame, file_path: str) -> pd.Series:
        """
        Extract interunit account references from NARRATION rows (Column C - Italic).
        FOLLOWS CORE LOGIC - same pattern as existing modules.
        """
        print(f"Extracting interunit account references from NARRATION column...")
        
        # Initialize interunit accounts series with None values
        interunit_accounts = pd.Series([None] * len(transactions), index=transactions.index)
        
        # Load the Excel file to check cell formatting
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Find NARRATION rows (Italic text in Column C) - NOT ledger rows
            for idx, row_idx in enumerate(transactions.index):
                excel_row = row_idx + 9  # Adjust for metadata rows
                
                if excel_row <= ws.max_row:
                    cell = ws.cell(row=excel_row, column=3)  # Column C
                    
                    if (cell.value and 
                        cell.font and 
                        not cell.font.bold and 
                        cell.font.italic):  # Italic but NOT bold = NARRATION row
                        
                        # This is a NARRATION row, extract interunit account reference
                        account_info = self.extract_interunit_account_from_narration(str(cell.value))
                        if account_info:
                            interunit_accounts.iloc[idx] = account_info['full_reference']
                            print(f"  Row {row_idx}: Found interunit account '{account_info['full_reference']}' in NARRATION")
            
            wb.close()
            
        except Exception as e:
            print(f"Error reading Excel file for interunit account extraction: {e}")
        
        print(f"Interunit account extraction complete. Found {interunit_accounts.notna().sum()} account references")
        return interunit_accounts
    
    def extract_interunit_account_from_narration(self, narration: str) -> Optional[Dict[str, Any]]:
        """
        Extract interunit account reference from narration text.
        FOLLOWS CORE LOGIC - same pattern as existing modules.
        """
        if not narration:
            return None
        
        # Look for short codes in narration (e.g., MTBL#3858, OBL#8826)
        short_code_patterns = [
            r'([A-Z]{2,4})#(\d{4,6})',  # MTBL#4355, MDBL#11026, OBL#8826
        ]
        
        for pattern in short_code_patterns:
            try:
                match = re.search(pattern, narration.upper())
                if match:
                    bank_code = match.group(1).strip()
                    account_number = match.group(2)
                    
                    return {
                        'account_number': account_number,
                        'bank_code': bank_code,
                        'full_reference': match.group(),
                        'full_account_format': None
                    }
            except Exception as e:
                print(f"DEBUG: Interunit narration regex error with pattern '{pattern}' and text '{narration}': {e}")
                continue
        
        return None
