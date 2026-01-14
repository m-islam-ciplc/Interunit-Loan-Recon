"""
Manual Matching Window

GUI dialog for manually reviewing and confirming/rejecting potential matches
based on amount matching only.
"""

import os
import traceback
from datetime import datetime
from openpyxl import load_workbook
import pandas as pd
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QTableWidget, QTableWidgetItem, QHeaderView, QMessageBox,
    QGroupBox, QSplitter, QSizePolicy
)
from PySide6.QtCore import Qt, Signal, QTimer
from PySide6.QtGui import QFont, QFontMetrics


class ManualMatchingWindow(QDialog):
    """Dialog window for manual matching of transaction blocks."""
    
    def __init__(self, potential_matches, file1_path, file2_path, parent=None):
        """
        Initialize manual matching window.
        
        Args:
            potential_matches: List of tuples (block1_rows, block2_rows, amount, file1_is_lender)
            file1_path: Path to File 1 Excel file
            file2_path: Path to File 2 Excel file
            parent: Parent widget
        """
        super().__init__(parent)
        self.potential_matches = potential_matches
        self.file1_path = file1_path
        self.file2_path = file2_path
        self.confirmed_matches = []
        self.current_index = 0
        
        # Extract ledger names from metadata (row 3, column 0)
        self.ledger_name1 = self._extract_ledger_name(file1_path)
        self.ledger_name2 = self._extract_ledger_name(file2_path)
        
        # Load workbooks for getting actual cell values
        # Note: Using data_only=False to get actual cell values (not stale calculated values)
        self.wb1 = load_workbook(file1_path, data_only=False)
        self.ws1 = self.wb1.active
        self.wb2 = load_workbook(file2_path, data_only=False)
        self.ws2 = self.wb2.active
        
        self.init_ui()
        self.load_match_pair(0)
    
    def init_ui(self):
        """Initialize the user interface."""
        self.setWindowTitle("Manual Matching - Review Potential Matches")
        
        # Enable maximize and minimize buttons (QDialog doesn't show them by default)
        self.setWindowFlags(
            self.windowFlags() | 
            Qt.WindowType.WindowMaximizeButtonHint | 
            Qt.WindowType.WindowMinimizeButtonHint
        )
        
        # Set minimum size and open maximized
        self.setMinimumSize(800, 600)  # Increased minimum height for taller tables
        # Open window maximized
        self.showMaximized()
        
        # Main layout
        main_layout = QVBoxLayout()
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(15, 15, 15, 15)
        
        # Progress indicator
        progress_label = QLabel()
        progress_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        progress_font = QFont()
        progress_font.setPointSize(12)
        progress_font.setBold(True)
        progress_label.setFont(progress_font)
        self.progress_label = progress_label
        main_layout.addWidget(progress_label)
        
        # Match info (amount and types)
        info_label = QLabel()
        info_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        info_font = QFont()
        info_font.setPointSize(10)
        info_label.setFont(info_font)
        self.info_label = info_label
        main_layout.addWidget(info_label)
        
        # Splitter for side-by-side display
        splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # File 1 panel - use ledger name if available, otherwise fallback to default
        file1_title = self.ledger_name1 if self.ledger_name1 else "File 1 Transaction Block"
        file1_group = QGroupBox(file1_title)
        # Make group box title bold
        group_font = QFont()
        group_font.setBold(True)
        file1_group.setFont(group_font)
        file1_layout = QVBoxLayout()
        file1_layout.setContentsMargins(5, 5, 5, 5)
        file1_table = QTableWidget()
        file1_table.setAlternatingRowColors(True)
        file1_table.horizontalHeader().setStretchLastSection(False)  # We'll control stretching manually
        file1_table.verticalHeader().setVisible(False)
        file1_table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        # Set scrollbar policies
        file1_table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        file1_table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)  # Disable vertical scrollbar - table will auto-resize to fit all rows
        # Set font size to 8pt for table content (not bold)
        table_font = QFont()
        table_font.setPointSize(8)
        table_font.setBold(False)  # Ensure table content is not bold
        file1_table.setFont(table_font)
        # Make ONLY table headers bold (not the table content)
        header_font = QFont()
        header_font.setPointSize(8)
        header_font.setBold(True)
        file1_table.horizontalHeader().setFont(header_font)
        file1_layout.addWidget(file1_table)
        file1_group.setLayout(file1_layout)
        self.file1_table = file1_table
        splitter.addWidget(file1_group)
        
        # File 2 panel - use ledger name if available, otherwise fallback to default
        file2_title = self.ledger_name2 if self.ledger_name2 else "File 2 Transaction Block"
        file2_group = QGroupBox(file2_title)
        # Make group box title bold
        group_font = QFont()
        group_font.setBold(True)
        file2_group.setFont(group_font)
        file2_layout = QVBoxLayout()
        file2_layout.setContentsMargins(5, 5, 5, 5)
        file2_table = QTableWidget()
        file2_table.setAlternatingRowColors(True)
        file2_table.horizontalHeader().setStretchLastSection(False)  # We'll control stretching manually
        file2_table.verticalHeader().setVisible(False)
        file2_table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        # Set scrollbar policies - EXACTLY like file1_table
        file2_table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        file2_table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)  # Disable vertical scrollbar - table will auto-resize to fit all rows
        # Set font size to 8pt for table content (not bold)
        table_font = QFont()
        table_font.setPointSize(8)
        table_font.setBold(False)  # Ensure table content is not bold
        file2_table.setFont(table_font)
        # Make ONLY table headers bold (not the table content)
        header_font = QFont()
        header_font.setPointSize(8)
        header_font.setBold(True)
        file2_table.horizontalHeader().setFont(header_font)
        file2_layout.addWidget(file2_table)
        file2_group.setLayout(file2_layout)
        self.file2_table = file2_table
        splitter.addWidget(file2_group)
        
        # Set equal widths initially, but allow dynamic resizing
        splitter.setSizes([600, 600])
        splitter.setChildrenCollapsible(False)  # Prevent panels from collapsing completely
        main_layout.addWidget(splitter)
        
        # Button layout
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        # Skip button
        skip_button = QPushButton("Skip Manual Matching")
        skip_button.setMinimumWidth(150)
        skip_button.clicked.connect(self.on_skip_manual_matching)
        button_layout.addWidget(skip_button)
        
        # Reject button
        reject_button = QPushButton("Reject Match")
        reject_button.setMinimumWidth(120)
        reject_button.clicked.connect(self.on_reject_match)
        button_layout.addWidget(reject_button)
        
        # Confirm button
        confirm_button = QPushButton("Confirm Match")
        confirm_button.setMinimumWidth(120)
        confirm_button.setDefault(True)
        confirm_button.clicked.connect(self.on_confirm_match)
        button_layout.addWidget(confirm_button)
        
        button_layout.addStretch()
        main_layout.addLayout(button_layout)
        
        self.setLayout(main_layout)
    
    def _extract_ledger_name(self, file_path):
        """Extract ledger name from metadata (row 3, column 0)."""
        try:
            # Read metadata rows (0-7, which are Excel rows 1-8)
            full_df = pd.read_excel(file_path, header=None, nrows=8, dtype=str)
            # Row 3 (index 3) contains the ledger name in column 0
            ledger_name = full_df.iloc[3, 0] if pd.notna(full_df.iloc[3, 0]) else None
            # Clean up the value (remove extra whitespace, handle NaN)
            if ledger_name and str(ledger_name).strip() and str(ledger_name).strip().lower() != 'nan':
                return str(ledger_name).strip()
            return None
        except Exception as e:
            print(f"Error extracting ledger name from {os.path.basename(file_path)}: {e}")
            return None
    
    def load_match_pair(self, index):
        """Load and display a match pair."""
        if index < 0 or index >= len(self.potential_matches):
            return
        
        self.current_index = index
        block1_rows, block2_rows, amount, file1_is_lender = self.potential_matches[index]
        
        # Update progress label
        self.progress_label.setText(f"Match {index + 1} of {len(self.potential_matches)}")
        
        # Update info label - use ledger names instead of "File 1" and "File 2"
        file1_type = "Lender (Debit)" if file1_is_lender else "Borrower (Credit)"
        file2_type = "Borrower (Credit)" if file1_is_lender else "Lender (Debit)"
        # Use ledger names if available, otherwise fallback to "File 1" and "File 2"
        file1_display = self.ledger_name1 if self.ledger_name1 else "File 1"
        file2_display = self.ledger_name2 if self.ledger_name2 else "File 2"
        self.info_label.setText(f"Amount: {amount:,.2f} | {file1_display}: {file1_type} | {file2_display}: {file2_type}")
        
        # Load File 1 block
        self.load_block_to_table(self.file1_table, block1_rows, self.ws1)
        
        # Load File 2 block - exactly like File 1, no differences
        self.load_block_to_table(self.file2_table, block2_rows, self.ws2)
    
    def load_block_to_table(self, table, block_rows, worksheet):
        """Load a transaction block into a table widget."""
        if not block_rows:
            table.setRowCount(0)
            table.setColumnCount(0)
            return
        
        # Get block data
        rows_data = []
        for df_row_idx in block_rows:
            excel_row = df_row_idx + 10  # DataFrame rows start at 0, Excel rows start at 10
            row_data = []
            for col in range(1, 10):  # Columns A-I (1-9)
                cell = worksheet.cell(row=excel_row, column=col)
                value = cell.value
                if value is None:
                    row_data.append("")
                else:
                    # Format dates as dd/MMM/YYYY (Column A is Date)
                    if col == 1:
                        # Column A is Date column
                        if isinstance(value, datetime):
                            formatted_date = value.strftime("%d/%b/%Y")
                            row_data.append(formatted_date)
                        else:
                            # Try to parse as date string if it's a string
                            str_value = str(value).strip()
                            if str_value and str_value.lower() not in ['none', 'nan', '']:
                                try:
                                    # Try parsing common date formats
                                    if '/' in str_value or '-' in str_value:
                                        # Try to parse and reformat
                                        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%Y/%m/%d']:
                                            try:
                                                dt = datetime.strptime(str_value, fmt)
                                                formatted_date = dt.strftime("%d/%b/%Y")
                                                row_data.append(formatted_date)
                                                break
                                            except ValueError:
                                                continue
                                        else:
                                            # If no format matched, use as-is
                                            row_data.append(str_value)
                                    else:
                                        row_data.append(str_value)
                                except Exception:
                                    row_data.append(str_value)
                            else:
                                row_data.append("")
                    else:
                        row_data.append(str(value))
            rows_data.append(row_data)
        
        # Set table dimensions
        num_rows = len(rows_data)
        num_cols = len(rows_data[0]) if rows_data else 9
        
        table.setRowCount(num_rows)
        table.setColumnCount(num_cols)
        
        # Column headers - must match Excel columns A-I (1-9)
        # Excel Column Mapping:
        # Column A (1) = Date
        # Column B (2) = Dr/Cr
        # Column C (3) = Particulars
        # Column D (4) = (empty)
        # Column E (5) = (empty)
        # Column F (6) = Vch Type
        # Column G (7) = Vch No.
        # Column H (8) = Debit
        # Column I (9) = Credit
        headers = ["Date", "Dr/Cr", "Particulars", "", "", "Vch Type", "Vch No.", "Debit", "Credit"]
        if num_cols <= len(headers):
            table.setHorizontalHeaderLabels(headers[:num_cols])
        
        # Populate table
        for row_idx, row_data in enumerate(rows_data):
            for col_idx, value in enumerate(row_data):
                item = QTableWidgetItem(str(value))
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                # Set all columns to middle alignment
                item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                table.setItem(row_idx, col_idx, item)
        
        # Hide headerless columns first (Column D and E - indices 3 and 4)
        if num_cols > 4:
            table.setColumnHidden(3, True)  # Column D (empty header)
            table.setColumnHidden(4, True)  # Column E (empty header)
        
        # Resize columns to content - all columns auto-fit consistently with no extra padding
        table.resizeColumnsToContents()
        
        # Make Particulars column (index 2) stretch to fill remaining space
        if num_cols > 2:
            table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        
        # Force layout update to get accurate Particulars column width after stretching
        table.updateGeometry()
        
        # Calculate fixed base row height - same as "Entered By" row (reliable, no extra padding)
        table_font = table.font()
        font_metrics = QFontMetrics(table_font)
        # Base row height = font height + minimal padding (2px top + 2px bottom = 4px total)
        base_row_height = font_metrics.height() + 4
        
        # Set ALL rows to fixed base height first (like "Entered By" row)
        # This ensures consistent height - no extra padding, reliable like "Entered By" row
        table.verticalHeader().setDefaultSectionSize(base_row_height)
        table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        for row_idx in range(num_rows):
            table.setRowHeight(row_idx, base_row_height)
        
        # Enable word wrap and selectively expand only rows where Particulars actually wraps
        table.setWordWrap(True)
        
        # Get Particulars column width after stretch is applied (use a small delay to ensure layout is complete)
        # Schedule row height adjustment after layout is complete
        def adjust_row_heights():
            particulars_width = table.columnWidth(2) if num_cols > 2 else 0
            
            # Check each row and expand only if Particulars text would actually wrap
            if particulars_width > 0:
                for row_idx in range(num_rows):
                    particulars_item = table.item(row_idx, 2)  # Particulars column
                    if particulars_item:
                        text = str(particulars_item.text()).strip()
                        if text:
                            # Calculate if text would wrap at current column width
                            text_width = font_metrics.horizontalAdvance(text)
                            # If text is wider than column, it will wrap - expand row
                            if text_width > particulars_width:
                                # Calculate how many lines needed (account for word boundaries and padding)
                                available_width = particulars_width - 20  # Account for cell padding
                                lines_needed = max(1, (text_width // max(available_width, 1)) + 1)
                                # Set row height to accommodate wrapped text
                                wrapped_height = (font_metrics.height() * lines_needed) + 4
                                table.setRowHeight(row_idx, wrapped_height)
                            # If text fits, keep base height (already set above)
        
        # Use QTimer to adjust row heights after layout is complete
        QTimer.singleShot(10, adjust_row_heights)
        
        # Calculate total table height needed to fit all rows (no scrollbar)
        def set_table_height_to_fit_all_rows():
            # Get header height
            header_height = table.horizontalHeader().height()
            if header_height == 0:
                # If header height is 0, use a default estimate
                header_height = font_metrics.height() + 8
            # Calculate total row heights
            total_row_height = 0
            for row_idx in range(num_rows):
                total_row_height += table.rowHeight(row_idx)
            # Set minimum and maximum height to fit all rows + header + small margin
            total_height = header_height + total_row_height + 2  # +2 for small margin
            table.setMinimumHeight(total_height)
            table.setMaximumHeight(total_height)  # Also set max to prevent expansion beyond content
        
        # Schedule height adjustment after row heights are set (with longer delay to ensure all adjustments are complete)
        QTimer.singleShot(100, set_table_height_to_fit_all_rows)
    
    def on_confirm_match(self):
        """Handle user confirmation of a match."""
        try:
            if self.current_index >= len(self.potential_matches):
                return
            
            block1_rows, block2_rows, amount, file1_is_lender = self.potential_matches[self.current_index]
            
            # Get header row indices
            file1_header_row = block1_rows[0] if block1_rows else None
            file2_header_row = block2_rows[0] if block2_rows else None
            
            if file1_header_row is None or file2_header_row is None:
                return
            
            # Get amounts from header rows using already-loaded workbooks
            # Extract amounts directly from the worksheets we already have open
            file1_header_row_idx = block1_rows[0]
            file2_header_row_idx = block2_rows[0]
            excel_row1 = file1_header_row_idx + 10
            excel_row2 = file2_header_row_idx + 10
            
            # Get debit and credit from File 1 header row
            debit_cell1 = self.ws1.cell(row=excel_row1, column=8)  # Column H
            credit_cell1 = self.ws1.cell(row=excel_row1, column=9)  # Column I
            file1_debit = float(debit_cell1.value) if debit_cell1.value is not None and debit_cell1.value != 0 else 0.0
            file1_credit = float(credit_cell1.value) if credit_cell1.value is not None and credit_cell1.value != 0 else 0.0
            
            # Get debit and credit from File 2 header row
            debit_cell2 = self.ws2.cell(row=excel_row2, column=8)  # Column H
            credit_cell2 = self.ws2.cell(row=excel_row2, column=9)  # Column I
            file2_debit = float(debit_cell2.value) if debit_cell2.value is not None and debit_cell2.value != 0 else 0.0
            file2_credit = float(credit_cell2.value) if credit_cell2.value is not None and credit_cell2.value != 0 else 0.0
            
            # Handle string values (in case data_only=True returns strings)
            try:
                if isinstance(file1_debit, str):
                    file1_debit = float(str(file1_debit).replace(',', ''))
            except (ValueError, TypeError):
                file1_debit = 0.0
            try:
                if isinstance(file1_credit, str):
                    file1_credit = float(str(file1_credit).replace(',', ''))
            except (ValueError, TypeError):
                file1_credit = 0.0
            try:
                if isinstance(file2_debit, str):
                    file2_debit = float(str(file2_debit).replace(',', ''))
            except (ValueError, TypeError):
                file2_debit = 0.0
            try:
                if isinstance(file2_credit, str):
                    file2_credit = float(str(file2_credit).replace(',', ''))
            except (ValueError, TypeError):
                file2_credit = 0.0
            
            # Create match dictionary (following standard match format)
            match = {
                'match_id': None,  # Will be assigned later
                'Match_Type': 'Manual',
                'File1_Index': file1_header_row,
                'File2_Index': file2_header_row,
                'File1_Debit': file1_debit,
                'File1_Credit': file1_credit,
                'File2_Debit': file2_debit,
                'File2_Credit': file2_credit,
                'Amount': amount,
                'File1_Amount': file1_debit if file1_debit > 0 else file1_credit,
                'File2_Amount': file2_debit if file2_debit > 0 else file2_credit
            }
            
            self.confirmed_matches.append(match)
            self.move_to_next()
        except Exception as e:
            # Show error message if something goes wrong
            QMessageBox.warning(self, "Error", f"Error confirming match: {str(e)}")
            print(f"Error in on_confirm_match: {e}")
            print(traceback.format_exc())
    
    def on_reject_match(self):
        """Handle user rejection of a match."""
        self.move_to_next()
    
    def on_skip_manual_matching(self):
        """Handle user skipping manual matching entirely."""
        # Close window and indicate skip (return code 2 for skip)
        self.done(2)
    
    def move_to_next(self):
        """Move to the next match pair."""
        self.current_index += 1
        
        if self.current_index >= len(self.potential_matches):
            # All matches reviewed
            self.accept()
        else:
            self.load_match_pair(self.current_index)
    
    def get_confirmed_matches(self):
        """Get the list of confirmed matches."""
        return self.confirmed_matches
    
    def keyPressEvent(self, event):
        """Override key press event to prevent accidental closing on Esc."""
        if event.key() == Qt.Key.Key_Escape:
            # Trigger close event which has the confirmation dialog
            self.close()
        else:
            super().keyPressEvent(event)
    
    def closeEvent(self, event):
        """Handle window close event with confirmation."""
        # Ask for confirmation before closing
        reply = QMessageBox.question(
            self,
            "Confirm Close",
            "Are you sure you want to close the manual matching window?\n\nAny unconfirmed matches will be lost.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            # User confirmed, close the workbooks and accept the close event
            self.wb1.close()
            self.wb2.close()
            super().closeEvent(event)
        else:
            # User cancelled, ignore the close event
            event.ignore()
