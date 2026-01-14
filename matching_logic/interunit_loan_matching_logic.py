#!/usr/bin/env python3
"""
Interunit Loan Matcher Module

This module implements the CORRECT interunit loan matching logic:
1. DEBIT block (Lender): Full account number in ledger â†’ extract short code
2. CREDIT block (Borrower): Short code appears in narration only
3. Debit/Credit amounts must match exactly (no tolerance)
4. ONE-WAY cross-reference: Borrower's narration should contain Lender's short code
5. Accounts must be different (lender vs borrower)
6. FOLLOWS CORE LOGIC AND FORMAT - cannot deviate
"""

import pandas as pd
import re
from openpyxl import load_workbook
from typing import Dict, List, Optional, Any
from transaction_block_identifier import TransactionBlockIdentifier
import sys
import os

# Add parent directory to path to import bank_account_mapping
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Import unmatched tracker
try:
    from ..unmatched_tracker import get_unmatched_tracker
except ImportError:
    from unmatched_tracker import get_unmatched_tracker

class InterunitLoanMatcher:
    """
    Matches interunit loan transactions between two files based on:
    - Full account numbers in ledger
    - Short codes in narration
    - Exact amount matching
    - Cross-referenced short codes (BIDIRECTIONAL)
    - FOLLOWS CORE LOGIC AND FORMAT EXACTLY
    """
    
    def __init__(self, block_identifier):
        """
        Initialize with a shared TransactionBlockIdentifier instance.
        
        Args:
            block_identifier: Shared instance of TransactionBlockIdentifier for consistent transaction block logic
        """
        self.block_identifier = block_identifier
        # Interunit account mapping (Full Format → List of Short Codes)
        # Imported from bank_account_mapping.py - reload on each initialization
        self.reload_mapping()
        
        # No amount tolerance - exact matching required
    
    def reload_mapping(self):
        """Reload the bank account mapping from file"""
        from bank_account_mapping import load_mapping
        self.interunit_account_mapping = load_mapping()
    
    def find_potential_matches(
        self, 
        transactions1: pd.DataFrame, 
        transactions2: pd.DataFrame, 
        interunit_accounts1: pd.Series, 
        interunit_accounts2: pd.Series, 
        file1_path: str, 
        file2_path: str, 
        existing_matches: Dict = None, 
        match_id_manager = None
    ) -> List[Dict]:
        """
        Find interunit loan matches between two transaction files.
        IMPLEMENTS THE CORRECT CROSS-REFERENCING LOGIC.
        FOLLOWS CORE LOGIC AND FORMAT EXACTLY - cannot deviate.
        
        Args:
            transactions1: DataFrame of transactions from first file
            transactions2: DataFrame of transactions from second file
            interunit_accounts1: Series of interunit accounts from first file
            interunit_accounts2: Series of interunit accounts from second file
            file1_path: Path to first file (for openpyxl access)
            file2_path: Path to second file (for openpyxl access)
            existing_matches: Dictionary of existing matches (shared state)
            match_counter: Counter for generating unique match IDs (shared state)
            
        Returns:
            List of match dictionaries following core format
        """
        matches = []
        
        # Use shared state if provided, otherwise create new
        if existing_matches is None:
            existing_matches = {}
        if match_id_manager is None:
            try:
                from ..match_id_manager import get_match_id_manager
            except ImportError:
                from match_id_manager import get_match_id_manager
            match_id_manager = get_match_id_manager()
        
        print(f"\n=== INTERUNIT LOAN MATCHING LOGIC (ONE-WAY) ===")
        print(f"1. DEBIT block (Lender): Full account number in ledger => extract short code")
        print(f"2. CREDIT block (Borrower): Short code appears in narration only")
        print(f"3. Debit/Credit amounts must match exactly (no tolerance)")
        print(f"4. ONE-WAY cross-reference: Borrower's narration should contain Lender's short code")
        print(f"5. Accounts must be different (lender vs borrower)")
        print(f"6. FOLLOWS CORE LOGIC: Uses universal M001 format, shared state, same structure")
        
        # Identify transaction blocks in both files
        blocks1 = self.block_identifier.identify_transaction_blocks(transactions1, file1_path)
        blocks2 = self.block_identifier.identify_transaction_blocks(transactions2, file2_path)
        
        print(f"\nFile 1: {len(blocks1)} transaction blocks")
        print(f"File 2: {len(blocks2)} transaction blocks")
        
        # OPTIMIZATION: Filter by matching amounts FIRST using universal template
        print(f"\n--- Filtering blocks by matching amounts (universal template) ---")
        matching_pairs = self.block_identifier.filter_blocks_by_matching_amounts(
            blocks1, blocks2, file1_path, file2_path
        )
        print(f"Found {len(matching_pairs)} block pairs with matching amounts and opposite types")
        
        if not matching_pairs:
            print("No matching amount pairs found. No interunit matches possible.")
            return []
        
        # OPTIMIZATION: Use cached workbooks if available, otherwise load
        if file1_path in self.block_identifier._cached_workbooks:
            wb1, ws1 = self.block_identifier._cached_workbooks[file1_path]
        else:
            wb1 = load_workbook(file1_path)
            ws1 = wb1.active
            self.block_identifier._cached_workbooks[file1_path] = (wb1, ws1)
        
        if file2_path in self.block_identifier._cached_workbooks:
            wb2, ws2 = self.block_identifier._cached_workbooks[file2_path]
        else:
            wb2 = load_workbook(file2_path)
            ws2 = wb2.active
            self.block_identifier._cached_workbooks[file2_path] = (wb2, ws2)
        
        # Create sets of blocks that are in matching pairs (for efficient lookup)
        blocks1_in_pairs = {tuple(pair[0]) for pair in matching_pairs}
        blocks2_in_pairs = {tuple(pair[1]) for pair in matching_pairs}
        
        # Collect interunit account information ONLY from blocks in matching pairs
        file1_interunit_data = []
        file2_interunit_data = []
        
        # Create mapping from block_rows tuple to block_data for quick lookup
        file1_block_data_map = {}
        file2_block_data_map = {}
        
        print(f"\n--- Scanning File 1 (only matching blocks) for interunit data ---")
        for i, block in enumerate(blocks1):
            if tuple(block) in blocks1_in_pairs:
                block_data = self._analyze_block_for_interunit_data(ws1, block, i, file1_path)
                if block_data['ledger_accounts'] or block_data['narration_short_codes']:
                    file1_interunit_data.append(block_data)
                    file1_block_data_map[tuple(block)] = block_data
                    if len(file1_interunit_data) <= 5:  # Show first 5
                        print(f"Block {i+1}: {len(block_data['ledger_accounts'])} ledger accounts, {len(block_data['narration_short_codes'])} short codes")
        
        print(f"\n--- Scanning File 2 (only matching blocks) for interunit data ---")
        for i, block in enumerate(blocks2):
            if tuple(block) in blocks2_in_pairs:
                block_data = self._analyze_block_for_interunit_data(ws2, block, i, file2_path)
                if block_data['ledger_accounts'] or block_data['narration_short_codes']:
                    file2_interunit_data.append(block_data)
                    file2_block_data_map[tuple(block)] = block_data
                    if len(file2_interunit_data) <= 5:  # Show first 5
                        print(f"Block {i+1}: {len(block_data['ledger_accounts'])} ledger accounts, {len(block_data['narration_short_codes'])} short codes")
        
        print(f"\nFile 1: {len(file1_interunit_data)} blocks with interunit data (from {len(blocks1_in_pairs)} matching blocks)")
        print(f"File 2: {len(file2_interunit_data)} blocks with interunit data (from {len(blocks2_in_pairs)} matching blocks)")
        
        # Look for cross-referenced matches using pre-filtered pairs
        print(f"\n--- Looking for cross-referenced matches ---")
        potential_matches = []
        
        # Initialize unmatched tracker for audit info
        unmatched_tracker = get_unmatched_tracker()
        
        # Process only the matching pairs (already filtered by amount and type)
        for block1_rows, block2_rows, amount, file1_is_lender in matching_pairs:
            block1_key = tuple(block1_rows)
            block2_key = tuple(block2_rows)
            
            # Get block data if it has interunit information
            block1 = file1_block_data_map.get(block1_key)
            block2 = file2_block_data_map.get(block2_key)
            
            if not block1 or not block2:
                # Track why this pair didn't match
                if not block1:
                    reason = "No interunit account found in File 1 block"
                    unmatched_tracker.add_unmatched_reason(block1_rows[0] if block1_rows else None, reason, 1)
                if not block2:
                    reason = "No interunit account found in File 2 block"
                    unmatched_tracker.add_unmatched_reason(block2_rows[0] if block2_rows else None, reason, 2)
                continue  # Skip if no interunit data
            
            # Amounts already match (from filtering), so proceed directly to cross-reference check
            # Check for cross-referenced short codes (ONE-WAY LOGIC)
            cross_reference_found = False
            matched_short_code = None
            
            # Get lender/borrower status (already known from filtering)
            file1_is_borrower = not file1_is_lender
            file2_is_lender = not file1_is_lender
            file2_is_borrower = file1_is_lender
            
            # Determine short codes for matching (ONE-WAY LOGIC)
            # LENDER (DEBIT block): Full account number in ledger → extract short code
            # BORROWER (CREDIT block): Short code appears in narration only
            # Matching: Borrower's narration should contain Lender's short code (ONE-WAY ONLY)
            
            # Determine Lender's short codes (from ledger accounts)
            lender_short_codes_to_check = []
            borrower_block = None
            lender_block = None
            
            if file1_is_lender and file2_is_borrower:
                # File 1 is lender (DEBIT), File 2 is borrower (CREDIT)
                lender_block = block1
                borrower_block = block2
                if block1.get('header_has_as_per_details') and block1.get('primary_bank_account_short_codes'):
                    lender_short_codes_to_check = block1['primary_bank_account_short_codes']
                else:
                    lender_short_codes_to_check = [ledger['short_code'] for ledger in block1['ledger_accounts']]
            elif file1_is_borrower and file2_is_lender:
                # File 1 is borrower (CREDIT), File 2 is lender (DEBIT)
                lender_block = block2
                borrower_block = block1
                if block2.get('header_has_as_per_details') and block2.get('primary_bank_account_short_codes'):
                    lender_short_codes_to_check = block2['primary_bank_account_short_codes']
                else:
                    lender_short_codes_to_check = [ledger['short_code'] for ledger in block2['ledger_accounts']]
            
            # ONE-WAY CHECK: Borrower's narration contains Lender's short code
            borrower_narration_contains_lender_code = False
            matched_short_code = None
            
            if lender_block and borrower_block and lender_short_codes_to_check:
                # Check if borrower's narration contains any of the lender's short codes
                for narration in borrower_block['narration_short_codes']:
                    for lender_short_code in lender_short_codes_to_check:
                        if narration['short_code'] == lender_short_code:
                            borrower_narration_contains_lender_code = True
                            matched_short_code = lender_short_code
                            break
                    if borrower_narration_contains_lender_code:
                        break
            
            # ONE-WAY: Only borrower's narration containing lender's short code is required
            cross_reference_found = borrower_narration_contains_lender_code
            
            # If cross-reference found, we have a match
            if cross_reference_found:
                # Mark as matched in tracker
                unmatched_tracker.mark_as_matched(block1_rows[0], block2_rows[0])
                
                # We have a match! Generate next sequential Match ID using centralized manager
                context = f"Interunit_{amount}_File1_Block_{block1['block_index']}_File2_Block_{block2['block_index']}"
                # Match ID will be assigned later in post-processing
                match_id = None
                
                # Create match following CORE FORMAT exactly
                match = {
                    'match_id': match_id,
                    'Match_Type': 'Interunit',  # Add explicit match type
                    'Interunit_Account': matched_short_code if matched_short_code else 'Matched',
                    'File1_Index': block1['amounts']['row'],
                    'File2_Index': block2['amounts']['row'],
                    'File1_Debit': block1['amounts']['debit'],
                    'File1_Credit': block1['amounts']['credit'],
                    'File2_Debit': block2['amounts']['debit'],
                    'File2_Credit': block2['amounts']['credit'],
                    'File1_Amount': amount,  # Add File1_Amount for audit info
                    'File2_Amount': amount,  # Add File2_Amount for audit info
                    'Amount': amount
                }
                
                potential_matches.append(match)
                print(f"  MATCH {match_id}: Amount {amount}")
                print(f"    Cross-reference: Borrower's narration contains lender's short code: {matched_short_code}")
            else:
                # Track why cross-reference failed (ONE-WAY LOGIC)
                if not borrower_narration_contains_lender_code:
                    reason = "Borrower's narration does not contain lender's short code"
                    unmatched_tracker.add_unmatched_reason(block1_rows[0] if block1_rows else None, reason, 1)
                    unmatched_tracker.add_unmatched_reason(block2_rows[0] if block2_rows else None, reason, 2)
        
        # OPTIMIZATION: Don't close workbooks if they're cached (keep them open for performance)
        if file1_path not in self.block_identifier._cached_workbooks:
            wb1.close()
        if file2_path not in self.block_identifier._cached_workbooks:
            wb2.close()
        
        print(f"\nInterunit Loan Matching Complete: {len(potential_matches)} matches found")
        print(f"FOLLOWS CORE LOGIC: Uses universal M001 format, integrates with shared state")
        return potential_matches
    
    def _analyze_block_for_interunit_data(self, worksheet, block_rows, block_index, file_path):
        """Analyze a transaction block for interunit account data."""
        block_data = {
            'block_index': block_index,
            'block_rows': block_rows,
            'ledger_accounts': [],
            'narration_short_codes': [],
            'amounts': {},
            'header_has_as_per_details': False,
            'primary_bank_account_short_codes': []  # For blocks with "(as per details)" - list of all short codes
        }
        
        # Get amounts from the header row using universal method
        # The header row is guaranteed to be the first element in block_rows
        block_data['amounts'] = self.block_identifier.get_block_header_amounts(block_rows, file_path)
        
        # Check header row for "(as per details)"
        header_row_idx = block_rows[0]
        excel_header_row = header_row_idx + 10
        if excel_header_row <= worksheet.max_row:
            header_cell_c = worksheet.cell(row=excel_header_row, column=3)  # Column C
            if header_cell_c.value and "(as per details)" in str(header_cell_c.value):
                block_data['header_has_as_per_details'] = True
        
        # Check each row in the block for ledger accounts and narration
        for row_idx in block_rows:
            excel_row = row_idx + 10  # Convert to Excel row number
            
            if excel_row <= worksheet.max_row:
                cell_c = worksheet.cell(row=excel_row, column=3)  # Column C
                
                # Check for ledger accounts (Bold but not italic)
                if (cell_c.value and 
                    cell_c.font and 
                    cell_c.font.bold and 
                    not cell_c.font.italic):
                    
                    # OPTIMIZATION: Check if this is an interunit account - stop at first match
                    cell_value_upper = str(cell_c.value).upper()
                    found_account = None
                    
                    for full_account, short_codes in self.interunit_account_mapping.items():
                        if full_account.upper() in cell_value_upper:
                            found_account = (full_account, short_codes)
                            break  # Stop at first match
                    
                    if found_account:
                        full_account, short_codes = found_account
                        # Handle both list and single string formats for backward compatibility
                        codes_list = short_codes if isinstance(short_codes, list) else [short_codes]
                        # Add all short codes for this account
                        for short_code in codes_list:
                            block_data['ledger_accounts'].append({
                                'full_account': full_account,
                                'short_code': short_code,
                                'cell_value': cell_c.value
                            })
                            
                            # If header has "(as per details)" and this is a bank account (not header row),
                            # add all short codes from this account to primary_bank_account_short_codes
                            if (block_data['header_has_as_per_details'] and 
                                row_idx != header_row_idx and  # Not the header row itself
                                len(block_data['primary_bank_account_short_codes']) == 0):  # Only set once
                                # Add all short codes from this account (codes_list already contains all codes)
                                block_data['primary_bank_account_short_codes'] = codes_list.copy()
                                # If we found the primary account, we can stop searching this row
                                break
                
                # Check for narration rows (Italic but not bold)
                elif (cell_c.value and 
                      cell_c.font and 
                      not cell_c.font.bold and 
                      cell_c.font.italic):
                    
                    narration_text = str(cell_c.value).upper()
                    
                    # Look for short codes in narration
                    for short_codes in self.interunit_account_mapping.values():
                        # Handle both list and single string formats for backward compatibility
                        codes_list = short_codes if isinstance(short_codes, list) else [short_codes]
                        for short_code in codes_list:
                            # Check for exact match first
                            if short_code.upper() in narration_text:
                                block_data['narration_short_codes'].append({
                                    'short_code': short_code,
                                    'narration': cell_c.value
                                })
                            else:
                                # Also check for account number without prefix (e.g., "4056" matches "EBL#4056")
                                # Extract account number from short code (format: "EBL#4056" -> "4056")
                                if '#' in short_code:
                                    account_number = short_code.split('#')[1]
                                    # Check if account number appears in narration (with or without prefix variations)
                                    # Look for patterns like: "EBL#4056", "EBL# 4056", "4056", "EBL 4056", etc.
                                    import re
                                    # Pattern to match the account number with optional prefix and spacing
                                    pattern = r'(?:[A-Z]{2,4}[#\s]*)?' + re.escape(account_number) + r'(?:\s|&|,|$)'
                                    if re.search(pattern, narration_text):
                                        block_data['narration_short_codes'].append({
                                            'short_code': short_code,
                                            'narration': cell_c.value
                                        })
                

        
        return block_data
    
    def extract_interunit_accounts_from_narration(self, transactions: pd.DataFrame, file_path: str) -> pd.Series:
        """
        Extract interunit account references from NARRATION rows (Column C - Italic).
        FOLLOWS CORE LOGIC - same pattern as existing modules.
        """
        print(f"Extracting interunit account references from NARRATION column...")
        
        # Initialize interunit accounts series with None values
        interunit_accounts = pd.Series([None] * len(transactions), index=transactions.index)
        
        # OPTIMIZATION: Use cached workbook if available, otherwise load
        try:
            if file_path in self.block_identifier._cached_workbooks:
                wb, ws = self.block_identifier._cached_workbooks[file_path]
            else:
                wb = load_workbook(file_path)
                ws = wb.active
                self.block_identifier._cached_workbooks[file_path] = (wb, ws)
            
            # Find NARRATION rows (Italic text in Column C) - NOT ledger rows
            for idx, row_idx in enumerate(transactions.index):
                excel_row = row_idx + 9  # Adjust for metadata rows
                
                if excel_row <= ws.max_row:
                    cell = ws.cell(row=excel_row, column=3)  # Column C
                    
                    if (cell.value and 
                        cell.font and 
                        not cell.font.bold and 
                        cell.font.italic):  # Italic but NOT bold = NARRATION row
                        
                        # This is a NARRATION row, extract interunit account reference
                        account_info = self.extract_interunit_account_from_narration(str(cell.value))
                        if account_info:
                            interunit_accounts.iloc[idx] = account_info['full_reference']
                            print(f"  Row {row_idx}: Found interunit account '{account_info['full_reference']}' in NARRATION")
            
            # OPTIMIZATION: Don't close workbook if it's cached
            if file_path not in self.block_identifier._cached_workbooks:
                wb.close()
            
        except Exception as e:
            print(f"Error reading Excel file for interunit account extraction: {e}")
        
        print(f"Interunit account extraction complete. Found {interunit_accounts.notna().sum()} account references")
        return interunit_accounts
    
    def extract_interunit_account_from_narration(self, narration: str) -> Optional[Dict[str, Any]]:
        """
        Extract interunit account reference from narration text.
        FOLLOWS CORE LOGIC - same pattern as existing modules.
        """
        if not narration:
            return None
        
        # Look for short codes in narration (e.g., MTBL#3858, OBL#8826)
        short_code_patterns = [
            r'([A-Z]{2,4})#(\d{4,6})',  # MTBL#4355, MDBL#11026, OBL#8826
        ]
        
        for pattern in short_code_patterns:
            try:
                match = re.search(pattern, narration.upper())
                if match:
                    bank_code = match.group(1).strip()
                    account_number = match.group(2)
                    
                    return {
                        'account_number': account_number,
                        'bank_code': bank_code,
                        'full_reference': match.group(),
                        'full_account_format': None
                    }
            except Exception as e:
                pass  # Regex error - skip this pattern
                continue
        
        return None

        return None
