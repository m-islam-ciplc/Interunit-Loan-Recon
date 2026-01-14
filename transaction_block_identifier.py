"""
Transaction Block Identifier Module

This module contains the logic for identifying transaction blocks in Excel files
based on specific formatting and content criteria.
"""

import openpyxl
import pandas as pd
import os


class TransactionBlockIdentifier:
    """
    Identifies transaction blocks in Excel files based on formatting and content.
    
    Transaction Block Structure:
    1. Block Start: Date + Dr/Cr + Ledger + Vch Type (Bold) + Vch No. (Regular) + Debit/Credit (Bold)
    2. Block Content: Ledger entries (Bold) and Narration (Regular text - not bold, not italic)
    3. Block End: "Entered By :" + Person name (Bold + Italic)
    """
    
    def __init__(self):
        """Initialize the TransactionBlockIdentifier."""
        # Cache for block mappings: {file_path: {row_index: block_rows}}
        self._block_cache = {}
        # Cache for workbooks: {file_path: (wb, ws)} - keep workbooks open for performance
        self._cached_workbooks = {}
        # Cache for amount lookups: {file_path: {header_row_idx: amounts_dict}}
        self._amount_cache = {}
        # Cache for block headers: {file_path: {description_row_idx: block_header_idx}}
        self._block_header_cache = {}
        # Cache for narration index: {file_path: {narration: [(block_header_idx, description_idx, amounts_dict)]}}
        self._narration_index_cache = {}
    
    def precompute_all_blocks(self, transactions_df, file_path):
        """
        Pre-compute all block mappings for a file and cache them.
        This eliminates the need to reload Excel files repeatedly.
        
        Args:
            transactions_df: DataFrame containing transaction data
            file_path: Path to the Excel file
        """
        if file_path in self._block_cache:
            # Already cached
            return
        
        print(f"Pre-computing block mappings for {os.path.basename(file_path)}...")
        
        # Load workbook once and cache it
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        self._cached_workbooks[file_path] = (wb, ws)
        
        # Identify all blocks using the cached workbook (don't reload)
        all_blocks = self._identify_transaction_blocks_with_worksheet(transactions_df, ws)
        
        # Create reverse mapping: for each row, which block does it belong to?
        row_to_block = {}
        for block_rows in all_blocks:
            for row_idx in block_rows:
                row_to_block[row_idx] = block_rows
        
        # Cache all mappings
        self._block_cache[file_path] = row_to_block
        
        print(f"  Cached {len(row_to_block)} row-to-block mappings for {len(all_blocks)} blocks")
    
    def _identify_transaction_blocks_with_worksheet(self, transactions_df, ws):
        """
        Internal helper method to identify blocks using an already-loaded worksheet.
        This avoids reloading the workbook when we already have it cached.
        
        Args:
            transactions_df: DataFrame containing transaction data
            ws: Already-loaded worksheet object
        
        Returns:
            List of transaction block row indices
        """
        transaction_blocks = []
        current_block = []
        in_block = False
        
        for row_idx in range(10, ws.max_row + 1):  # Start from row 10 (after headers)
            # Convert Excel row to DataFrame row index
            df_row_idx = row_idx - 10
            
            if df_row_idx < 0:
                continue
            
            # Check if this row starts a new transaction block
            date_cell = ws.cell(row=row_idx, column=1)  # Column A (Date)
            particulars_cell = ws.cell(row=row_idx, column=2)  # Column B (Particulars)
            vch_type_cell = ws.cell(row=row_idx, column=6)  # Column F (Vch Type)
            vch_no_cell = ws.cell(row=row_idx, column=7)  # Column G (Vch No.)
            debit_cell = ws.cell(row=row_idx, column=8)  # Column H (Debit)
            credit_cell = ws.cell(row=row_idx, column=9)  # Column I (Credit)
            
            # Check if this row has a real date and Dr/Cr
            has_real_date = (date_cell.value and 
                           str(date_cell.value).strip() and 
                           str(date_cell.value).strip() != 'None' and
                           str(date_cell.value).strip() != '')
            has_dr_cr = particulars_cell.value and str(particulars_cell.value).strip() in ['Dr', 'Cr']
            
            # Check if this row has Vch Type (Bold) and Vch No. (Regular) - required for transaction blocks
            has_vch_type = vch_type_cell.value and vch_type_cell.font and vch_type_cell.font.bold
            has_vch_no = vch_no_cell.value and vch_no_cell.font and not vch_no_cell.font.bold and not vch_no_cell.font.italic
            
            # Check if this row has Debit or Credit amount (Bold) - required for transaction blocks
            has_debit = debit_cell.value and debit_cell.font and debit_cell.font.bold
            has_credit = credit_cell.value and credit_cell.font and credit_cell.font.bold
            
            # Check if this is NOT an Opening Balance row (which is not a transaction block)
            is_not_opening_balance = not (particulars_cell.value and 'Opening Balance' in str(particulars_cell.value))
            
            # Check if this row ends a transaction block ("Entered By :")
            is_block_end = particulars_cell.value and str(particulars_cell.value).strip() == 'Entered By :'
            
            # Transaction block start requires: Date + Dr/Cr + Vch Type + Vch No. + Debit/Credit + NOT Opening Balance
            is_block_start = (has_real_date and has_dr_cr and has_vch_type and has_vch_no and 
                             (has_debit or has_credit) and is_not_opening_balance)
            
            if is_block_start:
                # If we're already in a block, end the current one
                if in_block and current_block:
                    transaction_blocks.append(current_block)
                
                # Start new block
                current_block = [df_row_idx]
                in_block = True
            elif in_block:
                # Continue adding rows to current block
                current_block.append(df_row_idx)
                
                # Check if this row ends the block
                if is_block_end:
                    # End the current block
                    transaction_blocks.append(current_block)
                    current_block = []
                    in_block = False
        
        # Add the last block if it exists
        if in_block and current_block:
            transaction_blocks.append(current_block)
        
        return transaction_blocks
    
    def clear_cache(self, file_path=None):
        """
        Clear cache for a specific file or all files.
        
        Args:
            file_path: If provided, clear cache for this file only. If None, clear all caches.
        """
        if file_path:
            if file_path in self._block_cache:
                del self._block_cache[file_path]
            if file_path in self._amount_cache:
                del self._amount_cache[file_path]
            if file_path in self._block_header_cache:
                del self._block_header_cache[file_path]
            if file_path in self._narration_index_cache:
                del self._narration_index_cache[file_path]
            if file_path in self._cached_workbooks:
                wb, ws = self._cached_workbooks[file_path]
                wb.close()
                del self._cached_workbooks[file_path]
        else:
            # Close all workbooks
            for wb, ws in self._cached_workbooks.values():
                wb.close()
            self._block_cache.clear()
            self._amount_cache.clear()
            self._block_header_cache.clear()
            self._narration_index_cache.clear()
            self._cached_workbooks.clear()
    
    def get_transaction_block_rows(self, lc_match_row, file_path):
        """
        Get all row indices that belong to the transaction block containing the LC match.
        NOW WITH CACHING - uses pre-computed mappings if available.
        
        Args:
            lc_match_row: The row index where the LC match was found
            file_path: Path to the Excel file to analyze
        
        Returns:
            List of row indices that belong to the transaction block (from start to "Entered By :")
        """
        # Check cache first
        if file_path in self._block_cache:
            if lc_match_row in self._block_cache[file_path]:
                # Cache hit - return immediately
                return self._block_cache[file_path][lc_match_row]
            else:
                # Row not in cache - might be outside transaction blocks (e.g., opening balance)
                # Return empty list to maintain compatibility
                return []
        
        # Cache miss - fall back to original logic (for backward compatibility)
        # This should rarely happen if precompute_all_blocks() is called first
        block_rows = []
        
        # Use cached workbook if available, otherwise load it
        if file_path in self._cached_workbooks:
            wb, ws = self._cached_workbooks[file_path]
        else:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            self._cached_workbooks[file_path] = (wb, ws)
        
        # Convert DataFrame row index to Excel row number 
        # DataFrame starts at 0, but Excel has metadata rows 1-8, then headers at row 9, then data starts at row 10
        # So DataFrame row 0 = Excel row 10, DataFrame row 1 = Excel row 11, etc.
        excel_lc_row = lc_match_row + 10
        
        # FIRST: Look BACKWARDS from the LC match row to find the ACTUAL start of the transaction block
        # Transaction block starts where we find Date + Dr/Cr + Vch Type + Vch No. + Debit/Credit
        block_start_row = excel_lc_row
        for row_idx in range(excel_lc_row, 8, -1):  # Go backwards from LC row to row 9
            date_cell = ws.cell(row=row_idx, column=1)  # Column A (Date)
            particulars_cell = ws.cell(row=row_idx, column=2)  # Column B (Particulars)
            vch_type_cell = ws.cell(row=row_idx, column=6)  # Column F (Vch Type)
            vch_no_cell = ws.cell(row=row_idx, column=7)  # Column G (Vch No.)
            debit_cell = ws.cell(row=row_idx, column=8)  # Column H (Debit)
            credit_cell = ws.cell(row=row_idx, column=9)  # Column I (Credit)
            
            # Check if this row has a real date and Dr/Cr (transaction block start)
            has_real_date = (date_cell.value and 
                           str(date_cell.value).strip() and 
                           str(date_cell.value).strip() != 'None' and
                           str(date_cell.value).strip() != '')
            has_dr_cr = particulars_cell.value and str(particulars_cell.value).strip() in ['Dr', 'Cr']
            
            # Check if this row has Vch Type (Bold) and Vch No. (Regular) - required for transaction blocks
            has_vch_type = vch_type_cell.value and vch_type_cell.font and vch_type_cell.font.bold
            has_vch_no = vch_no_cell.value and vch_no_cell.font and not vch_no_cell.font.bold and not vch_no_cell.font.italic
            
            # Check if this row has Debit or Credit amount (Bold) - required for transaction blocks
            has_debit = debit_cell.value and debit_cell.font and debit_cell.font.bold
            has_credit = credit_cell.value and credit_cell.font and credit_cell.font.bold
            
            # Check if this is NOT an Opening Balance row (which is not a transaction block)
            is_not_opening_balance = not (particulars_cell.value and 'Opening Balance' in str(particulars_cell.value))
            
            # Transaction block start requires: Date + Dr/Cr + Vch Type + Vch No. + Debit/Credit + NOT Opening Balance
            if (has_real_date and has_dr_cr and has_vch_type and has_vch_no and 
                (has_debit or has_credit) and is_not_opening_balance):
                # Found the start of the transaction block
                block_start_row = row_idx
                break
        
        # Convert back to DataFrame row index
        df_block_start = block_start_row - 10
        
        # SECOND: Look FORWARDS from the block start to find where it ends ("Entered By :")
        current_row = block_start_row
        while current_row <= ws.max_row:
            # Convert Excel row number to DataFrame row index
            df_row_index = current_row - 10
            if df_row_index >= 0:
                block_rows.append(df_row_index)
            
            # Check if this row contains "Entered By :" in the Particulars column (Column B)
            particulars_cell = ws.cell(row=current_row, column=2)
            if particulars_cell.value and str(particulars_cell.value).strip() == 'Entered By :':
                # Found "Entered By :", this is the end of the block
                break
            
            # Check if this row starts a NEW transaction block (Date + Dr/Cr + Vch Type + Vch No. + Debit/Credit)
            date_cell = ws.cell(row=current_row, column=1)  # Column A (Date)
            particulars_cell = ws.cell(row=current_row, column=2)  # Column B (Particulars)
            vch_type_cell = ws.cell(row=current_row, column=6)  # Column F (Vch Type)
            vch_no_cell = ws.cell(row=current_row, column=7)  # Column G (Vch No.)
            debit_cell = ws.cell(row=current_row, column=8)  # Column H (Debit)
            credit_cell = ws.cell(row=current_row, column=9)  # Column I (Credit)
            
            # Only treat as new block if it has a REAL date (not 'None' or empty) AND Dr/Cr AND Vch Type + Vch No. + Debit/Credit
            has_real_date = (date_cell.value and 
                           str(date_cell.value).strip() and 
                           str(date_cell.value).strip() != 'None' and
                           str(date_cell.value).strip() != '')
            has_dr_cr = particulars_cell.value and str(particulars_cell.value).strip() in ['Dr', 'Cr']
            has_vch_type = vch_type_cell.value and vch_type_cell.font and vch_type_cell.font.bold
            has_vch_no = vch_no_cell.value and vch_no_cell.font and not vch_no_cell.font.bold and not vch_no_cell.font.italic
            has_debit = debit_cell.value and debit_cell.font and debit_cell.font.bold
            has_credit = credit_cell.value and credit_cell.font and credit_cell.font.bold
            is_not_opening_balance = not (particulars_cell.value and 'Opening Balance' in str(particulars_cell.value))
            
            # If we find a new transaction block start, stop here
            if (has_real_date and has_dr_cr and has_vch_type and has_vch_no and 
                (has_debit or has_credit) and is_not_opening_balance and current_row > block_start_row):
                # This row starts a new transaction block, so don't include it
                # The current block ends at the previous row
                break
            
            current_row += 1
        
        # Don't close workbook if it's cached - keep it open for performance
        if file_path not in self._cached_workbooks:
            wb.close()
        
        return block_rows
    
    def identify_transaction_blocks(self, transactions_df, file_path):
        """
        Identify all transaction blocks in the transactions DataFrame.
        
        Args:
            transactions_df: DataFrame containing transaction data
            file_path: Path to the Excel file to analyze formatting
        
        Returns:
            List of transaction block row indices
        """
        # Load workbook with openpyxl to access formatting
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        transaction_blocks = []
        current_block = []
        in_block = False
        
        for row_idx in range(10, ws.max_row + 1):  # Start from row 10 (after headers)
            # Convert Excel row to DataFrame row index
            df_row_idx = row_idx - 10
            
            if df_row_idx < 0:
                continue
            
            # Check if this row starts a new transaction block
            date_cell = ws.cell(row=row_idx, column=1)  # Column A (Date)
            particulars_cell = ws.cell(row=row_idx, column=2)  # Column B (Particulars)
            vch_type_cell = ws.cell(row=row_idx, column=6)  # Column F (Vch Type)
            vch_no_cell = ws.cell(row=row_idx, column=7)  # Column G (Vch No.)
            debit_cell = ws.cell(row=row_idx, column=8)  # Column H (Debit)
            credit_cell = ws.cell(row=row_idx, column=9)  # Column I (Credit)
            
            # Check if this row has a real date and Dr/Cr
            has_real_date = (date_cell.value and 
                           str(date_cell.value).strip() and 
                           str(date_cell.value).strip() != 'None' and
                           str(date_cell.value).strip() != '')
            has_dr_cr = particulars_cell.value and str(particulars_cell.value).strip() in ['Dr', 'Cr']
            
            # Check if this row has Vch Type (Bold) and Vch No. (Regular) - required for transaction blocks
            has_vch_type = vch_type_cell.value and vch_type_cell.font and vch_type_cell.font.bold
            has_vch_no = vch_no_cell.value and vch_no_cell.font and not vch_no_cell.font.bold and not vch_no_cell.font.italic
            
            # Check if this row has Debit or Credit amount (Bold) - required for transaction blocks
            has_debit = debit_cell.value and debit_cell.font and debit_cell.font.bold
            has_credit = credit_cell.value and credit_cell.font and credit_cell.font.bold
            
            # Check if this is NOT an Opening Balance row (which is not a transaction block)
            is_not_opening_balance = not (particulars_cell.value and 'Opening Balance' in str(particulars_cell.value))
            
            # Check if this row ends a transaction block ("Entered By :")
            is_block_end = particulars_cell.value and str(particulars_cell.value).strip() == 'Entered By :'
            
            # Transaction block start requires: Date + Dr/Cr + Vch Type + Vch No. + Debit/Credit + NOT Opening Balance
            is_block_start = (has_real_date and has_dr_cr and has_vch_type and has_vch_no and 
                             (has_debit or has_credit) and is_not_opening_balance)
            
            if is_block_start:
                # If we're already in a block, end the current one
                if in_block and current_block:
                    transaction_blocks.append(current_block)
                
                # Start new block
                current_block = [df_row_idx]
                in_block = True
            elif in_block:
                # Continue adding rows to current block
                current_block.append(df_row_idx)
                
                # Check if this row ends the block
                if is_block_end:
                    # End the current block
                    transaction_blocks.append(current_block)
                    current_block = []
                    in_block = False
        
        # Add the last block if it exists
        if in_block and current_block:
            transaction_blocks.append(current_block)
        
        wb.close()
        
        return transaction_blocks
    
    def find_transaction_block_header(self, description_row_idx, transactions_df, file_path=None):
        """
        Find the transaction block header row for a given description row.
        This is the UNIVERSAL method used by all matching modules.
        OPTIMIZED: Uses caching to avoid redundant scans.
        
        Args:
            description_row_idx: The row index of a description/narration row
            transactions_df: DataFrame containing transaction data
            file_path: Optional file path for caching (recommended for performance)
        
        Returns:
            Row index of the transaction block header (with date, particulars, and amounts)
        """
        # OPTIMIZATION: Check cache first if file_path provided
        if file_path and file_path in self._block_header_cache:
            if description_row_idx in self._block_header_cache[file_path]:
                return self._block_header_cache[file_path][description_row_idx]
        
        # Start from the description row and go backwards to find the block header
        # Block header is the row with date and particulars (Dr/Cr) and amounts
        for row_idx in range(description_row_idx, -1, -1):
            row = transactions_df.iloc[row_idx]
            
            # Check if this row has a date and particulars
            has_date = pd.notna(row.iloc[0]) and str(row.iloc[0]).strip() != ''
            has_debit = pd.notna(row.iloc[7]) and row.iloc[7] != 0
            has_credit = pd.notna(row.iloc[8]) and row.iloc[8] != 0
            
            # Transaction block header: has date, particulars, and either debit or credit
            if has_date and (has_debit or has_credit):
                block_header_idx = row_idx
                # OPTIMIZATION: Cache the result
                if file_path:
                    if file_path not in self._block_header_cache:
                        self._block_header_cache[file_path] = {}
                    self._block_header_cache[file_path][description_row_idx] = block_header_idx
                return block_header_idx
        
        # If no header found, return the description row itself
        block_header_idx = description_row_idx
        # OPTIMIZATION: Cache the result
        if file_path:
            if file_path not in self._block_header_cache:
                self._block_header_cache[file_path] = {}
            self._block_header_cache[file_path][description_row_idx] = block_header_idx
        return block_header_idx
    
    def build_narration_index(self, transactions_df, file_path):
        """
        Build an index of narrations for fast lookups.
        Index structure: {narration_text: [(block_header_idx, description_idx, amounts_dict), ...]}
        
        Args:
            transactions_df: DataFrame containing transaction data
            file_path: Path to Excel file for amount extraction
        
        Returns:
            Dictionary mapping narration text to list of transaction info tuples
        """
        # Check cache first
        if file_path in self._narration_index_cache:
            return self._narration_index_cache[file_path]
        
        narration_index = {}
        
        print(f"  - Building narration index for {len(transactions_df)} transactions...")
        
        for idx in range(len(transactions_df)):
            # Find block header
            block_header_idx = self.find_transaction_block_header(idx, transactions_df, file_path)
            
            # Find description row
            description_idx = self.find_description_row_in_block(idx, transactions_df)
            if description_idx is None:
                continue
            
            # Extract narration
            narration = str(transactions_df.iloc[description_idx, 2]).strip()
            
            # Skip empty or very short narrations
            if len(narration) < 10 or narration.lower() in ['nan', 'none', '']:
                continue
            
            # Get amounts
            amounts = self.get_header_row_amounts(block_header_idx, file_path)
            if not amounts or amounts.get('amount', 0) <= 0:
                continue
            
            # Add to index
            if narration not in narration_index:
                narration_index[narration] = []
            
            narration_index[narration].append((block_header_idx, description_idx, amounts))
        
        # Cache the index
        self._narration_index_cache[file_path] = narration_index
        
        print(f"  - Built index with {len(narration_index)} unique narrations")
        
        return narration_index
    
    def find_description_row_in_block(self, row_idx, transactions_df):
        """
        Find the description row (with narration) within a transaction block.
        This is the UNIVERSAL method used by all matching modules.
        
        Args:
            row_idx: The row index to start searching from
            transactions_df: DataFrame containing transaction data
        
        Returns:
            Row index of the description row with narration text, or None if not found
        """
        # Start from the current row and look for a description row
        # Description row has narration text in column 2 (iloc[2])
        
        # Look backwards first to find description row
        for i in range(row_idx, -1, -1):
            row = transactions_df.iloc[i]
            narration = str(row.iloc[2]).strip()
            
            # Check if this row has meaningful narration text
            if (len(narration) > 10 and 
                narration.lower() not in ['nan', 'none', ''] and
                not narration.startswith('Dr') and 
                not narration.startswith('Cr')):
                return i
        
        # If not found looking backwards, look forwards
        for i in range(row_idx + 1, len(transactions_df)):
            row = transactions_df.iloc[i]
            narration = str(row.iloc[2]).strip()
            
            # Check if this row has meaningful narration text
            if (len(narration) > 10 and 
                narration.lower() not in ['nan', 'none', ''] and
                not narration.startswith('Dr') and 
                not narration.startswith('Cr')):
                return i
        
        return None
    
    def find_parent_transaction_row(self, current_row, transactions_df):
        """
        Find the parent transaction row for a description row.
        This is the UNIVERSAL method used by all matching modules.
        
        Args:
            current_row: The row index to start searching from
            transactions_df: DataFrame containing transaction data
        
        Returns:
            Row index of the parent transaction row (with date and amounts), or None if not found
        """
        # Look backwards from current row to find the most recent transaction row
        for row_idx in range(current_row - 1, -1, -1):
            row = transactions_df.iloc[row_idx]
            has_date = pd.notna(row.iloc[0])  # Date column
            has_debit = pd.notna(row.iloc[7]) and float(row.iloc[7]) > 0  # Debit column
            has_credit = pd.notna(row.iloc[8]) and float(row.iloc[8]) > 0  # Credit column
            
            if has_date and (has_debit or has_credit):
                return row_idx
        
        # If no parent found looking backwards, look forwards
        for row_idx in range(current_row + 1, len(transactions_df)):
            row = transactions_df.iloc[row_idx]
            has_date = pd.notna(row.iloc[0])  # Date column
            has_debit = pd.notna(row.iloc[7]) and float(row.iloc[7]) > 0  # Debit column
            has_credit = pd.notna(row.iloc[8]) and float(row.iloc[8]) > 0  # Credit column
            
            if has_date and (has_debit or has_credit):
                return row_idx
        
        return None
    
    def get_block_header_amounts(self, block_rows, file_path):
        """
        Get debit and credit amounts from the header row of a transaction block.
        This is a UNIVERSAL method used by all matching modules.
        OPTIMIZED: Uses caching to avoid reloading Excel files.
        
        The header row is guaranteed to be the first element in block_rows
        (as identified by identify_transaction_blocks).
        
        Args:
            block_rows: List of row indices in the transaction block (header row is first)
            file_path: Path to the Excel file to read amounts from
        
        Returns:
            Dictionary with 'debit', 'credit', 'row', 'is_lender', 'is_borrower', and 'amount' keys.
            Returns empty dict if not found.
            
            Universal Lender/Borrower Definition:
            - Lender: Column H has amount (debit > 0)
            - Borrower: Column I has amount (credit > 0)
        """
        if not block_rows:
            return {}
        
        # The first row in block_rows is the header row (guaranteed by identify_transaction_blocks)
        header_row_idx = block_rows[0]
        
        # OPTIMIZATION: Check cache first
        if file_path in self._amount_cache:
            if header_row_idx in self._amount_cache[file_path]:
                return self._amount_cache[file_path][header_row_idx]
        
        excel_header_row = header_row_idx + 10  # Convert DataFrame row to Excel row
        
        try:
            # OPTIMIZATION: Use cached workbook if available, otherwise load
            if file_path in self._cached_workbooks:
                wb, ws = self._cached_workbooks[file_path]
            else:
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                self._cached_workbooks[file_path] = (wb, ws)
            
            if excel_header_row <= ws.max_row:
                debit_cell = ws.cell(row=excel_header_row, column=8)  # Column H
                credit_cell = ws.cell(row=excel_header_row, column=9)  # Column I
                
                if (debit_cell.value is not None and debit_cell.value != 0) or \
                   (credit_cell.value is not None and credit_cell.value != 0):
                    # Extract debit and credit values
                    debit_value = debit_cell.value if debit_cell.value else None
                    credit_value = credit_cell.value if credit_cell.value else None
                    
                    # Convert to float for comparison (handle string numbers)
                    try:
                        debit_float = float(debit_value) if debit_value is not None else 0.0
                    except (ValueError, TypeError):
                        debit_float = 0.0
                    
                    try:
                        credit_float = float(credit_value) if credit_value is not None else 0.0
                    except (ValueError, TypeError):
                        credit_float = 0.0
                    
                    # Universal Lender/Borrower Identification
                    # Lender = Has Debit (Column H has amount > 0)
                    # Borrower = Has Credit (Column I has amount > 0)
                    is_lender = debit_float > 0
                    is_borrower = credit_float > 0
                    
                    # Determine the transaction amount (the non-zero value)
                    amount = debit_float if is_lender else credit_float
                    
                    amounts = {
                        'debit': debit_value,
                        'credit': credit_value,
                        'row': header_row_idx,
                        'is_lender': is_lender,
                        'is_borrower': is_borrower,
                        'amount': amount
                    }
                    
                    # OPTIMIZATION: Cache the result
                    if file_path not in self._amount_cache:
                        self._amount_cache[file_path] = {}
                    self._amount_cache[file_path][header_row_idx] = amounts
                    
                    # Don't close workbook if it's cached
                    if file_path not in self._cached_workbooks:
                        wb.close()
                    
                    return amounts
            
            # Don't close workbook if it's cached
            if file_path not in self._cached_workbooks:
                wb.close()
        except Exception as e:
            print(f"Error reading amounts from header row: {e}")
        
        return {}
    
    def get_header_row_amounts(self, header_row_idx, file_path):
        """
        Get debit and credit amounts from a specific header row.
        This is a UNIVERSAL method used by all matching modules.
        
        This is a convenience wrapper around get_block_header_amounts() for cases
        where you only have a single header row index (not a full block).
        
        Args:
            header_row_idx: DataFrame row index of the header row
            file_path: Path to the Excel file to read amounts from
        
        Returns:
            Dictionary with 'debit', 'credit', 'row', 'is_lender', 'is_borrower', and 'amount' keys.
            Returns empty dict if not found.
            
            Universal Lender/Borrower Definition:
            - Lender: Column H has amount (debit > 0)
            - Borrower: Column I has amount (credit > 0)
        """
        # Convert single row index to block_rows format (list with one element)
        return self.get_block_header_amounts([header_row_idx], file_path)
    
    def filter_blocks_by_matching_amounts(self, blocks1, blocks2, file1_path, file2_path):
        """
        Universal method to filter blocks by matching amounts and opposite transaction types.
        This is used by all matching modules to narrow down the scope before expensive operations.
        
        Args:
            blocks1: List of block row indices from File 1 (from identify_transaction_blocks)
            blocks2: List of block row indices from File 2 (from identify_transaction_blocks)
            file1_path: Path to File 1 Excel file
            file2_path: Path to File 2 Excel file
        
        Returns:
            List of tuples: [(block1_rows, block2_rows, amount, file1_is_lender), ...]
            Only returns pairs with:
            - Matching amounts (exact match, no tolerance)
            - Opposite transaction types (one lender, one borrower)
        """
        matching_pairs = []
        
        # Extract amounts for all blocks in File 1
        file1_block_amounts = {}
        for block_rows in blocks1:
            amounts = self.get_block_header_amounts(block_rows, file1_path)
            if amounts and amounts.get('amount', 0) > 0:
                file1_block_amounts[tuple(block_rows)] = {
                    'amounts': amounts,
                    'amount': amounts.get('amount', 0),
                    'is_lender': amounts.get('is_lender', False),
                    'is_borrower': amounts.get('is_borrower', False)
                }
        
        # Extract amounts for all blocks in File 2
        file2_block_amounts = {}
        for block_rows in blocks2:
            amounts = self.get_block_header_amounts(block_rows, file2_path)
            if amounts and amounts.get('amount', 0) > 0:
                file2_block_amounts[tuple(block_rows)] = {
                    'amounts': amounts,
                    'amount': amounts.get('amount', 0),
                    'is_lender': amounts.get('is_lender', False),
                    'is_borrower': amounts.get('is_borrower', False)
                }
        
        # Find matching pairs: same amount + opposite types
        for block1_key, block1_data in file1_block_amounts.items():
            amount1 = block1_data['amount']
            file1_is_lender = block1_data['is_lender']
            file1_is_borrower = block1_data['is_borrower']
            
            for block2_key, block2_data in file2_block_amounts.items():
                amount2 = block2_data['amount']
                file2_is_lender = block2_data['is_lender']
                file2_is_borrower = block2_data['is_borrower']
                
                # Check: amounts match exactly AND opposite transaction types
                if (amount1 == amount2 and 
                    ((file1_is_lender and file2_is_borrower) or 
                     (file1_is_borrower and file2_is_lender))):
                    
                    matching_pairs.append((
                        list(block1_key),  # block1_rows
                        list(block2_key),  # block2_rows
                        amount1,           # amount (same for both)
                        file1_is_lender    # file1_is_lender (for convenience)
                    ))
        
        return matching_pairs
    
    def group_blocks_by_amount(self, blocks, file_path):
        """
        Universal method to group blocks by their transaction amounts.
        This creates an efficient lookup structure for amount-based filtering.
        
        Args:
            blocks: List of block row indices (from identify_transaction_blocks)
            file_path: Path to Excel file
        
        Returns:
            Dictionary: {
                amount: {
                    'lender_blocks': [list of block_rows],
                    'borrower_blocks': [list of block_rows]
                }
            }
        """
        grouped = {}
        
        for block_rows in blocks:
            amounts = self.get_block_header_amounts(block_rows, file_path)
            if amounts and amounts.get('amount', 0) > 0:
                amount = amounts.get('amount', 0)
                is_lender = amounts.get('is_lender', False)
                is_borrower = amounts.get('is_borrower', False)
                
                if amount not in grouped:
                    grouped[amount] = {
                        'lender_blocks': [],
                        'borrower_blocks': []
                    }
                
                if is_lender:
                    grouped[amount]['lender_blocks'].append(block_rows)
                elif is_borrower:
                    grouped[amount]['borrower_blocks'].append(block_rows)
        
        return grouped
    
    def get_matching_pairs_from_grouped(self, grouped1, grouped2):
        """
        Get matching block pairs from pre-grouped amount dictionaries.
        More efficient than filtering all combinations.
        
        Args:
            grouped1: Dictionary from group_blocks_by_amount for File 1
            grouped2: Dictionary from group_blocks_by_amount for File 2
        
        Returns:
            List of tuples: [(block1_rows, block2_rows, amount, file1_is_lender), ...]
        """
        matching_pairs = []
        
        # Only check amounts that exist in both files
        common_amounts = set(grouped1.keys()) & set(grouped2.keys())
        
        for amount in common_amounts:
            # File 1 lenders match with File 2 borrowers
            for block1 in grouped1[amount]['lender_blocks']:
                for block2 in grouped2[amount]['borrower_blocks']:
                    matching_pairs.append((block1, block2, amount, True))
            
            # File 1 borrowers match with File 2 lenders
            for block1 in grouped1[amount]['borrower_blocks']:
                for block2 in grouped2[amount]['lender_blocks']:
                    matching_pairs.append((block1, block2, amount, False))
        
        return matching_pairs